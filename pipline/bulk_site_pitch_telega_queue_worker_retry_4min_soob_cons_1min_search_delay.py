"""
bulk_site_pitch_telega.py

Что делает:
1. Читает Excel-файл ip_fio_links.xlsx или путь из --xlsx.
2. Берёт последние два столбца каждой строки: телефон и сайт.
3. Пропускает строку, если телефон/сайт пустые или телефон не строго в формате +79999999999.
4. Загружает текст сайта.
5. Отправляет в OpenRouter ровно тот же промт, что в site_pitch.py, но с сайтом из Excel.
6. Находит получателя в Telegram по номеру телефона так же, как в telega.py.
7. Показывает сообщение и отправляет его только после подтверждения.

Интерактивный режим включён по умолчанию.
Чтобы убрать вопросы и выполнять всё автоматически: добавь флаг --yes.
Для безопасной проверки без отправки: добавь флаг --dry-run.

Перед запуском создай .env рядом со скриптом:
OPENROUTER_API_KEY=...
TELEGRAM_API_ID=...
TELEGRAM_API_HASH=...
TELEGRAM_PHONE=+79999999999

Установка зависимостей:
pip install openpyxl requests beautifulsoup4 python-dotenv telethon

Примеры запуска:
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --dry-run
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --yes
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --yes --limit 3
"""

import asyncio
import argparse
import asyncio
import json
import os
import random
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from telethon import TelegramClient, functions, types
from telethon.errors import (
    FloodWaitError,
    PeerFloodError,
    PhoneNumberInvalidError,
    SessionPasswordNeededError,
)


load_dotenv()

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
MODEL = "google/gemini-3-flash-preview"
SESSION_NAME = os.getenv("TELEGRAM_SESSION_NAME", "my_telegram_session")
PHONE_RE = re.compile(r"^\+7\d{10}$")
CRITICAL_TELEGRAM_ERRORS = (FloodWaitError, PeerFloodError)
TELEGRAM_RETRY_PAUSE_SECONDS = 4 * 60
TELEGRAM_PHONE_SEARCH_DELAY_SECONDS = 60
SOOB_CONS_SCRIPT_NAME = "start_clicker.sh"


class TelegramRepeatedError(RuntimeError):
    """Telegram снова вернул ошибку сразу после 4-минутной паузы."""


@dataclass
class LeadRow:
    excel_row: int
    phone: str
    site: str


class StepController:
    def __init__(self, auto_yes: bool) -> None:
        self.auto_yes = auto_yes

    def ask_continue(self, title: str) -> bool:
        if self.auto_yes:
            print(f"[AUTO] {title}: продолжаю без вопроса, потому что включён --yes.")
            return True

        print("\n" + "=" * 90)
        print(title)
        print("Введите:")
        print("  да / y / продолжаем  — продолжить")
        print("  нет / n / skip       — пропустить эту строку")
        print("  стоп / q / exit      — завершить программу")

        while True:
            answer = input("Продолжаем? > ").strip().lower()
            if answer in {"да", "д", "y", "yes", "продолжаем", "go", "ok", "ок"}:
                return True
            if answer in {"нет", "н", "n", "no", "skip", "пропустить"}:
                print("[USER] Строка пропущена по вашему решению.")
                return False
            if answer in {"стоп", "stop", "q", "quit", "exit", "выход"}:
                print("[USER] Остановка программы по вашему решению.")
                raise KeyboardInterrupt
            print("Не понял ответ. Напишите 'да', 'нет' или 'стоп'.")

    def ask_send(self) -> bool:
        if self.auto_yes:
            print("[AUTO] Отправляю без финального вопроса, потому что включён --yes.")
            return True

        print("\n" + "=" * 90)
        print("ФИНАЛЬНОЕ ПОДТВЕРЖДЕНИЕ ОТПРАВКИ")
        print("Чтобы реально отправить сообщение, введите строго: отправляем")
        print("Любой другой ответ НЕ отправит сообщение и программа перейдёт к следующей строке.")
        answer = input("Отправляем? > ").strip().lower()
        if answer == "отправляем":
            return True
        print("[USER] Сообщение не отправлено, строка пропущена.")
        return False


def log_block(title: str, text: str = "") -> None:
    print("\n" + "#" * 90)
    print(f"# {title}")
    print("#" * 90)
    if text:
        print(text)


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_phone(phone: str) -> Optional[str]:
    phone = normalize_cell(phone)
    phone = re.sub(r"[\s\-()]+", "", phone)
    if PHONE_RE.fullmatch(phone):
        return phone
    return None


def normalize_url(raw_url: str) -> str:
    url = normalize_cell(raw_url)
    if not url:
        return ""
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        url = "https://" + url
    return url


def read_leads_from_excel(xlsx_path: str, sheet_name: Optional[str] = None) -> list[LeadRow]:
    log_block("ЧТЕНИЕ EXCEL", f"Файл: {xlsx_path}")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel-файл не найден: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    print(f"Активный лист: {worksheet.title}")
    print(f"Всего строк: {worksheet.max_row}")
    print(f"Всего столбцов: {worksheet.max_column}")

    if worksheet.max_column < 2:
        raise RuntimeError("В Excel должно быть минимум два столбца, потому что нужны последние два: телефон и сайт.")

    phone_col = worksheet.max_column - 1
    site_col = worksheet.max_column

    print(f"Телефон будет взят из предпоследнего столбца: #{phone_col}")
    print(f"Сайт будет взят из последнего столбца: #{site_col}")

    leads: list[LeadRow] = []

    for row_idx in range(2, worksheet.max_row + 1):
        raw_phone = normalize_cell(worksheet.cell(row=row_idx, column=phone_col).value)
        raw_site = normalize_cell(worksheet.cell(row=row_idx, column=site_col).value)

        print("\n" + "-" * 90)
        print(f"Строка Excel #{row_idx}")
        print(f"Сырой телефон: {raw_phone!r}")
        print(f"Сырой сайт:    {raw_site!r}")

        if not raw_phone and not raw_site:
            print("[SKIP] Пропуск: оба поля пустые — нет ни телефона, ни сайта.")
            continue
        if not raw_phone:
            print("[SKIP] Пропуск: пустой телефон.")
            continue
        if not raw_site:
            print("[SKIP] Пропуск: пустой сайт.")
            continue

        phone = validate_phone(raw_phone)
        if not phone:
            print("[SKIP] Пропуск: телефон не в формате +79999999999.")
            print("       Требуется строго: плюс, цифра 7 и ещё 10 цифр. Пример: +79161234567")
            continue

        site = normalize_url(raw_site)
        parsed = urlparse(site)
        if not parsed.netloc:
            print("[SKIP] Пропуск: сайт выглядит некорректно после нормализации.")
            print(f"       Нормализованный сайт: {site!r}")
            continue

        print("[OK] Строка прошла первичную проверку.")
        print(f"     Телефон: {phone}")
        print(f"     Сайт:    {site}")
        leads.append(LeadRow(excel_row=row_idx, phone=phone, site=site))

    print("\n" + "=" * 90)
    print(f"Готово. Валидных строк: {len(leads)}")
    workbook.close()
    return leads


def fetch_site_text(url: str, max_chars: int = 8000) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; SiteAnalyzer/1.0)",
    }

    print(f"[SITE] Загружаю сайт: {url}")
    response = requests.get(url, headers=headers, timeout=15)
    response.raise_for_status()
    print(f"[SITE] HTTP статус: {response.status_code}")
    print(f"[SITE] Content-Type: {response.headers.get('Content-Type', 'не указан')}")

    soup = BeautifulSoup(response.text, "html.parser")

    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title and soup.title.string else ""

    meta_description = ""
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        meta_description = meta["content"].strip()

    text = soup.get_text(separator=" ")
    text = re.sub(r"\s+", " ", text).strip()

    result = f"""
URL: {url}
Title: {title}
Meta description: {meta_description}
Page text: {text}
""".strip()

    print(f"[SITE] Title: {title!r}")
    print(f"[SITE] Meta description: {meta_description!r}")
    print(f"[SITE] Длина очищенного текста до обрезки: {len(result)} символов")
    print(f"[SITE] В нейросеть уйдёт максимум: {max_chars} символов")

    return result[:max_chars]


def generate_pitch(site_text: str = "") -> str:
    prompt = f"""
Ты пишешь короткое рекламное сообщение потенциальному клиенту от веб-разработчика.

Ситуация:

* Мы не знаем точно, есть ли у клиента сайт.
* Сообщение должно одинаково хорошо подходить и тем, у кого сайта пока нет, и тем, у кого сайт уже есть.
* Не пиши фразы вроде: "я нашёл ваш сайт", "увидел ваш сайт", "наткнулся на ваш сайт", "у вас нет сайта", "ваш сайт плохой".
* Не утверждай то, чего мы точно не знаем.
* Можно использовать нейтральную формулировку: "если сайта пока нет — можно сделать его с нуля, если сайт уже есть — можно обновить его и доработать".

Стиль:

* Пиши просто, по-человечески.
* Без пафоса.
* Без слов: "статусный", "экспертиза", "упаковать", "презентабельный", "визуальная форма", "достойный", "серьезные клиенты".
* Не используй канцелярит.
* Не пиши слишком вежливо и корпоративно.
* Тон: прямой, уверенный, спокойный.
* Максимум 6-7 предложений.
* Ни слова не говори про цену.
* Обязательно поздоровайся.
* Прощаться не надо.

Что нужно сказать:

1. Меня зовут Владислав.
2. Я уже несколько лет разрабатываю сайты.
3. Также занимаюсь SEO-анализом — то есть слежу за тем, чтобы сайт могли увидеть как можно больше целевых клиентов.
4. Скажи, что я могу сделать сайт с нуля или обновить уже существующий.
5. Сайт должен быть красивым, понятным, современным и удобным для людей.
6. Хороший сайт помогает бизнесу выглядеть понятнее, вызывает больше доверия и помогает получать больше заявок.
7. Напиши уверенно, но честно: "сайт будет сделан так, чтобы он чаще попадал на первые страницы поиска".
8. Объясни, что чем чаще сайт появляется на первых страницах поиска, тем больше людей его видят, а значит у бизнеса может быть больше клиентов.
9. Упомяни, что я разрабатывал сайты для учебных заведений, адвокатских контор и интернет-магазинов.
10. Напиши, что если нужно, могу прислать свои работы.
11. Напиши, что работаю по договору.
12. Не дави на человека и не пугай его потерей клиентов.

Данные о бизнесе клиента, если они есть, чтобы аккуратно подстроить сообщение под сферу:
{site_text}

Выдай только готовое сообщение клиенту, без пояснений.
""".strip()

    if not OPENROUTER_API_KEY:
        raise RuntimeError(
            "Не найден OPENROUTER_API_KEY. Добавь его в .env или экспортируй переменную окружения."
        )

    print("[AI] Отправляю запрос в OpenRouter.")
    print(f"[AI] Модель: {MODEL}")
    print(f"[AI] Длина промта: {len(prompt)} символов")

    response = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://example.com",
            "X-Title": "Site Pitch Generator",
        },
        json={
            "model": MODEL,
            "messages": [
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
            "temperature": 0.7,
            "max_tokens": 300,
        },
        timeout=60,
    )

    print(f"[AI] HTTP статус OpenRouter: {response.status_code}")

    if response.status_code != 200:
        raise RuntimeError(f"OpenRouter error {response.status_code}: {response.text}")

    data = response.json()
    message = data["choices"][0]["message"]["content"].strip()
    print(f"[AI] Ответ получен. Длина сообщения: {len(message)} символов")
    return message


async def get_or_create_telegram_client() -> TelegramClient:
    api_id_raw = os.getenv("TELEGRAM_API_ID")
    api_hash = os.getenv("TELEGRAM_API_HASH")
    my_phone = os.getenv("TELEGRAM_PHONE")

    if not api_id_raw or not api_hash or not my_phone:
        raise RuntimeError(
            "Не найдены TELEGRAM_API_ID, TELEGRAM_API_HASH или TELEGRAM_PHONE. "
            "Добавь их в .env. API ID и HASH берутся на https://my.telegram.org"
        )

    try:
        api_id = int(api_id_raw)
    except ValueError as exc:
        raise RuntimeError("TELEGRAM_API_ID должен быть числом.") from exc

    client = TelegramClient(SESSION_NAME, api_id, api_hash)
    print(f"[TG] Подключаюсь к Telegram. Session name: {SESSION_NAME}")
    await client.connect()

    if not await client.is_user_authorized():
        print("[TG] Сессия ещё не авторизована. Отправляю код входа.")
        await client.send_code_request(my_phone)
        code = input("Введите код из Telegram/SMS: ").strip()

        try:
            await client.sign_in(my_phone, code)
        except SessionPasswordNeededError:
            password = getpass("Включена 2FA. Введите пароль Telegram: ")
            await client.sign_in(password=password)

    print("[TG] Telegram подключён и авторизован.")
    return client


async def wait_before_telegram_phone_search(args) -> None:
    """Делает фиксированную паузу 1 минуту между поисками Telegram-аккаунтов по номеру телефона."""
    delay = TELEGRAM_PHONE_SEARCH_DELAY_SECONDS

    last_search_at = getattr(args, "_last_telegram_phone_search_at", None)
    now = time.monotonic()

    if last_search_at is not None:
        elapsed = now - last_search_at
        remaining = delay - elapsed
        if remaining > 0:
            print(f"[DELAY] Жду {remaining:.1f} секунд перед следующим поиском Telegram-аккаунта по номеру...")
            await asyncio.sleep(remaining)

    args._last_telegram_phone_search_at = time.monotonic()


async def find_telegram_user_by_phone(client: TelegramClient, target_phone: str):
    print(f"[TG] Ищу пользователя по номеру: {target_phone}")
    print("[TG] Для поиска временно импортирую один контакт, как в telega.py.")

    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=target_phone,
        first_name="Temporary",
        last_name="Contact",
    )

    result = await client(functions.contacts.ImportContactsRequest([contact]))

    if not result.users:
        print("[TG] Пользователь не найден или недоступен по настройкам приватности.")
        return None

    user = result.users[0]
    print("[TG] Пользователь найден.")
    print(f"     id: {getattr(user, 'id', None)}")
    print(f"     username: {getattr(user, 'username', None)}")
    print(f"     first_name: {getattr(user, 'first_name', None)}")
    print(f"     last_name: {getattr(user, 'last_name', None)}")
    return user


async def cleanup_telegram_contact(client: TelegramClient, user) -> None:
    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
        print("[TG] Временный контакт удалён из контактов Telegram.")
    except Exception as exc:
        print("[TG][WARN] Не удалось удалить временный контакт.")
        print(f"          Тип ошибки: {type(exc).__name__}")
        print(f"          Текст ошибки: {exc}")


async def send_telegram_message(client: TelegramClient, user, message: str) -> None:
    print("[TG] Отправляю сообщение...")
    await client.send_message(user, message)
    print("[TG] Сообщение отправлено.")


def print_telegram_error(exc: Exception, attempt: int) -> None:
    print("[TG][ERROR] Telegram вернул ошибку.")
    print(f"           Попытка: {attempt}")
    print(f"           Тип ошибки: {type(exc).__name__}")
    print(f"           Текст ошибки: {exc}")
    if isinstance(exc, FloodWaitError):
        print(f"           Telegram просит подождать секунд: {exc.seconds}")


def run_soob_cons_script() -> None:
    """Запускает bash-скрипт с логикой вывода сообщений в консоль."""
    script_path = Path(__file__).resolve().with_name(SOOB_CONS_SCRIPT_NAME)

    if not script_path.exists():
        print(f"[TG][WARN] Не найден файл {SOOB_CONS_SCRIPT_NAME} рядом с Python-скриптом.")
        print(f"          Ожидался путь: {script_path}")
        print("          Пауза и повторная попытка всё равно будут выполнены.")
        return

    print(f"[TG][SCRIPT] Запускаю bash-скрипт: {script_path}")

    try:
        completed = subprocess.run(["bash", str(script_path)], check=False)
    except Exception as script_exc:
        print(f"[TG][WARN] Не удалось запустить {SOOB_CONS_SCRIPT_NAME}.")
        print(f"          Тип ошибки: {type(script_exc).__name__}")
        print(f"          Текст ошибки: {script_exc}")
        print("          Пауза и повторная попытка всё равно будут выполнены.")
        return

    if completed.returncode != 0:
        print(f"[TG][WARN] {SOOB_CONS_SCRIPT_NAME} завершился с кодом {completed.returncode}.")
        print("          Пауза и повторная попытка всё равно будут выполнены.")


async def wait_after_first_telegram_error(exc: Exception) -> None:
    print_telegram_error(exc, attempt=1)

    run_soob_cons_script()

    print(f"[TG][PAUSE] Жду {TELEGRAM_RETRY_PAUSE_SECONDS} секунд, потом сделаю ровно одну повторную попытку.")
    await asyncio.sleep(TELEGRAM_RETRY_PAUSE_SECONDS)


async def send_pitch_to_telegram_once(
    args,
    controller: StepController,
    client: Optional[TelegramClient],
    lead: LeadRow,
    pitch: str,
) -> Optional[TelegramClient]:
    log_block("ЭТАП 4: TELEGRAM — ПОДКЛЮЧЕНИЕ И ПОИСК ПОЛЬЗОВАТЕЛЯ")
    if client is None:
        client = await get_or_create_telegram_client()
    else:
        print("[TG] Уже подключён к Telegram, использую существующее соединение.")

    await wait_before_telegram_phone_search(args)
    user = await find_telegram_user_by_phone(client, lead.phone)
    if user is None:
        print("[SKIP] Сообщение не будет отправлено, потому что пользователь не найден.")
        return client

    if not controller.ask_continue("Этап 4: Telegram-пользователь найден и показан"):
        await cleanup_telegram_contact(client, user)
        return client

    log_block("ЭТАП 5: ТЕКСТ, КОТОРЫЙ БУДЕТ ОТПРАВЛЕН", pitch)

    if args.dry_run:
        print("[DRY-RUN] Включён --dry-run, поэтому сообщение НЕ отправляется.")
        await cleanup_telegram_contact(client, user)
        return client

    message_was_sent = False

    if controller.ask_send():
        await send_telegram_message(client, user, pitch)
        message_was_sent = True
    else:
        print("[SKIP] Отправка отменена пользователем.")

    await cleanup_telegram_contact(client, user)

    if message_was_sent and args.yes and args.delay > 0:
        print(f"[DELAY] Жду {args.delay} секунд перед следующей отправкой...")
        await asyncio.sleep(args.delay)

    return client


async def send_pitch_to_telegram_with_retry(
    args,
    controller: StepController,
    client: Optional[TelegramClient],
    lead: LeadRow,
    pitch: str,
) -> Optional[TelegramClient]:
    try:
        return await send_pitch_to_telegram_once(args, controller, client, lead, pitch)
    except KeyboardInterrupt:
        raise
    except Exception as exc:
        await wait_after_first_telegram_error(exc)

    try:
        return await send_pitch_to_telegram_once(args, controller, client, lead, pitch)
    except KeyboardInterrupt:
        raise
    except Exception as exc:
        print_telegram_error(exc, attempt=2)
        print("[TG][STOP] Telegram снова вернул ошибку сразу после 4-минутной паузы. Останавливаю программу.")
        raise TelegramRepeatedError("Telegram повторно вернул ошибку после паузы") from exc

def load_processed_keys(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()}


def append_processed_key(path: Path, key: str) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(key + "\n")
        f.flush()


def read_queue_leads(queue_path: Path) -> list[LeadRow]:
    """Читает JSONL-очередь, которую дописывает addres-sayta_pipeline.py."""
    if not queue_path.exists():
        return []

    leads: list[LeadRow] = []
    with queue_path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                print(f"[QUEUE][WARN] Не смог прочитать строку очереди #{line_no}, пропускаю.")
                continue

            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            excel_row = int(data.get("excel_row") or 0)
            if not phone or not site or not excel_row:
                continue
            leads.append(LeadRow(excel_row=excel_row, phone=phone, site=site))
    return leads


def queue_key(lead: LeadRow) -> str:
    return f"{lead.excel_row}|{lead.phone}|{lead.site}"


async def process_one_lead(args, controller: StepController, client: Optional[TelegramClient], lead: LeadRow, index_title: str) -> Optional[TelegramClient]:
    log_block(
        f"ОЧЕРЕДЬ {index_title} | EXCEL #{lead.excel_row}",
        f"Телефон: {lead.phone}\nСайт:    {lead.site}",
    )

    if not controller.ask_continue("Этап 1: данные строки взяты из очереди и показаны"):
        return client

    try:
        log_block("ЭТАП 2: ЗАГРУЗКА САЙТА")

        try:
            site_text = fetch_site_text(lead.site, max_chars=args.max_site_chars)
            print("[OK] Сайт успешно загружен.")
            print(f"[SITE] Длина текста сайта для нейросети: {len(site_text)} символов")

        except requests.RequestException as exc:
            print("[WARNING] Сайт не удалось загрузить, но строка НЕ будет пропущена.")
            print(f"          Строка Excel: {lead.excel_row}")
            print(f"          Сайт: {lead.site}")
            print(f"          Тип ошибки: {type(exc).__name__}")
            print("          В нейросеть НЕ передаю текст ошибки, чтобы она не писала клиенту про недоступность сайта.")

            site_text = f"""
        URL: {lead.site}
        Title:
        Meta description:
        Page text: Информации со страницы мало. Напиши сообщение в общем стиле: сайт выглядит слабым, его можно сделать понятнее, современнее и лучше подготовить под поиск.
        """.strip()

        log_block("ЭТАП 3: ЗАПРОС В НЕЙРОСЕТЬ")
        pitch = generate_pitch(site_text)

        log_block("ОТВЕТ НЕЙРОСЕТИ", pitch)
        if not controller.ask_continue("Этап 3: ответ нейросети получен и показан"):
            return client

        client = await send_pitch_to_telegram_with_retry(args, controller, client, lead, pitch)

    except TelegramRepeatedError:
        raise
    except PhoneNumberInvalidError:
        print("[ERROR] Telegram считает номер телефона неверным.")
        print(f"        Строка Excel: {lead.excel_row}")
        print(f"        Телефон: {lead.phone}")
        print("        Перехожу к следующей строке.")
    except CRITICAL_TELEGRAM_ERRORS as exc:
        print("[CRITICAL] Telegram ограничил отправку сообщений.")
        print(f"           Тип ошибки: {type(exc).__name__}")
        print(f"           Текст ошибки: {exc}")
        if isinstance(exc, FloodWaitError):
            print(f"           Нужно подождать секунд: {exc.seconds}")
        print("           Программа останавливается, чтобы не усиливать ограничение.")
        raise
    except Exception as exc:
        print("[ERROR] Непредвиденная ошибка на строке.")
        print(f"        Строка Excel: {lead.excel_row}")
        print(f"        Телефон: {lead.phone}")
        print(f"        Сайт: {lead.site}")
        print(f"        Тип ошибки: {type(exc).__name__}")
        print(f"        Текст ошибки: {exc}")
        print("        Перехожу к следующей строке.")

    return client


async def process_queue_leads(args) -> None:
    controller = StepController(auto_yes=args.yes)
    queue_path = Path(args.queue)
    state_path = Path(args.processed_state)
    client: Optional[TelegramClient] = None

    print(f"[QUEUE] Слушаю очередь: {queue_path.resolve()}")
    print(f"[QUEUE] Уже обработанные ключи храню тут: {state_path.resolve()}")

    try:
        while True:
            processed_keys = load_processed_keys(state_path)
            leads = read_queue_leads(queue_path)
            new_leads = [lead for lead in leads if queue_key(lead) not in processed_keys]

            if not new_leads:
                if not args.watch:
                    print("[QUEUE] Новых лидов нет. Завершаю работу, потому что --watch не включён.")
                    return
                print(f"[QUEUE] Новых лидов нет. Жду {args.poll} секунд...")
                await asyncio.sleep(args.poll)
                continue

            for idx, lead in enumerate(new_leads, start=1):
                key = queue_key(lead)
                try:
                    client = await process_one_lead(args, controller, client, lead, f"{idx}/{len(new_leads)}")
                except CRITICAL_TELEGRAM_ERRORS + (TelegramRepeatedError,):
                    print("[QUEUE] Остановка из-за повторной Telegram-ошибки. Лид НЕ отмечаю обработанным.")
                    return

                append_processed_key(state_path, key)
                print(f"[QUEUE] Отметил как обработанное: {key}")

    finally:
        if client is not None:
            print("[TG] Отключаюсь от Telegram.")
            await client.disconnect()


async def process_leads(args) -> None:
    controller = StepController(auto_yes=args.yes)
    leads = read_leads_from_excel(args.xlsx, args.sheet)

    if args.start_row:
        leads = [lead for lead in leads if lead.excel_row >= args.start_row]
        print(f"[FILTER] После --start-row {args.start_row} осталось строк: {len(leads)}")

    if args.limit is not None:
        leads = leads[: args.limit]
        print(f"[FILTER] После --limit {args.limit} осталось строк: {len(leads)}")

    if not leads:
        print("Нет валидных строк для обработки. Завершаю работу.")
        return

    client: Optional[TelegramClient] = None

    try:
        for index, lead in enumerate(leads, start=1):
            log_block(
                f"СТРОКА {index}/{len(leads)} | EXCEL #{lead.excel_row}",
                f"Телефон: {lead.phone}\nСайт:    {lead.site}",
            )

            if not controller.ask_continue("Этап 1: данные строки прочитаны и показаны"):
                continue

            try:
                log_block("ЭТАП 2: ЗАГРУЗКА САЙТА")
                
                try:
                    site_text = fetch_site_text(lead.site, max_chars=args.max_site_chars)
                    print("[OK] Сайт успешно загружен.")
                    print(f"[SITE] Длина текста сайта для нейросети: {len(site_text)} символов")
                
                except requests.RequestException as exc:
                    print("[WARNING] Сайт не удалось загрузить, но строка НЕ будет пропущена.")
                    print(f"          Строка Excel: {lead.excel_row}")
                    print(f"          Сайт: {lead.site}")
                    print(f"          Тип ошибки: {type(exc).__name__}")
                    print(f"          Текст ошибки: {exc}")
                    print("          В нейросеть НЕ передаю текст ошибки, чтобы она не писала клиенту про недоступность сайта.")

                    site_text = f"""
                URL: {lead.site}
                Title:
                Meta description:
                Page text: Информации со страницы мало. Напиши сообщение в общем стиле: сайт выглядит слабым, его можно сделать понятнее, современнее и лучше подготовить под поиск.
                """.strip()
                
                log_block("ЭТАП 3: ЗАПРОС В НЕЙРОСЕТЬ")
                pitch = generate_pitch(site_text)

                log_block("ОТВЕТ НЕЙРОСЕТИ", pitch)
                if not controller.ask_continue("Этап 3: ответ нейросети получен и показан"):
                    continue

                client = await send_pitch_to_telegram_with_retry(args, controller, client, lead, pitch)

            except PhoneNumberInvalidError:
                print("[ERROR] Telegram считает номер телефона неверным.")
                print(f"        Строка Excel: {lead.excel_row}")
                print(f"        Телефон: {lead.phone}")
                print("        Перехожу к следующей строке.")
            except TelegramRepeatedError:
                break
            except CRITICAL_TELEGRAM_ERRORS as exc:
                print("[CRITICAL] Telegram ограничил отправку сообщений.")
                print(f"           Тип ошибки: {type(exc).__name__}")
                print(f"           Текст ошибки: {exc}")
                if isinstance(exc, FloodWaitError):
                    print(f"           Нужно подождать секунд: {exc.seconds}")
                print("           Программа останавливается, чтобы не усиливать ограничение.")
                break
            except Exception as exc:
                print("[ERROR] Непредвиденная ошибка на строке.")
                print(f"        Строка Excel: {lead.excel_row}")
                print(f"        Телефон: {lead.phone}")
                print(f"        Сайт: {lead.site}")
                print(f"        Тип ошибки: {type(exc).__name__}")
                print(f"        Текст ошибки: {exc}")
                print("        Перехожу к следующей строке.")

    finally:
        if client is not None:
            print("[TG] Отключаюсь от Telegram.")
            await client.disconnect()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Генерирует сообщения по сайтам из Excel и отправляет их в Telegram по номерам из Excel."
    )
    parser.add_argument(
        "--xlsx",
        default="ip_fio_links_11.xlsx",
        help="Путь к Excel-файлу. По умолчанию: ip_fio_links.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Название листа Excel. Если не указать, берётся активный лист.",
    )
    parser.add_argument(
        "--page",
        type=int,
        default=None,
        help="Страница поисковика, на которой сайт был найден. Передаётся в промт так же, как в site_pitch.py.",
    )
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Автоматический режим: не задавать вопросы 'продолжаем ли' и отправлять без финального подтверждения.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Проверочный режим: всё сделать, но НЕ отправлять сообщения.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Обработать только первые N валидных строк.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help="Начать обработку с указанного номера строки Excel.",
    )
    parser.add_argument(
        "--max-site-chars",
        type=int,
        default=8000,
        help="Сколько символов текста сайта максимум отправлять в нейросеть. По умолчанию 8000.",
    )
    parser.add_argument(
        "--delay",
        type=int,
        default=0,
        help="Пауза в секундах между отправками сообщений после успешной отправки. Поиск Telegram-аккаунтов всегда ждёт 60 секунд."
    )
    parser.add_argument(
        "--queue",
        default=None,
        help="JSONL-очередь лидов от addres-sayta_pipeline.py. Если указать, скрипт читает очередь, а не весь Excel.",
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Не завершаться, а ждать новые строки в --queue.",
    )
    parser.add_argument(
        "--poll",
        type=int,
        default=10,
        help="Как часто проверять очередь, если новых лидов нет. По умолчанию 10 секунд.",
    )
    parser.add_argument(
        "--processed-state",
        default="leads_queue_processed.txt",
        help="Файл, куда записываются уже обработанные лиды из очереди.",
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    print("\nbulk_site_pitch_telega.py запущен")
    print(f"Python: {sys.version}")
    print(f"Excel: {args.xlsx}")
    print(f"Интерактивный режим: {'выключен (--yes)' if args.yes else 'включён'}")
    print(f"Dry-run: {'да, сообщения не отправляются' if args.dry_run else 'нет'}")

    if args.yes and not args.dry_run:
        print("\n[WARN] Включён --yes без --dry-run: скрипт будет отправлять сообщения без вопросов.")
        print("       Используйте это только для контактов, которые дали согласие на такой эксперимент.")

    try:
        if args.queue:
            asyncio.run(process_queue_leads(args))
        else:
            asyncio.run(process_leads(args))
    except KeyboardInterrupt:
        print("\nПрограмма остановлена пользователем.")


if __name__ == "__main__":
    main()
