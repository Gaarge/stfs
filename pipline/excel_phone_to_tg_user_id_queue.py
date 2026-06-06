"""
excel_phone_to_tg_user_id_queue.py

Что делает:
1. НЕ ходит на Rusprofile и НЕ делает curl.
2. Читает уже готовые данные из Excel: телефон и сайт.
3. По телефону первым Telegram-аккаунтом ищет user_id/access_hash.
4. Дописывает найденных людей в leads_queue_user_ids.jsonl.
5. Учитывает старый leads_queue_processed.txt, чтобы не поставить в очередь тех,
   кому бот уже написал по старой схеме excel_row|phone|site.
6. Учитывает уже существующую очередь, чтобы при перезапуске не плодить дубли.

.env:
TELEGRAM_SEARCH_API_ID=...
TELEGRAM_SEARCH_API_HASH=...
TELEGRAM_SEARCH_PHONE=+79999999999
TELEGRAM_SEARCH_SESSION_NAME=tg_search_session

Можно также использовать старые переменные TELEGRAM_API_ID/TELEGRAM_API_HASH/TELEGRAM_PHONE,
если TELEGRAM_SEARCH_* не заданы.

Пример:
python excel_phone_to_tg_user_id_queue.py --xlsx "fio_links_250_unique_appended_from_320.xlsx" --start-row 300
python excel_phone_to_tg_user_id_queue.py --xlsx "fio_links_250_unique_appended_from_320.xlsx" --start-row 300 --limit 20
"""

import argparse
import asyncio
import json
import os
import random
import re
import sys
import time
from getpass import getpass
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from telethon import TelegramClient, functions, types
from telethon.errors import FloodWaitError, PeerFloodError, SessionPasswordNeededError


load_dotenv()

DEFAULT_QUEUE_FILE = "leads_queue_user_ids.jsonl"
DEFAULT_PROCESSED_STATE_FILE = "leads_queue_processed.txt"
DEFAULT_TELEGRAM_SEARCH_SESSION_NAME = os.getenv("TELEGRAM_SEARCH_SESSION_NAME", "my_telegram_search_session")
TELEGRAM_PHONE_SEARCH_DELAY_SECONDS = int(os.getenv("TELEGRAM_PHONE_SEARCH_DELAY_SECONDS", "60"))
CRITICAL_TELEGRAM_ERRORS = (FloodWaitError, PeerFloodError)

PHONE_RE = re.compile(r"^\+\d{10,15}$")


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_phone(raw_phone: str) -> str:
    phone = normalize_cell(raw_phone)
    if not phone:
        return ""
    if phone.startswith("+"):
        cleaned = "+" + re.sub(r"\D", "", phone)
    else:
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("8"):
            cleaned = "+7" + digits[1:]
        elif len(digits) == 11 and digits.startswith("7"):
            cleaned = "+" + digits
        elif len(digits) == 10:
            cleaned = "+7" + digits
        else:
            cleaned = "+" + digits if digits else ""
    return cleaned if PHONE_RE.fullmatch(cleaned) else ""


def normalize_site(raw_site: str, fallback_site: str = "") -> str:
    site = normalize_cell(raw_site)
    if not site:
        return fallback_site
    if not re.match(r"^https?://", site, flags=re.IGNORECASE):
        site = "https://" + site
    return site


def normalize_processed_site(site: str) -> str:
    return normalize_cell(site).rstrip("/")


def safe_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip().replace("'", ""))
    except (TypeError, ValueError):
        return None


def legacy_queue_key(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_cell(site)}"


def legacy_queue_key_normalized(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_processed_site(site)}"


def recipient_keys_for_values(excel_row: int, phone: str, site: str, user_id=None) -> set[str]:
    phone = normalize_cell(phone)
    site = normalize_cell(site)
    keys = {
        legacy_queue_key(excel_row, phone, site),
        legacy_queue_key_normalized(excel_row, phone, site),
        f"phone:{phone}",
    }
    parsed_user_id = safe_int(user_id)
    if parsed_user_id:
        keys.add(f"user_id:{parsed_user_id}")
    return {key for key in keys if key and key != "phone:"}


def load_processed_keys(path: Path) -> set[str]:
    """Читает и старый формат excel_row|phone|site, и новый JSONL-формат."""
    if not path.exists():
        return set()

    processed: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        processed.add(line)

        if line.startswith("{"):
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = safe_int(data.get("excel_row")) or 0
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id") or data.get("telegram_user_id")
            processed.update(recipient_keys_for_values(excel_row, phone, site, user_id))
        else:
            parts = line.split("|", 2)
            if len(parts) == 3:
                row_raw, phone, site = parts
                excel_row = safe_int(row_raw) or 0
                processed.update(recipient_keys_for_values(excel_row, phone, site))
    return processed


def load_queued_keys(path: Path) -> set[str]:
    """Читает текущую очередь, чтобы не добавлять дубль при перезапуске."""
    if not path.exists():
        return set()

    keys: set[str] = set()
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or not line.startswith("{"):
                continue
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = safe_int(data.get("excel_row")) or 0
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id") or data.get("telegram_user_id")
            keys.update(recipient_keys_for_values(excel_row, phone, site, user_id))
    return keys


def find_col_by_header(ws, header_name: str) -> Optional[int]:
    wanted = header_name.strip().lower()
    for col in range(1, ws.max_column + 1):
        value = normalize_cell(ws.cell(row=1, column=col).value).lower()
        if value == wanted:
            return col
    return None


def get_or_create_column(ws, title: str) -> int:
    existing = find_col_by_header(ws, title)
    if existing:
        return existing
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = title
    return col


def detect_last_data_row(ws, relevant_cols: list[int]) -> int:
    """Excel иногда считает max_row как 1048576 из-за форматирования. Ищем последнюю реально заполненную строку."""
    for row in range(ws.max_row, 1, -1):
        for col in relevant_cols:
            if normalize_cell(ws.cell(row=row, column=col).value):
                return row
    return 1


def append_lead_to_queue(
    queue_path: Path,
    excel_row: int,
    phone: str,
    site: str,
    user_id: int,
    access_hash: int,
    username: str = "",
    first_name: str = "",
    last_name: str = "",
) -> None:
    record = {
        "excel_row": excel_row,
        "phone": phone,
        "site": site,
        "user_id": int(user_id),
        "access_hash": int(access_hash),
        "username": username or None,
        "first_name": first_name or None,
        "last_name": last_name or None,
        "legacy_key": legacy_queue_key(excel_row, phone, site),
        "source": "excel_phone_to_tg_user_id_queue.py",
        "queued_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }

    with queue_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()

    print(f"    [QUEUE] Добавлен в очередь: user_id={user_id}, phone={phone}")


async def get_or_create_search_telegram_client(session_name: str) -> TelegramClient:
    api_id_raw = os.getenv("TELEGRAM_SEARCH_API_ID") or os.getenv("TELEGRAM_API_ID")
    api_hash = os.getenv("TELEGRAM_SEARCH_API_HASH") or os.getenv("TELEGRAM_API_HASH")
    my_phone = os.getenv("TELEGRAM_SEARCH_PHONE") or os.getenv("TELEGRAM_PHONE")

    if not api_id_raw or not api_hash or not my_phone:
        raise RuntimeError(
            "Не найдены TELEGRAM_SEARCH_API_ID/TELEGRAM_SEARCH_API_HASH/TELEGRAM_SEARCH_PHONE "
            "или запасные TELEGRAM_API_ID/TELEGRAM_API_HASH/TELEGRAM_PHONE."
        )

    try:
        api_id = int(api_id_raw)
    except ValueError as exc:
        raise RuntimeError("TELEGRAM_SEARCH_API_ID или TELEGRAM_API_ID должен быть числом.") from exc

    client = TelegramClient(session_name, api_id, api_hash)
    print(f"[TG-SEARCH] Подключаюсь к Telegram для поиска user_id. Session: {session_name}")
    await client.connect()

    if not await client.is_user_authorized():
        print("[TG-SEARCH] Сессия поиска ещё не авторизована. Отправляю код входа.")
        await client.send_code_request(my_phone)
        code = input("Введите код из Telegram/SMS для аккаунта поиска: ").strip()
        try:
            await client.sign_in(my_phone, code)
        except SessionPasswordNeededError:
            password = getpass("Включена 2FA. Введите пароль Telegram для аккаунта поиска: ")
            await client.sign_in(password=password)

    print("[TG-SEARCH] Telegram-аккаунт поиска подключён и авторизован.")
    return client


async def find_telegram_user_by_phone(client: TelegramClient, target_phone: str):
    print(f"    [TG-SEARCH] Ищу user_id по номеру: {target_phone}")

    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=target_phone,
        first_name="Temporary",
        last_name="Contact",
    )

    result = await client(functions.contacts.ImportContactsRequest([contact]))
    if not result.users:
        print("    [TG-SEARCH] Пользователь не найден или скрыт настройками приватности.")
        return None

    user = result.users[0]
    print(
        "    [TG-SEARCH] Найден: "
        f"id={getattr(user, 'id', None)}, "
        f"access_hash={getattr(user, 'access_hash', None)}, "
        f"username={getattr(user, 'username', None)}"
    )

    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
        print("    [TG-SEARCH] Временный контакт удалён.")
    except Exception as exc:
        print(f"    [TG-SEARCH][WARN] Не удалось удалить временный контакт: {type(exc).__name__}: {exc}")

    return user


async def maybe_wait_before_tg_search(last_search_at: float) -> float:
    if last_search_at > 0:
        elapsed = time.monotonic() - last_search_at
        remaining = TELEGRAM_PHONE_SEARCH_DELAY_SECONDS - elapsed
        if remaining > 0:
            print(f"    [TG-SEARCH][DELAY] Жду {remaining:.1f} секунд перед следующим поиском по номеру...")
            await asyncio.sleep(remaining)
    return time.monotonic()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Берёт телефон и сайт из Excel, ищет Telegram user_id по телефону и пишет очередь для sender-скрипта. Rusprofile не трогает."
    )
    parser.add_argument("--xlsx", default="fio_links_250_unique_appended_from_320.xlsx", help="Путь к Excel-файлу.")
    parser.add_argument("--sheet", default=None, help="Название листа. Если не указано, берётся активный лист.")
    parser.add_argument("--start-row", type=int, default=2, help="С какой строки Excel начинать. По умолчанию 2.")
    parser.add_argument("--limit", type=int, default=None, help="Максимум строк, которые реально попытаться обработать.")
    parser.add_argument("--phone-header", default="Телефон", help="Заголовок колонки с телефоном. По умолчанию: Телефон.")
    parser.add_argument("--site-header", default="Сайт", help="Заголовок колонки с сайтом. По умолчанию: Сайт.")
    parser.add_argument("--user-id-header", default="Telegram user_id", help="Колонка для user_id.")
    parser.add_argument("--access-hash-header", default="Telegram access_hash", help="Колонка для access_hash.")
    parser.add_argument("--username-header", default="Telegram username", help="Колонка для username.")
    parser.add_argument("--queue", default=DEFAULT_QUEUE_FILE, help="JSONL-очередь для второго скрипта.")
    parser.add_argument("--processed-state", default=DEFAULT_PROCESSED_STATE_FILE, help="Файл уже отправленных лидов.")
    parser.add_argument("--fallback-site", default="https://example.com", help="Что ставить в site, если сайт в Excel пустой. Пустая строка = пропускать без сайта.")
    parser.add_argument("--telegram-session-name", default=DEFAULT_TELEGRAM_SEARCH_SESSION_NAME, help="Имя Telegram session-файла для аккаунта поиска.")
    parser.add_argument("--dry-run", action="store_true", help="Ничего не писать в очередь и Excel, только показать что было бы сделано.")
    return parser


async def main() -> None:
    args = build_arg_parser().parse_args()

    xlsx_path = Path(args.xlsx)
    queue_path = Path(args.queue)
    processed_state_path = Path(args.processed_state)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {xlsx_path.resolve()}")

    processed_keys = load_processed_keys(processed_state_path)
    queued_keys = load_queued_keys(queue_path)

    print("\nexcel_phone_to_tg_user_id_queue.py запущен")
    print(f"Python: {sys.version}")
    print(f"Excel: {xlsx_path.resolve()}")
    print(f"Очередь: {queue_path.resolve()}")
    print(f"Processed: {processed_state_path.resolve()}")
    print(f"Dry-run: {'да' if args.dry_run else 'нет'}")
    print(f"[STATE] Уже отправленных ключей загружено: {len(processed_keys)}")
    print(f"[QUEUE] Уже стоящих в очереди ключей загружено: {len(queued_keys)}")

    wb = load_workbook(xlsx_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    phone_col = find_col_by_header(ws, args.phone_header)
    site_col = find_col_by_header(ws, args.site_header)
    if not phone_col:
        raise RuntimeError(f"Не нашёл колонку с заголовком {args.phone_header!r} в первой строке Excel.")
    if not site_col:
        raise RuntimeError(f"Не нашёл колонку с заголовком {args.site_header!r} в первой строке Excel.")

    user_id_col = get_or_create_column(ws, args.user_id_header)
    access_hash_col = get_or_create_column(ws, args.access_hash_header)
    username_col = get_or_create_column(ws, args.username_header)

    last_row = detect_last_data_row(ws, [phone_col, site_col, user_id_col, access_hash_col])
    print(f"[EXCEL] Лист: {ws.title}")
    print(f"[EXCEL] Колонка телефона: #{phone_col}")
    print(f"[EXCEL] Колонка сайта: #{site_col}")
    print(f"[EXCEL] Колонка user_id: #{user_id_col}")
    print(f"[EXCEL] Колонка access_hash: #{access_hash_col}")
    print(f"[EXCEL] Последняя строка с данными: {last_row}")

    if not args.dry_run:
        wb.save(xlsx_path)

    tg_client = await get_or_create_search_telegram_client(args.telegram_session_name)
    last_tg_search_at = 0.0
    attempted_rows = 0
    added_rows = 0
    skipped_rows = 0

    try:
        for row in range(args.start_row, last_row + 1):
            raw_phone = ws.cell(row=row, column=phone_col).value
            raw_site = ws.cell(row=row, column=site_col).value

            phone = normalize_phone(raw_phone)
            site = normalize_site(raw_site, fallback_site=args.fallback_site)

            if not phone:
                continue
            if not site:
                print(f"[{row}] [SKIP] Есть телефон, но нет сайта, а --fallback-site пустой: {raw_phone!r}")
                skipped_rows += 1
                continue

            attempted_rows += 1
            if args.limit is not None and attempted_rows > args.limit:
                print(f"[LIMIT] Достигнут --limit {args.limit}. Останавливаюсь.")
                break

            existing_user_id = safe_int(ws.cell(row=row, column=user_id_col).value)
            existing_access_hash = safe_int(ws.cell(row=row, column=access_hash_col).value)
            existing_username = normalize_cell(ws.cell(row=row, column=username_col).value)

            print(f"\n[{row}] phone={phone!r}, site={site!r}")

            base_keys = recipient_keys_for_values(row, phone, site, existing_user_id)
            if not base_keys.isdisjoint(processed_keys):
                print("    [SKIP] Уже есть в leads_queue_processed.txt по старой/новой схеме. В очередь не добавляю.")
                skipped_rows += 1
                continue
            if not base_keys.isdisjoint(queued_keys):
                print("    [SKIP] Уже есть в очереди. Дубль не добавляю.")
                skipped_rows += 1
                continue

            if existing_user_id and existing_access_hash:
                print(f"    [EXCEL] Уже есть user_id/access_hash в Excel: {existing_user_id}")
                if not args.dry_run:
                    append_lead_to_queue(
                        queue_path=queue_path,
                        excel_row=row,
                        phone=phone,
                        site=site,
                        user_id=existing_user_id,
                        access_hash=existing_access_hash,
                        username=existing_username,
                    )
                    queued_keys.update(recipient_keys_for_values(row, phone, site, existing_user_id))
                else:
                    print("    [DRY-RUN] В очередь не пишу.")
                added_rows += 1
                continue

            last_tg_search_at = await maybe_wait_before_tg_search(last_tg_search_at)
            user = await find_telegram_user_by_phone(tg_client, phone)
            last_tg_search_at = time.monotonic()

            if user is None:
                print("    [QUEUE] В очередь не добавляю: Telegram user_id не найден.")
                skipped_rows += 1
                continue

            user_id = getattr(user, "id", None)
            access_hash = getattr(user, "access_hash", None)
            username = getattr(user, "username", None)
            first_name = getattr(user, "first_name", None)
            last_name = getattr(user, "last_name", None)

            if not user_id or not access_hash:
                print("    [QUEUE] В очередь не добавляю: Telegram вернул пользователя без user_id/access_hash.")
                skipped_rows += 1
                continue

            if not args.dry_run:
                ws.cell(row=row, column=user_id_col).value = str(user_id)
                ws.cell(row=row, column=access_hash_col).value = str(access_hash)
                ws.cell(row=row, column=username_col).value = username
                wb.save(xlsx_path)

                append_lead_to_queue(
                    queue_path=queue_path,
                    excel_row=row,
                    phone=phone,
                    site=site,
                    user_id=int(user_id),
                    access_hash=int(access_hash),
                    username=username or "",
                    first_name=first_name or "",
                    last_name=last_name or "",
                )
                queued_keys.update(recipient_keys_for_values(row, phone, site, user_id))
            else:
                print(f"    [DRY-RUN] Нашёл бы и записал бы: user_id={user_id}, access_hash={access_hash}")

            added_rows += 1

    except CRITICAL_TELEGRAM_ERRORS as exc:
        print(f"\n[TG-SEARCH][CRITICAL] Telegram ограничил поиск: {type(exc).__name__}: {exc}")
        if isinstance(exc, FloodWaitError):
            print(f"[TG-SEARCH][CRITICAL] Telegram просит подождать секунд: {exc.seconds}")
        print("Останавливаю скрипт, чтобы не усиливать ограничение.")
    finally:
        if not args.dry_run:
            wb.save(xlsx_path)
        wb.close()
        await tg_client.disconnect()
        print("[TG-SEARCH] Отключился от Telegram.")

    print("\nГотово.")
    print(f"Строк с телефонами просмотрено: {attempted_rows}")
    print(f"Добавлено/готово к добавлению в очередь: {added_rows}")
    print(f"Пропущено: {skipped_rows}")


if __name__ == "__main__":
    asyncio.run(main())
