import json
import random
import re
import subprocess
import time
from pathlib import Path
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook


EXCEL_FILE = "ip_fio_links_str_15.xlsx"
BASE_URL = "https://www.rusprofile.ru"

# Нумерация колонок в openpyxl начинается с 1:
# 3 = колонка C
LINK_COL = 3

# Начиная со 2 строки, если в 1 строке заголовки
START_ROW = 2

# Пауза между обычными запросами.
# Лучше не ставить 1 секунду: Rusprofile быстро начинает показывать капчу.
REQUEST_DELAY_SECONDS_MIN = 4.0
REQUEST_DELAY_SECONDS_MAX = 12.0

# После каждых 49 реально обработанных строк спать 2 минуты
BATCH_ROWS = 25
BATCH_SLEEP_SECONDS = 120

# Что делать, если Rusprofile отдал капчу
CAPTCHA_SLEEP_SECONDS = 1800  # 30 минут
CAPTCHA_RETRIES_PER_ROW = 10   # 1 повтор после ожидания. Если снова капча — остановка.

# Очередь для второго скрипта. Первый скрипт дописывает сюда готовые лиды.
QUEUE_FILE = "leads_queue.jsonl"


class CaptchaDetected(RuntimeError):
    pass


def curl_html(url: str) -> str:
    """Скачивает HTML через curl и возвращает текст страницы."""
    result = subprocess.run(
        [
            "curl",
            "-L",
            "--compressed",
            "--connect-timeout", "15",
            "--max-time", "40",
            "-A",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36",
            url,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    if result.returncode != 0:
        raise RuntimeError(f"curl error {result.returncode}: {result.stderr}")

    return result.stdout


def is_captcha_page(html: str) -> bool:
    """Проверяет, что Rusprofile вместо карточки отдал страницу капчи."""
    markers = [
        "captcha-section",
        "g-recaptcha",
        "Активность с вашего IP-адреса была распознана как автоматическая",
        "captcha-validate",
        "Я не робот",
    ]
    return any(marker in html for marker in markers)


def clean_phone(phone: str) -> str:
    """Приводит телефон к виду +79999999999, если это возможно."""
    phone = phone.strip()
    if not phone:
        return ""

    # href tel обычно уже нормальный, например +74951607080
    if phone.startswith("+"):
        return "+" + re.sub(r"\D", "", phone)

    digits = re.sub(r"\D", "", phone)
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    return "+" + digits


def extract_phone_and_site(html: str) -> tuple[str, str]:
    """Извлекает телефон и сайт из HTML Rusprofile."""
    if is_captcha_page(html):
        raise CaptchaDetected("Rusprofile отдал страницу капчи вместо карточки")

    soup = BeautifulSoup(html, "html.parser")

    phone = ""
    site = ""

    # 1) Телефон: <a href="tel:+74951607080" itemprop="telephone">
    phone_tag = soup.select_one('a[href^="tel:"]')
    if phone_tag:
        phone = phone_tag.get("href", "").replace("tel:", "").strip()

    if not phone:
        phone_match = re.search(r'href=["\']tel:([^"\']+)', html, flags=re.IGNORECASE)
        if phone_match:
            phone = phone_match.group(1).strip()

    phone = clean_phone(phone)

    # 2) Сайт: <span class="company-info__contact site iconer"> ... <a href="http://...">
    site_tag = soup.select_one('.company-info__contact.site a[href^="http"]')
    if site_tag:
        site = site_tag.get("href", "").strip()

    if not site:
        for contact in soup.select(".company-info__contact"):
            classes = contact.get("class") or []
            text = contact.get_text(" ", strip=True)
            if "site" not in classes and "Сайт" not in text:
                continue
            a = contact.select_one('a[href^="http"]')
            if a:
                site = a.get("href", "").strip()
                break

    if not site:
        site_match = re.search(
            r'class=["\'][^"\']*company-info__contact[^"\']*site[^"\']*["\'][\s\S]*?href=["\'](https?://[^"\']+)',
            html,
            flags=re.IGNORECASE,
        )
        if site_match:
            site = site_match.group(1).strip()

    return phone, site


def make_full_url(value) -> str:
    """Из значения 3-й колонки делает полный URL."""
    if value is None:
        return ""

    link = str(value).strip()
    if not link:
        return ""

    if link.startswith("http://") or link.startswith("https://"):
        return link

    if link.startswith("/"):
        return urljoin(BASE_URL, link)

    return urljoin(BASE_URL + "/", link)


def append_lead_to_queue(queue_path: Path, excel_row: int, phone: str, site: str, rusprofile_url: str) -> None:
    """Дописывает готовую строку в очередь для bulk_site_pitch_telega.

    Формат JSONL: одна строка = один готовый лид.
    Это специально не Excel: второй скрипт может читать очередь, пока первый продолжает писать Excel.
    """
    if not phone or not site:
        return

    record = {
        "excel_row": excel_row,
        "phone": phone,
        "site": site,
        "rusprofile_url": rusprofile_url,
    }

    with queue_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()

    print(f"    [QUEUE] Лид добавлен в очередь: {queue_path}")


def get_or_create_column(ws, title: str) -> int:
    """Находит колонку по заголовку или создает новую, чтобы не плодить дубли при перезапуске."""
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col).value).strip() == title:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = title
    return col


def main():
    path = Path(EXCEL_FILE)
    queue_path = Path(QUEUE_FILE)

    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path.resolve()}")

    wb = load_workbook(path)
    ws = wb.active

    phone_col = get_or_create_column(ws, "Телефон")
    site_col = get_or_create_column(ws, "Сайт")

    processed_rows = 0

    for row in range(START_ROW, ws.max_row + 1):
        raw_link = ws.cell(row=row, column=LINK_COL).value
        url = make_full_url(raw_link)

        if not url:
            continue

        print(f"[{row}] {url}")

        row_done = False
        captcha_retries = 0

        while not row_done:
            try:
                html = curl_html(url)

                if is_captcha_page(html):
                    raise CaptchaDetected("Rusprofile отдал капчу")

                phone, site = extract_phone_and_site(html)

                ws.cell(row=row, column=phone_col).value = phone
                ws.cell(row=row, column=site_col).value = site
                wb.save(path)

                print(f"    phone={phone!r}, site={site!r}")

                if phone and site:
                    append_lead_to_queue(queue_path, row, phone, site, url)
                else:
                    print("    [QUEUE] В очередь не добавляю: нет телефона или сайта.")

                row_done = True

            except CaptchaDetected as e:
                wb.save(path)
                captcha_retries += 1

                print(f"    CAPTCHA: {e}")
                print("    ВАЖНО: пустые значения в Excel НЕ записываю, чтобы не затереть данные.")

                if captcha_retries > CAPTCHA_RETRIES_PER_ROW:
                    print("    Капча повторилась после ожидания. Останавливаю скрипт безопасно.")
                    print(f"    Потом продолжи с этой строки: START_ROW = {row}")
                    wb.save(path)
                    return

                print(f"    Сплю {CAPTCHA_SLEEP_SECONDS} секунд и попробую эту же строку еще раз...")
                time.sleep(CAPTCHA_SLEEP_SECONDS)

            except Exception as e:
                print(f"    ERROR: {e}")
                ws.cell(row=row, column=phone_col).value = ""
                ws.cell(row=row, column=site_col).value = ""
                wb.save(path)
                row_done = True

        processed_rows += 1

        if processed_rows % BATCH_ROWS == 0:
            print(
                f"[PAUSE] Обработано {processed_rows} строк с ссылками. "
                f"Сплю {BATCH_SLEEP_SECONDS} секунд..."
            )
            time.sleep(BATCH_SLEEP_SECONDS)

        delay = random.uniform(REQUEST_DELAY_SECONDS_MIN, REQUEST_DELAY_SECONDS_MAX)
        print(f"[DELAY] Сплю {delay:.1f} секунд перед следующим запросом...")
        time.sleep(delay)

    wb.save(path)
    print(f"Готово. Файл обновлён: {path.resolve()}")


if __name__ == "__main__":
    main()
