import asyncio
import json
import os
import random
import re
import subprocess
import time
from getpass import getpass
from pathlib import Path
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from telethon import TelegramClient, functions, types
from telethon.errors import FloodWaitError, PeerFloodError, SessionPasswordNeededError


load_dotenv()

EXCEL_FILE = "fio_links_250_unique_appended_from_320.xlsx"
BASE_URL = "https://www.rusprofile.ru"

# Нумерация колонок в openpyxl начинается с 1:
# 3 = колонка C
LINK_COL = 3

# Начиная со 48 строки, если в 1 строке заголовки
START_ROW = 300

# Пауза между обычными запросами.
# Лучше не ставить 1 секунду: Rusprofile быстро начинает показывать капчу.
REQUEST_DELAY_SECONDS_MIN = 4.0
REQUEST_DELAY_SECONDS_MAX = 12.0

USER_AGENTS = [
    # 1. Chrome / Windows 10
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36",

    # 2. Chrome / Windows 11
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36",

    # 3. Chrome / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36",

    # 4. Chrome / Linux
    "Mozilla/5.0 (X11; Linux x86_64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36",

    # 5. Firefox / Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) "
    "Gecko/20100101 Firefox/126.0",

    # 6. Firefox / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:126.0) "
    "Gecko/20100101 Firefox/126.0",

    # 7. Firefox / Linux
    "Mozilla/5.0 (X11; Linux x86_64; rv:126.0) "
    "Gecko/20100101 Firefox/126.0",

    # 8. Edge / Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",

    # 9. Edge / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",

    # 10. Yandex Browser / Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 YaBrowser/24.4.0.0 Safari/537.36",

    # 11. Yandex Browser / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 YaBrowser/24.4.0.0 Safari/537.36",

    # 12. Opera / Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36 OPR/109.0.0.0",

    # 13. Opera / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36 OPR/109.0.0.0",

    # 14. Safari / macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/17.0 Safari/605.1.15",

    # 15. Safari / older macOS-style desktop
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 11_7_10) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/16.6 Safari/605.1.15",
]


# После каждых 49 реально обработанных строк спать 2 минуты
BATCH_ROWS = 25
BATCH_SLEEP_SECONDS = 120

# Что делать, если Rusprofile отдал капчу
CAPTCHA_SLEEP_SECONDS = 1800  # 30 минут
CAPTCHA_RETRIES_PER_ROW = 10   # 1 повтор после ожидания. Если снова капча — остановка.

# Очередь для второго скрипта. Первый скрипт дописывает сюда готовые лиды.
# Новая очередь: первый аккаунт Telegram уже нашёл user_id по номеру.
QUEUE_FILE = "leads_queue_user_ids.jsonl"

# Старый/новый файл отправленных. Читаем его здесь, чтобы не ставить в очередь тех,
# кому второй аккаунт уже написал раньше по старой схеме.
PROCESSED_STATE_FILE = "leads_queue_processed.txt"

# Отдельный аккаунт для поиска user_id. Если переменные не заданы, используются общие TELEGRAM_*
TELEGRAM_SEARCH_SESSION_NAME = os.getenv("TELEGRAM_SEARCH_SESSION_NAME", "my_telegram_search_session")
TELEGRAM_PHONE_SEARCH_DELAY_SECONDS = int(os.getenv("TELEGRAM_PHONE_SEARCH_DELAY_SECONDS", "60"))
CRITICAL_TELEGRAM_ERRORS = (FloodWaitError, PeerFloodError)


class CaptchaDetected(RuntimeError):
    pass


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_processed_site(site: str) -> str:
    return normalize_cell(site).rstrip("/")


def legacy_queue_key(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_cell(site)}"


def legacy_queue_key_normalized(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_processed_site(site)}"


def recipient_keys_for_values(excel_row: int, phone: str, site: str, user_id=None) -> set[str]:
    keys = {
        legacy_queue_key(excel_row, phone, site),
        legacy_queue_key_normalized(excel_row, phone, site),
        f"phone:{normalize_cell(phone)}",
    }
    if user_id:
        keys.add(f"user_id:{int(user_id)}")
    return {key for key in keys if key and key != "phone:"}


def load_processed_keys(path: Path) -> set[str]:
    """Читает и старый формат excel_row|phone|site, и новый JSONL-формат."""
    if not path.exists():
        return set()

    processed: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        processed.add(line)
        if line.startswith("{"):
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = int(data.get("excel_row") or 0)
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id")
            try:
                user_id = int(user_id) if user_id else None
            except (TypeError, ValueError):
                user_id = None
            processed.update(recipient_keys_for_values(excel_row, phone, site, user_id))
        else:
            parts = line.split("|", 2)
            if len(parts) == 3:
                row_raw, phone, site = parts
                try:
                    excel_row = int(row_raw)
                except ValueError:
                    excel_row = 0
                processed.update(recipient_keys_for_values(excel_row, phone, site))
    return processed


def load_queued_keys(path: Path) -> set[str]:
    """Читает текущую очередь, чтобы не добавлять дубль при перезапуске."""
    if not path.exists():
        return set()
    keys: set[str] = set()
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or not line.startswith("{"):
                continue
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = int(data.get("excel_row") or 0)
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id") or data.get("telegram_user_id")
            try:
                user_id = int(user_id) if user_id else None
            except (TypeError, ValueError):
                user_id = None
            keys.update(recipient_keys_for_values(excel_row, phone, site, user_id))
    return keys


async def get_or_create_search_telegram_client() -> TelegramClient:
    api_id_raw = os.getenv("TELEGRAM_SEARCH_API_ID") or os.getenv("TELEGRAM_API_ID")
    api_hash = os.getenv("TELEGRAM_SEARCH_API_HASH") or os.getenv("TELEGRAM_API_HASH")
    my_phone = os.getenv("TELEGRAM_SEARCH_PHONE") or os.getenv("TELEGRAM_PHONE")

    if not api_id_raw or not api_hash or not my_phone:
        raise RuntimeError(
            "Не найдены TELEGRAM_SEARCH_API_ID/TELEGRAM_SEARCH_API_HASH/TELEGRAM_SEARCH_PHONE "
            "или запасные TELEGRAM_API_ID/TELEGRAM_API_HASH/TELEGRAM_PHONE."
        )

    try:
        api_id = int(api_id_raw)
    except ValueError as exc:
        raise RuntimeError("TELEGRAM_SEARCH_API_ID или TELEGRAM_API_ID должен быть числом.") from exc

    client = TelegramClient(TELEGRAM_SEARCH_SESSION_NAME, api_id, api_hash)
    print(f"[TG-SEARCH] Подключаюсь к Telegram для поиска user_id. Session: {TELEGRAM_SEARCH_SESSION_NAME}")
    await client.connect()

    if not await client.is_user_authorized():
        print("[TG-SEARCH] Сессия поиска ещё не авторизована. Отправляю код входа.")
        await client.send_code_request(my_phone)
        code = input("Введите код из Telegram/SMS для аккаунта поиска: ").strip()
        try:
            await client.sign_in(my_phone, code)
        except SessionPasswordNeededError:
            password = getpass("Включена 2FA. Введите пароль Telegram для аккаунта поиска: ")
            await client.sign_in(password=password)

    print("[TG-SEARCH] Telegram-аккаунт поиска подключён и авторизован.")
    return client


async def find_telegram_user_by_phone(client: TelegramClient, target_phone: str):
    print(f"[TG-SEARCH] Ищу Telegram user_id по номеру: {target_phone}")
    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=target_phone,
        first_name="Temporary",
        last_name="Contact",
    )
    result = await client(functions.contacts.ImportContactsRequest([contact]))
    if not result.users:
        print("[TG-SEARCH] Пользователь не найден или скрыт настройками приватности.")
        return None

    user = result.users[0]
    print(
        "[TG-SEARCH] Найден: "
        f"id={getattr(user, 'id', None)}, "
        f"access_hash={getattr(user, 'access_hash', None)}, "
        f"username={getattr(user, 'username', None)}"
    )
    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
        print("[TG-SEARCH] Временный контакт удалён.")
    except Exception as exc:
        print(f"[TG-SEARCH][WARN] Не удалось удалить временный контакт: {type(exc).__name__}: {exc}")
    return user


def curl_html(url: str) -> str:
    """Скачивает HTML через curl и возвращает текст страницы."""
    user_agent = random.choice(USER_AGENTS)

    result = subprocess.run(
        [
            "curl",
            "-L",
            "--compressed",
            "--connect-timeout", "15",
            "--max-time", "40",
            "-A", user_agent,
            url,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    if result.returncode != 0:
        raise RuntimeError(f"curl error {result.returncode}: {result.stderr}")

    return result.stdout


def is_captcha_page(html: str) -> bool:
    """Проверяет, что Rusprofile вместо карточки отдал страницу капчи."""
    markers = [
        "captcha-section",
        "g-recaptcha",
        "Активность с вашего IP-адреса была распознана как автоматическая",
        "captcha-validate",
        "Я не робот",
    ]
    return any(marker in html for marker in markers)


def clean_phone(phone: str) -> str:
    """Приводит телефон к виду +79999999999, если это возможно."""
    phone = phone.strip()
    if not phone:
        return ""

    # href tel обычно уже нормальный, например +74951607080
    if phone.startswith("+"):
        return "+" + re.sub(r"\D", "", phone)

    digits = re.sub(r"\D", "", phone)
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    return "+" + digits


def extract_phone_and_site(html: str) -> tuple[str, str]:
    """Извлекает телефон и сайт из HTML Rusprofile."""
    if is_captcha_page(html):
        raise CaptchaDetected("Rusprofile отдал страницу капчи вместо карточки")

    soup = BeautifulSoup(html, "html.parser")

    phone = ""
    site = ""

    # 1) Телефон: <a href="tel:+74951607080" itemprop="telephone">
    phone_tag = soup.select_one('a[href^="tel:"]')
    if phone_tag:
        phone = phone_tag.get("href", "").replace("tel:", "").strip()

    if not phone:
        phone_match = re.search(r'href=["\']tel:([^"\']+)', html, flags=re.IGNORECASE)
        if phone_match:
            phone = phone_match.group(1).strip()

    phone = clean_phone(phone)

    # 2) Сайт: <span class="company-info__contact site iconer"> ... <a href="http://...">
    site_tag = soup.select_one('.company-info__contact.site a[href^="http"]')
    if site_tag:
        site = site_tag.get("href", "").strip()

    if not site:
        for contact in soup.select(".company-info__contact"):
            classes = contact.get("class") or []
            text = contact.get_text(" ", strip=True)
            if "site" not in classes and "Сайт" not in text:
                continue
            a = contact.select_one('a[href^="http"]')
            if a:
                site = a.get("href", "").strip()
                break

    if not site:
        site_match = re.search(
            r'class=["\'][^"\']*company-info__contact[^"\']*site[^"\']*["\'][\s\S]*?href=["\'](https?://[^"\']+)',
            html,
            flags=re.IGNORECASE,
        )
        if site_match:
            site = site_match.group(1).strip()

    return phone, site


def make_full_url(value) -> str:
    """Из значения 3-й колонки делает полный URL."""
    if value is None:
        return ""

    link = str(value).strip()
    if not link:
        return ""

    if link.startswith("http://") or link.startswith("https://"):
        return link

    if link.startswith("/"):
        return urljoin(BASE_URL, link)

    return urljoin(BASE_URL + "/", link)


def append_lead_to_queue(queue_path: Path, excel_row: int, phone: str, site: str, rusprofile_url: str, telegram_user) -> None:
    """Дописывает готовую строку в очередь для второго аккаунта.

    Формат JSONL: одна строка = один лид, уже найденный первым Telegram-аккаунтом.
    Второй скрипт больше не ищет по телефону, а отправляет по user_id/access_hash.
    """
    if not phone or not site or telegram_user is None:
        return

    record = {
        "excel_row": excel_row,
        "phone": phone,
        "site": site,
        "rusprofile_url": rusprofile_url,
        "user_id": getattr(telegram_user, "id", None),
        "access_hash": getattr(telegram_user, "access_hash", None),
        "username": getattr(telegram_user, "username", None),
        "first_name": getattr(telegram_user, "first_name", None),
        "last_name": getattr(telegram_user, "last_name", None),
        "legacy_key": legacy_queue_key(excel_row, phone, site),
    }

    with queue_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()

    print(f"    [QUEUE] Лид с Telegram user_id добавлен в очередь: {queue_path}")


def get_or_create_column(ws, title: str) -> int:
    """Находит колонку по заголовку или создает новую, чтобы не плодить дубли при перезапуске."""
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col).value).strip() == title:
            return col

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = title
    return col


async def main():
    path = Path(EXCEL_FILE)
    queue_path = Path(QUEUE_FILE)
    processed_state_path = Path(PROCESSED_STATE_FILE)

    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path.resolve()}")

    processed_keys = load_processed_keys(processed_state_path)
    queued_keys = load_queued_keys(queue_path)

    print(f"[STATE] Уже отправленные читаю из: {processed_state_path.resolve()}")
    print(f"[STATE] Ключей отправленных/старого формата загружено: {len(processed_keys)}")
    print(f"[QUEUE] Новая очередь с user_id: {queue_path.resolve()}")
    print(f"[QUEUE] Ключей уже стоящих в очереди: {len(queued_keys)}")

    tg_client = await get_or_create_search_telegram_client()
    last_tg_search_at = 0.0

    try:
        wb = load_workbook(path)
        ws = wb.active

        phone_col = get_or_create_column(ws, "Телефон")
        site_col = get_or_create_column(ws, "Сайт")
        user_id_col = get_or_create_column(ws, "Telegram user_id")
        username_col = get_or_create_column(ws, "Telegram username")

        processed_rows = 0

        for row in range(START_ROW, ws.max_row + 1):
            raw_link = ws.cell(row=row, column=LINK_COL).value
            url = make_full_url(raw_link)

            if not url:
                continue

            print(f"[{row}] {url}")

            row_done = False
            captcha_retries = 0

            while not row_done:
                try:
                    html = curl_html(url)

                    if is_captcha_page(html):
                        raise CaptchaDetected("Rusprofile отдал капчу")

                    phone, site = extract_phone_and_site(html)
                    if phone and not site:
                        site = "https://example.com"

                    ws.cell(row=row, column=phone_col).value = phone
                    ws.cell(row=row, column=site_col).value = site
                    wb.save(path)

                    print(f"    phone={phone!r}, site={site!r}")

                    if phone and site:
                        base_keys = recipient_keys_for_values(row, phone, site)
                        if not base_keys.isdisjoint(processed_keys):
                            print("    [SKIP] Уже есть в leads_queue_processed.txt по старой/новой схеме. В очередь не добавляю.")
                        elif not base_keys.isdisjoint(queued_keys):
                            print("    [SKIP] Уже есть в очереди. Дубль не добавляю.")
                        else:
                            now = time.monotonic()
                            remaining = TELEGRAM_PHONE_SEARCH_DELAY_SECONDS - (now - last_tg_search_at)
                            if last_tg_search_at and remaining > 0:
                                print(f"    [TG-SEARCH][DELAY] Жду {remaining:.1f} секунд перед следующим поиском по номеру...")
                                await asyncio.sleep(remaining)

                            user = await find_telegram_user_by_phone(tg_client, phone)
                            last_tg_search_at = time.monotonic()

                            if user is None:
                                print("    [QUEUE] В очередь не добавляю: Telegram user_id не найден.")
                            else:
                                ws.cell(row=row, column=user_id_col).value = getattr(user, "id", None)
                                ws.cell(row=row, column=username_col).value = getattr(user, "username", None)
                                wb.save(path)
                                append_lead_to_queue(queue_path, row, phone, site, url, user)
                                queued_keys.update(recipient_keys_for_values(row, phone, site, getattr(user, "id", None)))
                    else:
                        print("    [QUEUE] В очередь не добавляю: нет телефона или сайта.")

                    row_done = True

                except CaptchaDetected as e:
                    wb.save(path)
                    captcha_retries += 1

                    print(f"    CAPTCHA: {e}")
                    print("    ВАЖНО: пустые значения в Excel НЕ записываю, чтобы не затереть данные.")

                    if captcha_retries > CAPTCHA_RETRIES_PER_ROW:
                        print("    Капча повторилась после ожидания. Останавливаю скрипт безопасно.")
                        print(f"    Потом продолжи с этой строки: START_ROW = {row}")
                        wb.save(path)
                        return

                    print(f"    Сплю {CAPTCHA_SLEEP_SECONDS} секунд и попробую эту же строку еще раз...")
                    time.sleep(CAPTCHA_SLEEP_SECONDS)

                except CRITICAL_TELEGRAM_ERRORS as e:
                    wb.save(path)
                    print(f"    [TG-SEARCH][CRITICAL] Telegram ограничил поиск: {type(e).__name__}: {e}")
                    print("    Останавливаю первый скрипт, чтобы не усиливать ограничение.")
                    return

                except Exception as e:
                    print(f"    ERROR: {e}")
                    ws.cell(row=row, column=phone_col).value = ""
                    ws.cell(row=row, column=site_col).value = ""
                    wb.save(path)
                    row_done = True

            processed_rows += 1

            if processed_rows % BATCH_ROWS == 0:
                print(
                    f"[PAUSE] Обработано {processed_rows} строк с ссылками. "
                    f"Сплю {BATCH_SLEEP_SECONDS} секунд..."
                )
                time.sleep(BATCH_SLEEP_SECONDS)

            delay = random.uniform(REQUEST_DELAY_SECONDS_MIN, REQUEST_DELAY_SECONDS_MAX)
            print(f"[DELAY] Сплю {delay:.1f} секунд перед следующим запросом...")
            time.sleep(delay)

        wb.save(path)
        print(f"Готово. Файл обновлён: {path.resolve()}")
    finally:
        await tg_client.disconnect()
        print("[TG-SEARCH] Отключился от Telegram.")


if __name__ == "__main__":
    asyncio.run(main())
