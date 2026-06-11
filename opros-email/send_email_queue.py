import argparse
import base64
import json
import os
import re
import smtplib
import socket
import ssl
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from getpass import getpass
from pathlib import Path

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency. Install requirements first:\n"
        "  python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR.parent / "users-from-rusprofile" / "rusprofile_users_from_test.xlsx"
DEFAULT_QUEUE = BASE_DIR.parent / "users-from-rusprofile" / "rusprofile_email_queue.txt"
DEFAULT_PROCESSED = BASE_DIR / "email_users_processed.jsonl"
DEFAULT_ERRORS_FILE = BASE_DIR / "send_errors.jsonl"
DEFAULT_MESSAGE_FILE = BASE_DIR / "message.txt"
DEFAULT_GMAIL_CREDENTIALS_FILE = BASE_DIR / "gmail_credentials.json"
DEFAULT_GMAIL_TOKEN_FILE = BASE_DIR / "gmail_token.json"
DEFAULT_OPENROUTER_MODEL = "google/gemini-3-flash-preview"
DEFAULT_FROM_EMAIL = "vladram3707@gmail.com"
DEFAULT_SUBJECT = "Разработка сайта"
DEFAULT_BATCH_SIZE = 30
DEFAULT_BATCH_PAUSE_SECONDS = 10 * 60
DEFAULT_SEND_DELAY_SECONDS = 120
DEFAULT_SMTP_TIMEOUT_SECONDS = 30
MAX_TOTAL_PER_RUN = 450
GMAIL_API_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
CONTACT_BLOCK = (
    "Если для Вас это актуально, то свяжитесь со мной: "
    "tg - @dev_all_sites или позвоните по телефону - 89162040241"
)

EMAIL_RE = re.compile(
    r"(?<![\w.+-])([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})(?![\w.-])",
    flags=re.IGNORECASE,
)

EMAIL_HEADERS = (
    "email",
    "e-mail",
    "mail",
    "почта",
    "электронная почта",
    "email адрес",
    "email-адрес",
)
FIO_HEADERS = ("фио", "name", "имя")
SITE_HEADERS = ("сайт", "site", "website", "web")
LINK_HEADERS = ("ссылка", "link", "rusprofile link", "rusprofile")


def load_env_files() -> list[Path]:
    if not load_dotenv:
        return []

    candidates = [
        BASE_DIR / ".env",
        BASE_DIR.parent / ".env",
        BASE_DIR.parent.parent / ".env",
        Path.cwd() / ".env",
    ]
    loaded = []
    seen = set()

    for path in candidates:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists() and load_dotenv(resolved, override=False):
            loaded.append(resolved)

    return loaded


LOADED_ENV_FILES = load_env_files()
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")


@dataclass
class EmailLead:
    row_no: int
    email: str
    fio: str = ""
    site: str = ""
    rusprofile_link: str = ""


@dataclass
class LeadSource:
    path: Path
    label: str


@dataclass
class RunStats:
    started_at: datetime
    finished_at: datetime | None = None
    source_label: str = "-"
    source_path: str = "-"
    processed_path: str = "-"
    errors_path: str = "-"
    loaded_count: int = 0
    processed_loaded_count: int = 0
    pending_count: int = 0
    run_limit: int = 0
    sent: int = 0
    would_send: int = 0
    skipped: int = 0
    errors: int = 0
    dry_run: bool = False
    stop_reason: str = "not started"
    exit_code: int = 0


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value) -> str:
    return normalize_cell(value).lower()


def normalize_email(value: str) -> str:
    email = normalize_cell(value)
    if email.lower().startswith("mailto:"):
        email = email[7:]
    email = email.strip(" \t\r\n,;<>")
    return email.lower()


def extract_emails_from_text(value: str) -> list[str]:
    emails = []
    seen = set()
    for match in EMAIL_RE.findall(normalize_cell(value).replace("mailto:", " ")):
        email = normalize_email(match)
        if email and email not in seen:
            seen.add(email)
            emails.append(email)
    return emails


def extract_emails_from_cell(cell) -> list[str]:
    parts = [normalize_cell(cell.value)]
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink:
        parts.append(normalize_cell(getattr(hyperlink, "target", "")))
        parts.append(normalize_cell(getattr(hyperlink, "display", "")))

    emails = []
    seen = set()
    for part in parts:
        for email in extract_emails_from_text(part):
            if email not in seen:
                seen.add(email)
                emails.append(email)
    return emails


def find_col_by_headers(ws, headers: tuple[str, ...]) -> int | None:
    wanted = {normalize_header(header) for header in headers}
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=1, column=col).value) in wanted:
            return col
    return None


def read_excel_leads(path: Path) -> list[EmailLead]:
    if not path.exists():
        raise SystemExit(f"Excel file not found: {path.resolve()}")

    workbook = load_workbook(path, data_only=True)
    ws = workbook.active

    email_col = find_col_by_headers(ws, EMAIL_HEADERS)
    if not email_col:
        raise SystemExit(f"Email column not found in {path.resolve()}")

    fio_col = find_col_by_headers(ws, FIO_HEADERS)
    site_col = find_col_by_headers(ws, SITE_HEADERS)
    link_col = find_col_by_headers(ws, LINK_HEADERS)

    leads = []
    seen_emails = set()
    for row_no in range(2, ws.max_row + 1):
        row_emails = extract_emails_from_cell(ws.cell(row=row_no, column=email_col))
        if not row_emails:
            continue

        fio = normalize_cell(ws.cell(row=row_no, column=fio_col).value) if fio_col else ""
        site = normalize_cell(ws.cell(row=row_no, column=site_col).value) if site_col else ""
        rusprofile_link = normalize_cell(ws.cell(row=row_no, column=link_col).value) if link_col else ""

        for email in row_emails:
            if email in seen_emails:
                continue
            seen_emails.add(email)
            leads.append(
                EmailLead(
                    row_no=row_no,
                    email=email,
                    fio=fio,
                    site=site,
                    rusprofile_link=rusprofile_link,
                )
            )

    return leads


def read_text_queue_leads(path: Path) -> list[EmailLead]:
    if not path.exists():
        raise SystemExit(f"Email queue file not found: {path.resolve()}")

    content = path.read_text(encoding="utf-8-sig")
    if content and not content.endswith(("\n", "\r")):
        print("[QUEUE][WARN] Последняя строка без перевода строки пропущена, чтобы не поймать недописанный email.")
        content = content.rsplitlines(keepends=True)
        content = "".join(content[:-1])

    leads = []
    seen_emails = set()
    for line_no, line in enumerate(content.splitlines(), start=1):
        raw = normalize_cell(line)
        if not raw or raw.startswith("#"):
            continue
        for email in extract_emails_from_text(raw):
            if email in seen_emails:
                continue
            seen_emails.add(email)
            leads.append(EmailLead(row_no=line_no, email=email))

    return leads


def read_leads_from_args(args) -> tuple[list[EmailLead], LeadSource]:
    if args.queue:
        queue_path = Path(args.queue)
        return read_text_queue_leads(queue_path), LeadSource(queue_path, "Queue")

    xlsx_path = Path(args.xlsx)
    return read_excel_leads(xlsx_path), LeadSource(xlsx_path, "Excel")


def local_time(value: datetime | None = None) -> datetime:
    return (value or datetime.now(timezone.utc)).astimezone()


def format_local_time(value: datetime | None = None) -> str:
    return local_time(value).strftime("%Y-%m-%d %H:%M:%S %Z")


def format_seconds(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    minutes, second = divmod(seconds, 60)
    hours, minute = divmod(minutes, 60)
    if hours:
        return f"{hours}ч {minute}м {second}с"
    if minute:
        return f"{minute}м {second}с"
    return f"{second}с"


def resolve_run_limit(args, pending_count: int) -> int:
    if args.max_total < 0:
        raise SystemExit("--max-total must be >= 0")
    if args.max_per_run is not None and args.max_per_run < 0:
        raise SystemExit("--max-per-run must be >= 0")

    requested = min(args.max_total, MAX_TOTAL_PER_RUN)
    if args.max_total > MAX_TOTAL_PER_RUN:
        print(f"[LIMIT][WARN] --max-total больше {MAX_TOTAL_PER_RUN}. Ограничиваю запуск до {MAX_TOTAL_PER_RUN} писем.")

    if args.max_per_run is not None:
        requested = min(requested, args.max_per_run)

    return min(pending_count, requested)


def batch_position(letter_no: int, batch_size: int) -> tuple[int, int, int]:
    if batch_size <= 0:
        return 1, letter_no, 0
    batch_no = ((letter_no - 1) // batch_size) + 1
    in_batch_no = ((letter_no - 1) % batch_size) + 1
    return batch_no, in_batch_no, batch_size


def print_sleep(kind: str, seconds: float) -> None:
    wake_at = local_time() + timedelta(seconds=max(0, seconds))
    print(f"[{kind}] Пауза {format_seconds(seconds)}. Продолжу примерно в {format_local_time(wake_at)}.")


def maybe_sleep_after_letter(args, processed_this_run: int, run_limit: int) -> None:
    if args.dry_run or processed_this_run >= run_limit:
        return

    if args.batch_size > 0 and processed_this_run % args.batch_size == 0:
        if args.batch_pause > 0:
            print_sleep("BATCH PAUSE", args.batch_pause)
            time.sleep(args.batch_pause)
        return

    if args.delay > 0:
        print_sleep("DELAY", args.delay)
        time.sleep(args.delay)


def smtp_error_text(exc: Exception) -> str:
    raw = getattr(exc, "smtp_error", "")
    if isinstance(raw, bytes):
        return raw.decode("utf-8", errors="replace")
    return str(raw or exc)


def is_smtp_policy_or_limit_error(exc: Exception) -> bool:
    if not isinstance(exc, smtplib.SMTPResponseException):
        return False
    text = smtp_error_text(exc).lower()
    markers = (
        "quota",
        "limit",
        "rate",
        "too many",
        "blocked",
        "suspicious",
        "policy",
        "spam",
        "temporarily rejected",
    )
    return any(marker in text for marker in markers)


def print_summary(stats: RunStats) -> None:
    stats.finished_at = datetime.now(timezone.utc)
    duration = (stats.finished_at - stats.started_at).total_seconds()
    sent_label = "Would send" if stats.dry_run else "Sent"

    print("\n" + "=" * 90)
    print("EMAIL RUN SUMMARY")
    print("=" * 90)
    print(f"Started: {format_local_time(stats.started_at)}")
    print(f"Finished: {format_local_time(stats.finished_at)}")
    print(f"Duration: {format_seconds(duration)}")
    print(f"Source: {stats.source_label} {stats.source_path}")
    print(f"Loaded emails: {stats.loaded_count}")
    print(f"Already processed at start: {stats.processed_loaded_count}")
    print(f"Pending at start: {stats.pending_count}")
    print(f"Run limit: {stats.run_limit}")
    print(f"{sent_label}: {stats.would_send if stats.dry_run else stats.sent}")
    if stats.dry_run:
        print("Sent: 0")
    print(f"Skipped: {stats.skipped}")
    print(f"Errors: {stats.errors}")
    print(f"Stop reason: {stats.stop_reason}")
    print(f"Processed file: {stats.processed_path}")
    print(f"Errors file: {stats.errors_path}")


def normalize_url(raw_url: str) -> str:
    url = normalize_cell(raw_url)
    if not url:
        return ""
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        url = "https://" + url
    return url


def fetch_site_text(url: str, max_chars: int = 8000) -> str:
    normalized_url = normalize_url(url)
    if not normalized_url:
        return ""

    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; SiteAnalyzer/1.0)",
    }

    print(f"[SITE] Загружаю сайт: {normalized_url}")
    response = requests.get(normalized_url, headers=headers, timeout=15)
    response.raise_for_status()
    print(f"[SITE] HTTP статус: {response.status_code}")
    print(f"[SITE] Content-Type: {response.headers.get('Content-Type', 'не указан')}")

    if response.apparent_encoding:
        response.encoding = response.apparent_encoding

    soup = BeautifulSoup(response.text, "html.parser")

    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title and soup.title.string else ""

    meta_description = ""
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        meta_description = meta["content"].strip()

    text = soup.get_text(separator=" ")
    text = re.sub(r"\s+", " ", text).strip()

    result = f"""
URL: {normalized_url}
Title: {title}
Meta description: {meta_description}
Page text: {text}
""".strip()

    print(f"[SITE] Title: {title!r}")
    print(f"[SITE] Meta description: {meta_description!r}")
    print(f"[SITE] Длина очищенного текста до обрезки: {len(result)} символов")
    print(f"[SITE] В нейросеть уйдёт максимум: {max_chars} символов")

    return result[:max_chars]


def generate_pitch(site_text: str = "", model: str = DEFAULT_OPENROUTER_MODEL) -> str:
    prompt = f"""
Ты пишешь короткое рекламное сообщение потенциальному клиенту от веб-разработчика.

Ситуация:

* Мы не знаем точно, есть ли у клиента сайт.
* Сообщение должно одинаково хорошо подходить и тем, у кого сайта пока нет, и тем, у кого сайт уже есть.
* Не пиши фразы вроде: "я нашёл ваш сайт", "увидел ваш сайт", "наткнулся на ваш сайт", "у вас нет сайта", "ваш сайт плохой".
* Не утверждай то, чего мы точно не знаем.
* Можно использовать нейтральную формулировку: "если сайта пока нет — можно сделать его с нуля, если сайт уже есть — можно обновить его и доработать".

Стиль:

* Пиши просто, по-человечески.
* Без пафоса.
* Без слов: "статусный", "экспертиза", "упаковать", "презентабельный", "визуальная форма", "достойный", "серьезные клиенты".
* Не используй канцелярит.
* Не пиши слишком вежливо и корпоративно.
* Тон: прямой, уверенный, спокойный.
* Максимум 6-7 предложений.
* Ни слова не говори про цену.
* Обязательно поздоровайся.
* Прощаться не надо.

Что нужно сказать:

1. Меня зовут Владислав.
2. Я уже несколько лет разрабатываю сайты.
3. Также занимаюсь SEO-анализом — то есть слежу за тем, чтобы сайт могли увидеть как можно больше целевых клиентов.
4. Скажи, что я могу сделать сайт с нуля или обновить уже существующий.
5. Сайт должен быть красивым, понятным, современным и удобным для людей.
6. Хороший сайт помогает бизнесу выглядеть понятнее, вызывает больше доверия и помогает получать больше заявок.
7. Напиши уверенно, но честно: "сайт будет сделан так, чтобы он чаще попадал на первые страницы поиска".
8. Объясни, что чем чаще сайт появляется на первых страницах поиска, тем больше людей его видят, а значит у бизнеса может быть больше клиентов.
9. Упомяни, что я разрабатывал сайты для учебных заведений, адвокатских контор и интернет-магазинов.
10. Напиши, что если нужно, могу прислать свои работы.
11. Напиши, что работаю по договору.
12. Не дави на человека и не пугай его потерей клиентов.

Данные о бизнесе клиента, если они есть, чтобы аккуратно подстроить сообщение под сферу:
{site_text}

Выдай только готовое сообщение клиенту, без пояснений.
""".strip()

    if not OPENROUTER_API_KEY:
        raise RuntimeError(
            "Не найден OPENROUTER_API_KEY. Добавь его в .env или экспортируй переменную окружения."
        )

    print("[AI] Отправляю запрос в OpenRouter.")
    print(f"[AI] Модель: {model}")
    print(f"[AI] Длина промта: {len(prompt)} символов")

    response = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://example.com",
            "X-Title": "Site Pitch Generator",
        },
        json={
            "model": model,
            "messages": [
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
            "temperature": 0.7,
            "max_tokens": 300,
        },
        timeout=60,
    )

    print(f"[AI] HTTP статус OpenRouter: {response.status_code}")

    if response.status_code != 200:
        raise RuntimeError(f"OpenRouter error {response.status_code}: {response.text}")

    data = response.json()
    message = data["choices"][0]["message"]["content"].strip()
    print(f"[AI] Ответ получен. Длина сообщения: {len(message)} символов")
    return message


def build_site_fallback_text(lead: EmailLead) -> str:
    return f"""
URL: {normalize_url(lead.site)}
Title:
Meta description:
Page text: Информации со страницы мало. Напиши сообщение в общем стиле: сайт выглядит слабым, его можно сделать понятнее, современнее и лучше подготовить под поиск.
""".strip()


def build_openrouter_message(lead: EmailLead, max_site_chars: int, model: str) -> str:
    site_text = ""

    if lead.site:
        try:
            site_text = fetch_site_text(lead.site, max_chars=max_site_chars)
            print("[OK] Сайт успешно загружен.")
            print(f"[SITE] Длина текста сайта для нейросети: {len(site_text)} символов")
        except requests.RequestException as exc:
            print("[WARNING] Сайт не удалось загрузить, но лид НЕ будет пропущен.")
            print(f"          Строка Excel: {lead.row_no}")
            print(f"          Сайт: {lead.site}")
            print(f"          Тип ошибки: {type(exc).__name__}")
            print("          В нейросеть НЕ передаю текст ошибки, чтобы она не писала клиенту про недоступность сайта.")
            site_text = build_site_fallback_text(lead)
    else:
        print("[SITE][WARN] В строке Excel нет сайта. Сгенерирую общее сообщение без данных о бизнесе.")

    return generate_pitch(site_text, model=model)


def add_contact_block(message: str) -> str:
    message = message.strip()
    if CONTACT_BLOCK.lower() in message.lower():
        return message
    return f"{message}\n\n{CONTACT_BLOCK}"


def load_static_message(args) -> str:
    if args.message:
        return args.message.strip()

    message_path = Path(args.message_file)
    if not message_path.exists():
        message_path.parent.mkdir(parents=True, exist_ok=True)
        message_path.touch()
    message = message_path.read_text(encoding="utf-8").strip()
    if not message:
        raise SystemExit(f"Static message is empty. Fill {message_path} or pass --message.")
    return message


def load_processed_emails(path: Path) -> set[str]:
    if not path.exists():
        return set()

    emails = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith("{"):
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                data = {}
            email = normalize_email(data.get("email", ""))
            if email:
                emails.add(email)
                continue
        emails.update(extract_emails_from_text(line))
    return emails


def append_processed(
    path: Path,
    lead: EmailLead,
    subject: str,
    message: str,
    run_letter_no: int | None = None,
    sent_at: datetime | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    sent_at = sent_at or datetime.now(timezone.utc)
    record = {
        "status": "sent",
        "sent_at": sent_at.isoformat(),
        "run_letter_no": run_letter_no,
        "row_no": lead.row_no,
        "email": lead.email,
        "fio": lead.fio or None,
        "site": lead.site or None,
        "rusprofile_link": lead.rusprofile_link or None,
        "subject": subject,
        "message_len": len(message),
    }
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()


def append_error(path: Path, lead: EmailLead, exc: Exception, run_letter_no: int | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    record = {
        "error_at": datetime.now(timezone.utc).isoformat(),
        "run_letter_no": run_letter_no,
        "row_no": lead.row_no,
        "email": lead.email,
        "fio": lead.fio or None,
        "site": lead.site or None,
        "rusprofile_link": lead.rusprofile_link or None,
        "error_type": type(exc).__name__,
        "error": str(exc),
    }
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()


def resolve_smtp_password(args) -> str:
    env_candidates = []
    if args.smtp_password_env:
        env_candidates.append(args.smtp_password_env)
    env_candidates.extend(["GMAIL_APP_PASSWORD", "SMTP_PASSWORD", "EMAIL_PASSWORD"])

    seen = set()
    for name in env_candidates:
        if not name or name in seen:
            continue
        seen.add(name)
        value = os.getenv(name)
        if value:
            return value

    if sys.stdin.isatty():
        return getpass(f"SMTP password/app password for {args.smtp_user}: ")
    return ""


def create_ipv4_connection(host: str, port: int, timeout: float) -> socket.socket:
    errors = []
    try:
        addresses = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
    except socket.gaierror as exc:
        raise OSError(f"Cannot resolve IPv4 address for {host}:{port}: {exc}") from exc

    for family, socktype, proto, _canonname, sockaddr in addresses:
        sock = socket.socket(family, socktype, proto)
        sock.settimeout(timeout)
        try:
            sock.connect(sockaddr)
            return sock
        except OSError as exc:
            errors.append(f"{sockaddr[0]}:{sockaddr[1]} -> {exc}")
            sock.close()

    raise OSError(f"Cannot connect to {host}:{port} over IPv4. Attempts: {'; '.join(errors) or '-'}")


class IPv4SMTP(smtplib.SMTP):
    def _get_socket(self, host, port, timeout):
        return create_ipv4_connection(host, port, timeout)


class IPv4SMTPSSL(smtplib.SMTP_SSL):
    def _get_socket(self, host, port, timeout):
        raw_sock = create_ipv4_connection(host, port, timeout)
        try:
            return self.context.wrap_socket(raw_sock, server_hostname=host)
        except Exception:
            raw_sock.close()
            raise


def smtp_attempts(args) -> list[tuple[str, int, bool, str]]:
    attempts = [(args.smtp_host, args.smtp_port, args.smtp_ssl, "primary")]
    if not args.smtp_fallback:
        return attempts

    fallbacks = [
        (args.smtp_host, 587, False, "fallback STARTTLS"),
        (args.smtp_host, 465, True, "fallback SSL"),
    ]
    seen = {(args.smtp_host, args.smtp_port, args.smtp_ssl)}
    for host, port, use_ssl, label in fallbacks:
        key = (host, port, use_ssl)
        if key not in seen:
            seen.add(key)
            attempts.append((host, port, use_ssl, label))
    return attempts


def open_smtp_connection(host: str, port: int, use_ssl: bool, timeout: float, force_ipv4: bool):
    context = ssl.create_default_context()
    if use_ssl:
        smtp_class = IPv4SMTPSSL if force_ipv4 else smtplib.SMTP_SSL
        return smtp_class(host, port, context=context, timeout=timeout)

    smtp_class = IPv4SMTP if force_ipv4 else smtplib.SMTP
    smtp = smtp_class(host, port, timeout=timeout)
    try:
        smtp.ehlo()
        smtp.starttls(context=context)
        smtp.ehlo()
        return smtp
    except Exception:
        try:
            smtp.quit()
        except Exception:
            pass
        raise


def connect_smtp(args, password: str):
    last_exc: Exception | None = None
    for host, port, use_ssl, label in smtp_attempts(args):
        mode = "SSL" if use_ssl else "STARTTLS"
        family = "IPv4" if args.smtp_ipv4 else "system DNS"
        print(f"[SMTP] Пробую {label}: {host}:{port} {mode}, {family}, timeout={args.smtp_timeout:.1f}s")
        smtp = None
        try:
            smtp = open_smtp_connection(host, port, use_ssl, args.smtp_timeout, args.smtp_ipv4)
            smtp.login(args.smtp_user, password)
            print(f"[SMTP] Подключение успешно: {host}:{port} {mode}")
            return smtp
        except smtplib.SMTPAuthenticationError:
            if smtp is not None:
                try:
                    smtp.quit()
                except Exception:
                    pass
            raise
        except Exception as exc:
            last_exc = exc
            if smtp is not None:
                try:
                    smtp.quit()
                except Exception:
                    pass
            print(f"[SMTP][WARN] Не получилось через {host}:{port} {mode}: {type(exc).__name__}: {exc}")

    if last_exc is not None:
        raise last_exc
    raise OSError("No SMTP connection attempts configured.")


def missing_gmail_api_dependency(exc: ImportError) -> SystemExit:
    return SystemExit(
        "Missing Gmail API dependencies. Install them first:\n"
        "  venv/bin/python -m pip install google-api-python-client google-auth-oauthlib google-auth-httplib2"
    )


def connect_gmail_api(args):
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise missing_gmail_api_dependency(exc) from exc

    credentials_path = Path(args.gmail_credentials)
    token_path = Path(args.gmail_token)

    if not credentials_path.exists():
        raise SystemExit(
            f"Gmail API credentials file not found: {credentials_path.resolve()}\n"
            "Create an OAuth client JSON in Google Cloud Console and save it there."
        )

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_API_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("[GMAIL API] Обновляю OAuth-токен.")
            creds.refresh(Request())
        else:
            print("[GMAIL API] Нужна первая авторизация Google в браузере.")
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), GMAIL_API_SCOPES)
            creds = flow.run_local_server(port=args.gmail_oauth_port, open_browser=True)

        token_path.parent.mkdir(parents=True, exist_ok=True)
        token_path.write_text(creds.to_json(), encoding="utf-8")
        print(f"[GMAIL API] Токен сохранен: {token_path}")

    service = build("gmail", "v1", credentials=creds)
    print("[GMAIL API] Подключение готово.")
    return service


def send_gmail_api_message(service, args, lead: EmailLead, body: str) -> None:
    email_message = make_email_message(args, lead, body)
    raw_message = base64.urlsafe_b64encode(email_message.as_bytes()).decode("ascii")
    result = service.users().messages().send(userId="me", body={"raw": raw_message}).execute()
    message_id = result.get("id", "-")
    print(f"[GMAIL API] Message ID: {message_id}")


def is_gmail_api_policy_or_limit_error(exc: Exception) -> bool:
    status = getattr(getattr(exc, "resp", None), "status", None)
    if status in {403, 429}:
        return True

    text = str(exc).lower()
    markers = (
        "ratelimitexceeded",
        "user ratelimitexceeded",
        "daily limit",
        "quota",
        "too many",
        "spam",
        "policy",
    )
    return any(marker in text for marker in markers)


def is_policy_or_limit_error(exc: Exception) -> bool:
    return is_smtp_policy_or_limit_error(exc) or is_gmail_api_policy_or_limit_error(exc)


def make_email_message(args, lead: EmailLead, body: str) -> EmailMessage:
    email_message = EmailMessage()
    email_message["From"] = args.from_email
    email_message["To"] = lead.email
    email_message["Subject"] = args.subject
    email_message["Reply-To"] = args.from_email
    email_message.set_content(body)
    return email_message


def send_email(smtp, args, lead: EmailLead, body: str) -> None:
    smtp.send_message(make_email_message(args, lead, body))


def send_email_with_retry(smtp, args, password: str, lead: EmailLead, body: str):
    try:
        send_email(smtp, args, lead, body)
        return smtp
    except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError, OSError) as first_exc:
        print(f"[SMTP][WARN] Соединение оборвалось: {type(first_exc).__name__}: {first_exc}")
        print("[SMTP][RETRY] Переподключаюсь и повторяю это же письмо один раз.")
        try:
            try:
                smtp.quit()
            except Exception:
                pass
            smtp = connect_smtp(args, password)
            send_email(smtp, args, lead, body)
            return smtp
        except Exception as second_exc:
            raise second_exc from first_exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate OpenRouter messages and send them to emails from Rusprofile Excel.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Excel file with Email column.")
    parser.add_argument("--queue", default="", help=f"Text queue with emails. Example: {DEFAULT_QUEUE}")
    parser.add_argument("--processed", default=str(DEFAULT_PROCESSED), help="Processed state JSONL.")
    parser.add_argument("--errors", default=str(DEFAULT_ERRORS_FILE), help="Error log JSONL.")
    parser.add_argument("--message-file", default=str(DEFAULT_MESSAGE_FILE), help="Static message text file for --use-static-message.")
    parser.add_argument("--message", default="", help="Static message text from command line for --use-static-message.")
    parser.add_argument("--use-static-message", action="store_true", help="Send --message or --message-file instead of generating through OpenRouter.")
    parser.add_argument("--openrouter-model", default=os.getenv("OPENROUTER_MODEL", DEFAULT_OPENROUTER_MODEL), help="OpenRouter model.")
    parser.add_argument("--max-site-chars", type=int, default=8000, help="Max site text chars to send to OpenRouter.")
    parser.add_argument("--send-method", choices=("smtp", "gmail-api"), default=os.getenv("EMAIL_SEND_METHOD", "smtp"), help="Delivery method: smtp or gmail-api over HTTPS.")
    parser.add_argument("--from-email", default=os.getenv("EMAIL_FROM", DEFAULT_FROM_EMAIL), help="Email From header.")
    parser.add_argument("--smtp-user", default=os.getenv("SMTP_USER", os.getenv("EMAIL_USER", DEFAULT_FROM_EMAIL)), help="SMTP login.")
    parser.add_argument("--smtp-password-env", default=os.getenv("SMTP_PASSWORD_ENV", "GMAIL_APP_PASSWORD"), help="Env var with SMTP password/app password.")
    parser.add_argument("--smtp-host", default=os.getenv("SMTP_HOST", "smtp.gmail.com"), help="SMTP host.")
    parser.add_argument("--smtp-port", type=int, default=int(os.getenv("SMTP_PORT", "465")), help="SMTP port.")
    parser.add_argument("--smtp-ssl", action=argparse.BooleanOptionalAction, default=os.getenv("SMTP_SSL", "1") != "0", help="Use SMTP over SSL.")
    parser.add_argument("--smtp-timeout", type=float, default=float(os.getenv("SMTP_TIMEOUT_SECONDS", str(DEFAULT_SMTP_TIMEOUT_SECONDS))), help="SMTP connect timeout in seconds.")
    parser.add_argument("--smtp-ipv4", action=argparse.BooleanOptionalAction, default=os.getenv("SMTP_IPV4", "1") != "0", help="Force IPv4 for SMTP connections.")
    parser.add_argument("--smtp-fallback", action=argparse.BooleanOptionalAction, default=os.getenv("SMTP_FALLBACK", "1") != "0", help="Try 587 STARTTLS and 465 SSL fallback when the primary SMTP connection fails.")
    parser.add_argument("--gmail-credentials", default=os.getenv("GMAIL_CREDENTIALS_FILE", str(DEFAULT_GMAIL_CREDENTIALS_FILE)), help="OAuth client JSON for --send-method gmail-api.")
    parser.add_argument("--gmail-token", default=os.getenv("GMAIL_TOKEN_FILE", str(DEFAULT_GMAIL_TOKEN_FILE)), help="Saved OAuth token JSON for --send-method gmail-api.")
    parser.add_argument("--gmail-oauth-port", type=int, default=int(os.getenv("GMAIL_OAUTH_PORT", "0")), help="Local OAuth callback port. 0 means auto-pick.")
    parser.add_argument("--subject", default=os.getenv("EMAIL_SUBJECT", DEFAULT_SUBJECT), help="Email subject.")
    parser.add_argument("--delay", type=float, default=float(os.getenv("EMAIL_SEND_DELAY_SECONDS", str(DEFAULT_SEND_DELAY_SECONDS))), help="Delay after each successful send inside a batch.")
    parser.add_argument("--batch-size", type=int, default=int(os.getenv("EMAIL_BATCH_SIZE", str(DEFAULT_BATCH_SIZE))), help="Emails to send before a long batch pause. Use 0 to disable batch pauses.")
    parser.add_argument("--batch-pause", type=float, default=float(os.getenv("EMAIL_BATCH_PAUSE_SECONDS", str(DEFAULT_BATCH_PAUSE_SECONDS))), help="Long pause after every full batch.")
    parser.add_argument("--max-total", type=int, default=int(os.getenv("EMAIL_MAX_TOTAL", str(MAX_TOTAL_PER_RUN))), help=f"Max emails to send this run. Hard capped at {MAX_TOTAL_PER_RUN}.")
    parser.add_argument("--error-sleep", type=float, default=float(os.getenv("EMAIL_ERROR_SLEEP_SECONDS", "60")), help="Sleep after send error.")
    parser.add_argument("--max-per-run", type=int, default=None, help="Deprecated alias limit. The effective limit is min(--max-total, --max-per-run, 450).")
    parser.add_argument("--dry-run", action="store_true", help="Do not send and do not write processed file.")
    parser.add_argument("--yes", action="store_true", help="Do not ask before each generation/send.")
    return parser


def run(args, stats: RunStats) -> None:
    if args.delay < 0:
        raise SystemExit("--delay must be >= 0")
    if args.batch_size < 0:
        raise SystemExit("--batch-size must be >= 0")
    if args.batch_pause < 0:
        raise SystemExit("--batch-pause must be >= 0")
    if args.error_sleep < 0:
        raise SystemExit("--error-sleep must be >= 0")
    if args.send_method == "smtp" and args.smtp_timeout <= 0:
        raise SystemExit("--smtp-timeout must be > 0")

    processed_path = Path(args.processed)
    errors_path = Path(args.errors)
    static_message = load_static_message(args) if args.use_static_message else None

    leads, lead_source = read_leads_from_args(args)
    processed_emails = load_processed_emails(processed_path)
    pending_count = sum(1 for lead in leads if lead.email not in processed_emails)
    run_limit = resolve_run_limit(args, pending_count)

    stats.source_label = lead_source.label
    stats.source_path = str(lead_source.path.resolve())
    stats.processed_path = str(processed_path.resolve())
    stats.errors_path = str(errors_path.resolve())
    stats.loaded_count = len(leads)
    stats.processed_loaded_count = len(processed_emails)
    stats.pending_count = pending_count
    stats.run_limit = run_limit
    stats.dry_run = args.dry_run
    stats.stop_reason = "running"

    print(f"{lead_source.label}: {lead_source.path.resolve()}")
    print(f"Processed: {processed_path.resolve()}")
    print(f"Errors: {errors_path.resolve()}")
    print(f"Loaded .env files: {', '.join(str(path) for path in LOADED_ENV_FILES) or '-'}")
    print(f"Emails loaded: {len(leads)}")
    print(f"Processed emails loaded: {len(processed_emails)}")
    print(f"Pending emails at start: {pending_count}")
    print(f"Max total this run: {run_limit} (hard cap: {MAX_TOTAL_PER_RUN})")
    print(f"Batch size: {args.batch_size or 'disabled'}")
    print(f"Delay between emails: {format_seconds(args.delay)}")
    print(f"Batch pause: {format_seconds(args.batch_pause) if args.batch_size else 'disabled'}")
    print(f"Dry-run: {'yes' if args.dry_run else 'no'}")
    print(f"Message source: {'static message' if static_message is not None else 'OpenRouter'}")
    if static_message is None:
        print(f"OpenRouter model: {args.openrouter_model}")
    print(f"Send method: {args.send_method}")
    if args.send_method == "smtp":
        print(f"SMTP: {args.smtp_user}@{args.smtp_host}:{args.smtp_port}")
        print(f"SMTP mode: {'SSL' if args.smtp_ssl else 'STARTTLS'}, IPv4: {'yes' if args.smtp_ipv4 else 'no'}, fallback: {'yes' if args.smtp_fallback else 'no'}")
    else:
        print(f"Gmail credentials: {Path(args.gmail_credentials).resolve()}")
        print(f"Gmail token: {Path(args.gmail_token).resolve()}")
    print(f"Subject: {args.subject}")

    if run_limit <= 0:
        stats.stop_reason = "nothing to send"
        print("[DONE] Нет новых email для отправки в этом запуске.")
        return

    if not args.use_static_message and not OPENROUTER_API_KEY:
        raise SystemExit("OPENROUTER_API_KEY is empty. Add it to .env or pass --use-static-message.")

    processed_this_run = 0

    smtp = None
    gmail_service = None
    smtp_password = ""
    if args.dry_run:
        print("[DRY-RUN] Email delivery connection is skipped.")
    elif args.send_method == "smtp":
        smtp_password = resolve_smtp_password(args)
        if not smtp_password:
            raise SystemExit(
                "SMTP password is empty. For Gmail, create an app password and put it into "
                "GMAIL_APP_PASSWORD, SMTP_PASSWORD, or EMAIL_PASSWORD."
            )
        print("[SMTP] Подключаюсь к почте.")
        smtp = connect_smtp(args, smtp_password)
        print("[SMTP] Подключение успешно.")
    else:
        print("[GMAIL API] Подключаюсь к Gmail API.")
        gmail_service = connect_gmail_api(args)

    try:
        for lead in leads:
            if lead.email in processed_emails:
                print(f"[SKIP][строка {lead.row_no}] Уже отправляли раньше: {lead.email}")
                stats.skipped += 1
                continue
            if processed_this_run >= run_limit:
                stats.stop_reason = f"run limit reached ({run_limit})"
                print(f"[LIMIT] Достигнут лимит запуска: {run_limit}")
                break

            letter_no = processed_this_run + 1
            batch_no, in_batch_no, batch_size = batch_position(letter_no, args.batch_size)
            batch_label = (
                f"batch={batch_no}, inside_batch={in_batch_no}/{batch_size}"
                if batch_size
                else f"batch=disabled, inside_batch={letter_no}"
            )
            title = (
                f"[{letter_no} письмо][строка {lead.row_no}] email={lead.email} "
                f"{batch_label} "
                f"fio={lead.fio or '-'} "
                f"site={lead.site or '-'} "
                f"rusprofile={lead.rusprofile_link or '-'}"
            )
            print("\n" + title)

            if static_message is None and not args.yes:
                answer = input("Generate OpenRouter email? Type 'generate' to continue, anything else to skip: ").strip().lower()
                if answer not in {"generate", "g", "yes", "y", "да", "send"}:
                    print(f"[{letter_no} письмо][SKIP] Пропущено до генерации OpenRouter.")
                    stats.skipped += 1
                    continue

            try:
                message = static_message or build_openrouter_message(
                    lead,
                    max_site_chars=args.max_site_chars,
                    model=args.openrouter_model,
                )
                message = add_contact_block(message)

                print("\n" + "#" * 90)
                print("# EMAIL MESSAGE")
                print("#" * 90)
                print(message)

                if args.dry_run:
                    stats.would_send += 1
                    processed_this_run += 1
                    print(
                        f"[{letter_no} письмо][DRY-RUN] Было бы отправлено в {format_local_time()} "
                        f"| email={lead.email} | строка={lead.row_no} | message_len={len(message)}"
                    )
                else:
                    if not args.yes:
                        answer = input("Send this email? Type 'send' to send, anything else to skip: ").strip().lower()
                        if answer != "send":
                            print(f"[{letter_no} письмо][SKIP] Пропущено пользователем.")
                            stats.skipped += 1
                            continue

                    if args.send_method == "smtp":
                        smtp = send_email_with_retry(smtp, args, smtp_password, lead, message)
                    else:
                        send_gmail_api_message(gmail_service, args, lead, message)
                    sent_at = datetime.now(timezone.utc)
                    append_processed(processed_path, lead, args.subject, message, run_letter_no=letter_no, sent_at=sent_at)
                    processed_emails.add(lead.email)
                    stats.sent += 1
                    processed_this_run += 1
                    print(
                        f"[{letter_no} письмо][OK] Отправлено в {format_local_time(sent_at)} "
                        f"| email={lead.email} | строка={lead.row_no} | message_len={len(message)}"
                    )
                    print(f"[{letter_no} письмо][STATE] Записано в processed: {processed_path}")

                if processed_this_run >= run_limit:
                    stats.stop_reason = f"run limit reached ({run_limit})"
                    print(f"[LIMIT] Достигнут лимит запуска: {run_limit}")
                    break

                maybe_sleep_after_letter(args, processed_this_run, run_limit)
            except smtplib.SMTPAuthenticationError as exc:
                stats.errors += 1
                append_error(errors_path, lead, exc, run_letter_no=letter_no)
                stats.stop_reason = "SMTP authentication error"
                print(f"[{letter_no} письмо][SMTP][STOP] Ошибка авторизации: {exc}")
                break
            except Exception as exc:
                stats.errors += 1
                append_error(errors_path, lead, exc, run_letter_no=letter_no)
                print(f"[{letter_no} письмо][ERROR] {type(exc).__name__}: {exc}")
                if is_policy_or_limit_error(exc):
                    stats.stop_reason = f"delivery policy/limit error: {type(exc).__name__}"
                    print("[EMAIL][STOP] Похоже на лимит, блокировку или policy/spam-ошибку отправки. Останавливаю запуск.")
                    break
                if args.error_sleep > 0 and not args.dry_run:
                    print_sleep("ERROR SLEEP", args.error_sleep)
                    time.sleep(args.error_sleep)
    finally:
        if smtp is not None:
            try:
                smtp.quit()
            except Exception:
                pass

    if stats.stop_reason == "running":
        stats.stop_reason = "queue finished"


def main() -> None:
    args = build_parser().parse_args()
    stats = RunStats(started_at=datetime.now(timezone.utc), dry_run=args.dry_run)
    try:
        run(args, stats)
    except KeyboardInterrupt:
        stats.exit_code = 130
        stats.stop_reason = "interrupted by user"
        print("\n[STOP] Остановлено пользователем.")
    except SystemExit as exc:
        stats.exit_code = exc.code if isinstance(exc.code, int) else 1
        if stats.stop_reason in {"not started", "running"}:
            stats.stop_reason = f"stopped: {exc}"
        if stats.exit_code:
            print(f"[STOP] {exc}")
    except Exception as exc:
        stats.exit_code = 1
        stats.stop_reason = f"fatal error: {type(exc).__name__}"
        print(f"[FATAL] {type(exc).__name__}: {exc}")
    finally:
        print_summary(stats)

    if stats.exit_code:
        raise SystemExit(stats.exit_code)


if __name__ == "__main__":
    main()
