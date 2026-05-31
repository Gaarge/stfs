import subprocess
import time
from pathlib import Path
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook


EXCEL_FILE = "ip_fio_links_56.xlsx"
BASE_URL = "https://www.rusprofile.ru"

# Нумерация колонок в openpyxl начинается с 1:
# 3 = колонка C
LINK_COL = 3

# Начиная со 2 строки, если в 1 строке заголовки
START_ROW = 2

# Пауза между запросами, чтобы не долбить сайт слишком часто
REQUEST_DELAY_SECONDS = 1.0


def curl_html(url: str) -> str:
    """Скачивает HTML через curl и возвращает текст страницы."""
    result = subprocess.run(
        [
            "curl",
            "-L",
            "--compressed",
            "--connect-timeout", "15",
            "--max-time", "40",
            "-A",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36",
            url,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    if result.returncode != 0:
        raise RuntimeError(f"curl error {result.returncode}: {result.stderr}")

    return result.stdout


def extract_phone_and_site(html: str) -> tuple[str, str]:
    """Извлекает телефон и сайт из HTML Rusprofile."""
    soup = BeautifulSoup(html, "html.parser")

    phone = ""
    site = ""

    # 1) Телефон: <a href="tel:+78003330882" itemprop="telephone">
    phone_tag = soup.select_one('a[href^="tel:"]')
    if phone_tag:
        phone = phone_tag.get("href", "").replace("tel:", "").strip()
        # если хотите красивый текст вместо +78003330882:
        # phone = phone_tag.get_text(" ", strip=True)

    # 2) Сайт: ищем контактный блок с подписью "Сайт"
    for contact in soup.select(".company-info__contact"):
        label = contact.get_text(" ", strip=True)
        if "Сайт" in label:
            a = contact.select_one('a[href^="http"]')
            if a:
                site = a.get("href", "").strip()
                break

    # Запасной вариант, если класс/структура поменяется
    if not site:
        for span in soup.find_all("span"):
            if span.get_text(strip=True) == "Сайт":
                parent = span.find_parent()
                if parent:
                    a = parent.find("a", href=True)
                    if a and a["href"].startswith("http"):
                        site = a["href"].strip()
                        break

    return phone, site


def make_full_url(value) -> str:
    """Из значения 3-й колонки делает полный URL."""
    if value is None:
        return ""

    link = str(value).strip()
    if not link:
        return ""

    # если в ячейке уже полный URL
    if link.startswith("http://") or link.startswith("https://"):
        return link

    # если в ячейке /ip/306213203100037
    if link.startswith("/"):
        return urljoin(BASE_URL, link)

    # если вдруг в ячейке ip/306213203100037
    return urljoin(BASE_URL + "/", link)


def main():
    path = Path(EXCEL_FILE)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path.resolve()}")

    wb = load_workbook(path)
    ws = wb.active

    phone_col = ws.max_column + 1
    site_col = ws.max_column + 2

    ws.cell(row=1, column=phone_col).value = "Телефон"
    ws.cell(row=1, column=site_col).value = "Сайт"

    for row in range(START_ROW, ws.max_row + 1):
        raw_link = ws.cell(row=row, column=LINK_COL).value
        url = make_full_url(raw_link)

        if not url:
            continue

        print(f"[{row}] {url}")

        try:
            html = curl_html(url)
            phone, site = extract_phone_and_site(html)

            ws.cell(row=row, column=phone_col).value = phone
            ws.cell(row=row, column=site_col).value = site

            print(f"    phone={phone!r}, site={site!r}")

        except Exception as e:
            print(f"    ERROR: {e}")
            ws.cell(row=row, column=phone_col).value = ""
            ws.cell(row=row, column=site_col).value = ""

        time.sleep(REQUEST_DELAY_SECONDS)

    wb.save(path)
    print(f"Готово. Файл обновлён: {path.resolve()}")


if __name__ == "__main__":
    main()
