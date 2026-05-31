"""
bulk_site_pitch_telega.py

Что делает:
1. Читает Excel-файл ip_fio_links.xlsx или путь из --xlsx.
2. Берёт последние два столбца каждой строки: телефон и сайт.
3. Пропускает строку, если телефон/сайт пустые или телефон не строго в формате +79999999999.
4. Загружает текст сайта.
5. Отправляет в OpenRouter ровно тот же промт, что в site_pitch.py, но с сайтом из Excel.
6. Находит получателя в Telegram по номеру телефона так же, как в telega.py.
7. Показывает сообщение и отправляет его только после подтверждения.

Интерактивный режим включён по умолчанию.
Чтобы убрать вопросы и выполнять всё автоматически: добавь флаг --yes.
Для безопасной проверки без отправки: добавь флаг --dry-run.

Перед запуском создай .env рядом со скриптом:
OPENROUTER_API_KEY=...
TELEGRAM_API_ID=...
TELEGRAM_API_HASH=...
TELEGRAM_PHONE=+79999999999

Установка зависимостей:
pip install openpyxl requests beautifulsoup4 python-dotenv telethon

Примеры запуска:
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --dry-run
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --yes
python bulk_site_pitch_telega.py --xlsx ip_fio_links.xlsx --yes --limit 3
"""

import asyncio
import argparse
import asyncio
import os
import random
import re
import sys
from dataclasses import dataclass
from getpass import getpass
from typing import Optional
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from telethon import TelegramClient, functions, types
from telethon.errors import (
    FloodWaitError,
    PhoneNumberInvalidError,
    SessionPasswordNeededError,
)


load_dotenv()

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
MODEL = "google/gemini-3-flash-preview"
SESSION_NAME = os.getenv("TELEGRAM_SESSION_NAME", "my_telegram_session")
PHONE_RE = re.compile(r"^\+7\d{10}$")


@dataclass
class LeadRow:
    excel_row: int
    phone: str
    site: str


class StepController:
    def __init__(self, auto_yes: bool) -> None:
        self.auto_yes = auto_yes

    def ask_continue(self, title: str) -> bool:
        if self.auto_yes:
            print(f"[AUTO] {title}: продолжаю без вопроса, потому что включён --yes.")
            return True

        print("\n" + "=" * 90)
        print(title)
        print("Введите:")
        print("  да / y / продолжаем  — продолжить")
        print("  нет / n / skip       — пропустить эту строку")
        print("  стоп / q / exit      — завершить программу")

        while True:
            answer = input("Продолжаем? > ").strip().lower()
            if answer in {"да", "д", "y", "yes", "продолжаем", "go", "ok", "ок"}:
                return True
            if answer in {"нет", "н", "n", "no", "skip", "пропустить"}:
                print("[USER] Строка пропущена по вашему решению.")
                return False
            if answer in {"стоп", "stop", "q", "quit", "exit", "выход"}:
                print("[USER] Остановка программы по вашему решению.")
                raise KeyboardInterrupt
            print("Не понял ответ. Напишите 'да', 'нет' или 'стоп'.")

    def ask_send(self) -> bool:
        if self.auto_yes:
            print("[AUTO] Отправляю без финального вопроса, потому что включён --yes.")
            return True

        print("\n" + "=" * 90)
        print("ФИНАЛЬНОЕ ПОДТВЕРЖДЕНИЕ ОТПРАВКИ")
        print("Чтобы реально отправить сообщение, введите строго: отправляем")
        print("Любой другой ответ НЕ отправит сообщение и программа перейдёт к следующей строке.")
        answer = input("Отправляем? > ").strip().lower()
        if answer == "отправляем":
            return True
        print("[USER] Сообщение не отправлено, строка пропущена.")
        return False


def log_block(title: str, text: str = "") -> None:
    print("\n" + "#" * 90)
    print(f"# {title}")
    print("#" * 90)
    if text:
        print(text)


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_phone(phone: str) -> Optional[str]:
    phone = normalize_cell(phone)
    phone = re.sub(r"[\s\-()]+", "", phone)
    if PHONE_RE.fullmatch(phone):
        return phone
    return None


def normalize_url(raw_url: str) -> str:
    url = normalize_cell(raw_url)
    if not url:
        return ""
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        url = "https://" + url
    return url


def read_leads_from_excel(xlsx_path: str, sheet_name: Optional[str] = None) -> list[LeadRow]:
    log_block("ЧТЕНИЕ EXCEL", f"Файл: {xlsx_path}")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel-файл не найден: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    print(f"Активный лист: {worksheet.title}")
    print(f"Всего строк: {worksheet.max_row}")
    print(f"Всего столбцов: {worksheet.max_column}")

    if worksheet.max_column < 2:
        raise RuntimeError("В Excel должно быть минимум два столбца, потому что нужны последние два: телефон и сайт.")

    phone_col = worksheet.max_column - 1
    site_col = worksheet.max_column

    print(f"Телефон будет взят из предпоследнего столбца: #{phone_col}")
    print(f"Сайт будет взят из последнего столбца: #{site_col}")

    leads: list[LeadRow] = []

    for row_idx in range(2, worksheet.max_row + 1):
        raw_phone = normalize_cell(worksheet.cell(row=row_idx, column=phone_col).value)
        raw_site = normalize_cell(worksheet.cell(row=row_idx, column=site_col).value)

        print("\n" + "-" * 90)
        print(f"Строка Excel #{row_idx}")
        print(f"Сырой телефон: {raw_phone!r}")
        print(f"Сырой сайт:    {raw_site!r}")

        if not raw_phone and not raw_site:
            print("[SKIP] Пропуск: оба поля пустые — нет ни телефона, ни сайта.")
            continue
        if not raw_phone:
            print("[SKIP] Пропуск: пустой телефон.")
            continue
        if not raw_site:
            print("[SKIP] Пропуск: пустой сайт.")
            continue

        phone = validate_phone(raw_phone)
        if not phone:
            print("[SKIP] Пропуск: телефон не в формате +79999999999.")
            print("       Требуется строго: плюс, цифра 7 и ещё 10 цифр. Пример: +79161234567")
            continue

        site = normalize_url(raw_site)
        parsed = urlparse(site)
        if not parsed.netloc:
            print("[SKIP] Пропуск: сайт выглядит некорректно после нормализации.")
            print(f"       Нормализованный сайт: {site!r}")
            continue

        print("[OK] Строка прошла первичную проверку.")
        print(f"     Телефон: {phone}")
        print(f"     Сайт:    {site}")
        leads.append(LeadRow(excel_row=row_idx, phone=phone, site=site))

    print("\n" + "=" * 90)
    print(f"Готово. Валидных строк: {len(leads)}")
    workbook.close()
    return leads


def fetch_site_text(url: str, max_chars: int = 8000) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; SiteAnalyzer/1.0)",
    }

    print(f"[SITE] Загружаю сайт: {url}")
    response = requests.get(url, headers=headers, timeout=15)
    response.raise_for_status()
    print(f"[SITE] HTTP статус: {response.status_code}")
    print(f"[SITE] Content-Type: {response.headers.get('Content-Type', 'не указан')}")

    soup = BeautifulSoup(response.text, "html.parser")

    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title and soup.title.string else ""

    meta_description = ""
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        meta_description = meta["content"].strip()

    text = soup.get_text(separator=" ")
    text = re.sub(r"\s+", " ", text).strip()

    result = f"""
URL: {url}
Title: {title}
Meta description: {meta_description}
Page text: {text}
""".strip()

    print(f"[SITE] Title: {title!r}")
    print(f"[SITE] Meta description: {meta_description!r}")
    print(f"[SITE] Длина очищенного текста до обрезки: {len(result)} символов")
    print(f"[SITE] В нейросеть уйдёт максимум: {max_chars} символов")

    return result[:max_chars]


def generate_pitch(url: str, site_text: str, found_page: int | None = None) -> str:
    if found_page:
        search_context = (
            f"Я нашёл этот сайт, когда копался в интернете, "
            f"и он попался мне только на {found_page}-й странице поисковика."
        )
    else:
        search_context = "Я нашёл этот сайт, когда копался в интернете."

    prompt = f"""
Ты пишешь короткое сообщение потенциальному клиенту от веб-разработчика.

Стиль:
- Пиши просто, по-человечески.
- Без пафоса.
- Без слов: "статусный", "экспертиза", "упаковать", "презентабельный", "визуальная форма", "достойный", "серьезные клиенты".
- Не используй канцелярит.
- Не пиши слишком вежливо и корпоративно.
- Тон: прямой, уверенный, но без оскорблений.
- Максимум 5-6 предложений.

Что нужно сказать:
1. Я нашёл сайт, когда копался в интернете.
2. Если указан номер страницы поиска, скажи, что сайт показывается только на этой странице.
3. Объясни простыми словами, что сайт выглядит плохо/устаревше/неудобно и из-за этого может отталкивать клиентов.
4. Скажи, что такая низкая позиция в поиске часто бывает из-за того, что сайт изначально плохо спроектирован: плохая структура, слабый текст, старый дизайн, мало понятных блоков.
5. Скажи, что я могу сделать новый сайт: красивый, понятный, современный и нормально подготовленный под поиск.
6. Ни слова не говори про цену.
7. Напиши, что меня зовут Владислав, я уже несколько лет разрабатываю сайты, занимаюсь SEO-анализом - тоесть слежу за тем, чтобы сайт увидели как можно больше целевых клиентов
8. Разрабатывал сайты для учебных заведений, адвокатских контор, интернет магазинов, если нужно, то могу прислать свои работы
9. Работаю по договору
Напиши уверенно, но честно: "сайт будет сделан так, чтобы он часто попадал на первые страницах поиска".
И объясни, что из за того, что он будет чаще попадать на первые страницы, то его сможет увидеть больше людей, поэтому у Вас будет больше клиентов
Еще можешь немного, пару слов совсем рассказать почему именно этот сайт плохой
НЕ ЗАБУДЬ ПОЗДОРОВАТЬСЯ! НО ПРОЩАТЬСЯ НЕ НАДО!
Контекст поиска:
{search_context}

Данные сайта:
{site_text}

Выдай только готовое сообщение клиенту, без пояснений.
""".strip()

    if not OPENROUTER_API_KEY:
        raise RuntimeError(
            "Не найден OPENROUTER_API_KEY. Добавь его в .env или экспортируй переменную окружения."
        )

    print("[AI] Отправляю запрос в OpenRouter.")
    print(f"[AI] Модель: {MODEL}")
    print(f"[AI] Длина промта: {len(prompt)} символов")

    response = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://example.com",
            "X-Title": "Site Pitch Generator",
        },
        json={
            "model": MODEL,
            "messages": [
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
            "temperature": 0.7,
            "max_tokens": 300,
        },
        timeout=60,
    )

    print(f"[AI] HTTP статус OpenRouter: {response.status_code}")

    if response.status_code != 200:
        raise RuntimeError(f"OpenRouter error {response.status_code}: {response.text}")

    data = response.json()
    message = data["choices"][0]["message"]["content"].strip()
    print(f"[AI] Ответ получен. Длина сообщения: {len(message)} символов")
    return message


async def get_or_create_telegram_client() -> TelegramClient:
    api_id_raw = os.getenv("TELEGRAM_API_ID")
    api_hash = os.getenv("TELEGRAM_API_HASH")
    my_phone = os.getenv("TELEGRAM_PHONE")

    if not api_id_raw or not api_hash or not my_phone:
        raise RuntimeError(
            "Не найдены TELEGRAM_API_ID, TELEGRAM_API_HASH или TELEGRAM_PHONE. "
            "Добавь их в .env. API ID и HASH берутся на https://my.telegram.org"
        )

    try:
        api_id = int(api_id_raw)
    except ValueError as exc:
        raise RuntimeError("TELEGRAM_API_ID должен быть числом.") from exc

    client = TelegramClient(SESSION_NAME, api_id, api_hash)
    print(f"[TG] Подключаюсь к Telegram. Session name: {SESSION_NAME}")
    await client.connect()

    if not await client.is_user_authorized():
        print("[TG] Сессия ещё не авторизована. Отправляю код входа.")
        await client.send_code_request(my_phone)
        code = input("Введите код из Telegram/SMS: ").strip()

        try:
            await client.sign_in(my_phone, code)
        except SessionPasswordNeededError:
            password = getpass("Включена 2FA. Введите пароль Telegram: ")
            await client.sign_in(password=password)

    print("[TG] Telegram подключён и авторизован.")
    return client


async def find_telegram_user_by_phone(client: TelegramClient, target_phone: str):
    print(f"[TG] Ищу пользователя по номеру: {target_phone}")
    print("[TG] Для поиска временно импортирую один контакт, как в telega.py.")

    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=target_phone,
        first_name="Temporary",
        last_name="Contact",
    )

    result = await client(functions.contacts.ImportContactsRequest([contact]))

    if not result.users:
        print("[TG] Пользователь не найден или недоступен по настройкам приватности.")
        return None

    user = result.users[0]
    print("[TG] Пользователь найден.")
    print(f"     id: {getattr(user, 'id', None)}")
    print(f"     username: {getattr(user, 'username', None)}")
    print(f"     first_name: {getattr(user, 'first_name', None)}")
    print(f"     last_name: {getattr(user, 'last_name', None)}")
    return user


async def cleanup_telegram_contact(client: TelegramClient, user) -> None:
    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
        print("[TG] Временный контакт удалён из контактов Telegram.")
    except Exception as exc:
        print("[TG][WARN] Не удалось удалить временный контакт.")
        print(f"          Тип ошибки: {type(exc).__name__}")
        print(f"          Текст ошибки: {exc}")


async def send_telegram_message(client: TelegramClient, user, message: str) -> None:
    print("[TG] Отправляю сообщение...")
    await client.send_message(user, message)
    print("[TG] Сообщение отправлено.")

async def process_leads(args) -> None:
    controller = StepController(auto_yes=args.yes)
    leads = read_leads_from_excel(args.xlsx, args.sheet)

    if args.start_row:
        leads = [lead for lead in leads if lead.excel_row >= args.start_row]
        print(f"[FILTER] После --start-row {args.start_row} осталось строк: {len(leads)}")

    if args.limit is not None:
        leads = leads[: args.limit]
        print(f"[FILTER] После --limit {args.limit} осталось строк: {len(leads)}")

    if not leads:
        print("Нет валидных строк для обработки. Завершаю работу.")
        return

    client: Optional[TelegramClient] = None

    try:
        for index, lead in enumerate(leads, start=1):
            log_block(
                f"СТРОКА {index}/{len(leads)} | EXCEL #{lead.excel_row}",
                f"Телефон: {lead.phone}\nСайт:    {lead.site}",
            )

            if not controller.ask_continue("Этап 1: данные строки прочитаны и показаны"):
                continue

            try:
                log_block("ЭТАП 2: ЗАГРУЗКА САЙТА")
                
                try:
                    site_text = fetch_site_text(lead.site, max_chars=args.max_site_chars)
                    print("[OK] Сайт успешно загружен.")
                    print(f"[SITE] Длина текста сайта для нейросети: {len(site_text)} символов")
                
                except requests.RequestException as exc:
                    print("[WARNING] Сайт не удалось загрузить, но строка НЕ будет пропущена.")
                    print(f"          Строка Excel: {lead.excel_row}")
                    print(f"          Сайт: {lead.site}")
                    print(f"          Тип ошибки: {type(exc).__name__}")
                    print(f"          Текст ошибки: {exc}")
                    print("          Вместо текста сайта передаю в нейросеть описание ошибки.")
                
                    site_text = f"""
                URL: {lead.site}
                Title:
                Meta description:
                Page text: Сайт не удалось открыть при автоматической проверке.
                
                Техническая ошибка:
                {type(exc).__name__}: {exc}
                
                Важно:
                Сайт может быть временно недоступен, неправильно настроен, не иметь DNS-записи,
                не открываться по HTTP/HTTPS или блокировать автоматические запросы.
                """.strip()
                
                log_block("ЭТАП 3: ЗАПРОС В НЕЙРОСЕТЬ")
                pitch = generate_pitch(lead.site, site_text, args.page)

                log_block("ОТВЕТ НЕЙРОСЕТИ", pitch)
                if not controller.ask_continue("Этап 3: ответ нейросети получен и показан"):
                    continue

                log_block("ЭТАП 4: TELEGRAM — ПОДКЛЮЧЕНИЕ И ПОИСК ПОЛЬЗОВАТЕЛЯ")
                if client is None:
                    client = await get_or_create_telegram_client()
                else:
                    print("[TG] Уже подключён к Telegram, использую существующее соединение.")

                user = await find_telegram_user_by_phone(client, lead.phone)
                if user is None:
                    print("[SKIP] Сообщение не будет отправлено, потому что пользователь не найден.")
                    continue

                if not controller.ask_continue("Этап 4: Telegram-пользователь найден и показан"):
                    await cleanup_telegram_contact(client, user)
                    continue

                log_block("ЭТАП 5: ТЕКСТ, КОТОРЫЙ БУДЕТ ОТПРАВЛЕН", pitch)

                if args.dry_run:
                    print("[DRY-RUN] Включён --dry-run, поэтому сообщение НЕ отправляется.")
                    await cleanup_telegram_contact(client, user)
                    continue

                message_was_sent = False
                
                if controller.ask_send():
                    await send_telegram_message(client, user, pitch)
                    message_was_sent = True
                else:
                    print("[SKIP] Отправка отменена пользователем.")
                
                await cleanup_telegram_contact(client, user)
                
                if message_was_sent and args.yes and args.delay > 0:
                    print(f"[DELAY] Жду {args.delay} секунд перед следующей отправкой...")
                    await asyncio.sleep(args.delay)

            except PhoneNumberInvalidError:
                print("[ERROR] Telegram считает номер телефона неверным.")
                print(f"        Строка Excel: {lead.excel_row}")
                print(f"        Телефон: {lead.phone}")
                print("        Перехожу к следующей строке.")
            except FloodWaitError as exc:
                print("[ERROR] Telegram временно ограничил запросы.")
                print(f"        Нужно подождать секунд: {exc.seconds}")
                print("        Программа останавливается, чтобы не усиливать ограничение.")
                break
            except Exception as exc:
                print("[ERROR] Непредвиденная ошибка на строке.")
                print(f"        Строка Excel: {lead.excel_row}")
                print(f"        Телефон: {lead.phone}")
                print(f"        Сайт: {lead.site}")
                print(f"        Тип ошибки: {type(exc).__name__}")
                print(f"        Текст ошибки: {exc}")
                print("        Перехожу к следующей строке.")

    finally:
        if client is not None:
            print("[TG] Отключаюсь от Telegram.")
            await client.disconnect()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Генерирует сообщения по сайтам из Excel и отправляет их в Telegram по номерам из Excel."
    )
    parser.add_argument(
        "--xlsx",
        default="ip_fio_links_11.xlsx",
        help="Путь к Excel-файлу. По умолчанию: ip_fio_links.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Название листа Excel. Если не указать, берётся активный лист.",
    )
    parser.add_argument(
        "--page",
        type=int,
        default=None,
        help="Страница поисковика, на которой сайт был найден. Передаётся в промт так же, как в site_pitch.py.",
    )
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Автоматический режим: не задавать вопросы 'продолжаем ли' и отправлять без финального подтверждения.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Проверочный режим: всё сделать, но НЕ отправлять сообщения.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Обработать только первые N валидных строк.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help="Начать обработку с указанного номера строки Excel.",
    )
    parser.add_argument(
        "--max-site-chars",
        type=int,
        default=8000,
        help="Сколько символов текста сайта максимум отправлять в нейросеть. По умолчанию 8000.",
    )
    parser.add_argument(
        "--delay",
        type=int,
        default=0,
        help="Пауза в секундах между отправками сообщений. Например: --delay 30"
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    print("\nbulk_site_pitch_telega.py запущен")
    print(f"Python: {sys.version}")
    print(f"Excel: {args.xlsx}")
    print(f"Интерактивный режим: {'выключен (--yes)' if args.yes else 'включён'}")
    print(f"Dry-run: {'да, сообщения не отправляются' if args.dry_run else 'нет'}")

    if args.yes and not args.dry_run:
        print("\n[WARN] Включён --yes без --dry-run: скрипт будет отправлять сообщения без вопросов.")
        print("       Используйте это только для контактов, которые дали согласие на такой эксперимент.")

    try:
        asyncio.run(process_leads(args))
    except KeyboardInterrupt:
        print("\nПрограмма остановлена пользователем.")


if __name__ == "__main__":
    main()
