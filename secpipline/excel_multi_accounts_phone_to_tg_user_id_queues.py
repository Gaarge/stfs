"""
excel_multi_accounts_phone_to_tg_user_id_queues.py

Что делает:
1. НЕ ходит на Rusprofile и НЕ делает curl.
2. Запускает несколько Telegram-аккаунтов поиска параллельно.
3. Каждый аккаунт читает свой Excel-файл.
4. Каждый аккаунт пишет найденные Telegram user_id/access_hash в свою JSONL-очередь.
5. Учитывает leads_queue_processed.txt, чтобы не ставить в очередь тех, кому уже писали.
6. Учитывает уже существующую очередь каждого аккаунта, чтобы при перезапуске не плодить дубли.
7. Сохраняет найденные user_id/access_hash/username обратно в Excel аккаунта.

Установка зависимостей:
pip install telethon python-dotenv openpyxl

.env общий для всех аккаунтов:
TELEGRAM_SEARCH_API_ID=123456
TELEGRAM_SEARCH_API_HASH=abcdef123456abcdef123456abcdef12
TELEGRAM_PHONE_SEARCH_DELAY_SECONDS=60

Файл accounts.json:
[
  {
    "name": "acc1",
    "phone": "+79990000001",
    "session_name": "tg_search_acc1",
    "xlsx": "fio_links_acc1.xlsx",
    "queue": "leads_queue_acc1.jsonl",
    "processed_state": "leads_queue_processed.txt",
    "start_row": 2,
    "limit": null
  },
  {
    "name": "acc2",
    "phone": "+79990000002",
    "session_name": "tg_search_acc2",
    "xlsx": "fio_links_acc2.xlsx",
    "queue": "leads_queue_acc2.jsonl",
    "processed_state": "leads_queue_processed.txt",
    "start_row": 2,
    "limit": null
  }
]

Пример запуска:
python excel_multi_accounts_phone_to_tg_user_id_queues.py --accounts-config accounts.json
python excel_multi_accounts_phone_to_tg_user_id_queues.py --accounts-config accounts.json --dry-run

Важно:
- Первый запуск новых Telegram-сессий попросит код входа для каждого аккаунта. Авторизация выполняется аккуратно,
  через общий lock, чтобы запросы кода/пароля не перемешивались в консоли.
- После авторизации все аккаунты начинают обработку Excel параллельно.
"""

import argparse
import asyncio
import json
import os
import random
import re
import sys
import time
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from telethon import TelegramClient, functions, types
from telethon.errors import FloodWaitError, PeerFloodError, SessionPasswordNeededError


load_dotenv()

DEFAULT_QUEUE_FILE = "leads_queue_user_ids.jsonl"
DEFAULT_PROCESSED_STATE_FILE = "leads_queue_processed.txt"
TELEGRAM_PHONE_SEARCH_DELAY_SECONDS = int(os.getenv("TELEGRAM_PHONE_SEARCH_DELAY_SECONDS", "60"))
CRITICAL_TELEGRAM_ERRORS = (FloodWaitError, PeerFloodError)
PHONE_RE = re.compile(r"^\+\d{10,15}$")

# Нужен, чтобы при первом запуске нескольких аккаунтов prompts ввода кода/2FA не шли одновременно.
AUTH_INPUT_LOCK = asyncio.Lock()


@dataclass
class AccountConfig:
    name: str
    phone: str
    session_name: str
    xlsx: Path
    queue: Path
    processed_state: Path
    sheet: Optional[str]
    start_row: int
    limit: Optional[int]
    phone_header: str
    site_header: str
    user_id_header: str
    access_hash_header: str
    username_header: str
    fallback_site: str
    api_id: int
    api_hash: str
    delay_seconds: int


@dataclass
class WorkerStats:
    name: str
    attempted_rows: int = 0
    added_rows: int = 0
    skipped_rows: int = 0
    stopped_by_telegram_limit: bool = False
    error: Optional[str] = None


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_phone(raw_phone: str) -> str:
    phone = normalize_cell(raw_phone)
    if not phone:
        return ""
    if phone.startswith("+"):
        cleaned = "+" + re.sub(r"\D", "", phone)
    else:
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("8"):
            cleaned = "+7" + digits[1:]
        elif len(digits) == 11 and digits.startswith("7"):
            cleaned = "+" + digits
        elif len(digits) == 10:
            cleaned = "+7" + digits
        else:
            cleaned = "+" + digits if digits else ""
    return cleaned if PHONE_RE.fullmatch(cleaned) else ""


def normalize_site(raw_site: str, fallback_site: str = "") -> str:
    site = normalize_cell(raw_site)
    if not site:
        return fallback_site
    if not re.match(r"^https?://", site, flags=re.IGNORECASE):
        site = "https://" + site
    return site


def normalize_processed_site(site: str) -> str:
    return normalize_cell(site).rstrip("/")


def safe_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip().replace("'", ""))
    except (TypeError, ValueError):
        return None


def legacy_queue_key(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_cell(site)}"


def legacy_queue_key_normalized(excel_row: int, phone: str, site: str) -> str:
    return f"{excel_row}|{normalize_cell(phone)}|{normalize_processed_site(site)}"


def recipient_keys_for_values(excel_row: int, phone: str, site: str, user_id=None) -> set[str]:
    phone = normalize_cell(phone)
    site = normalize_cell(site)
    keys = {
        legacy_queue_key(excel_row, phone, site),
        legacy_queue_key_normalized(excel_row, phone, site),
        f"phone:{phone}",
    }
    parsed_user_id = safe_int(user_id)
    if parsed_user_id:
        keys.add(f"user_id:{parsed_user_id}")
    return {key for key in keys if key and key != "phone:"}


def load_processed_keys(path: Path) -> set[str]:
    """Читает и старый формат excel_row|phone|site, и новый JSONL-формат."""
    if not path.exists():
        return set()

    processed: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        processed.add(line)

        if line.startswith("{"):
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = safe_int(data.get("excel_row")) or 0
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id") or data.get("telegram_user_id")
            processed.update(recipient_keys_for_values(excel_row, phone, site, user_id))
        else:
            parts = line.split("|", 2)
            if len(parts) == 3:
                row_raw, phone, site = parts
                excel_row = safe_int(row_raw) or 0
                processed.update(recipient_keys_for_values(excel_row, phone, site))
    return processed


def load_queued_keys(path: Path) -> set[str]:
    """Читает текущую очередь, чтобы не добавлять дубль при перезапуске."""
    if not path.exists():
        return set()

    keys: set[str] = set()
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or not line.startswith("{"):
                continue
            try:
                data = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = safe_int(data.get("excel_row")) or 0
            phone = normalize_cell(data.get("phone"))
            site = normalize_cell(data.get("site"))
            user_id = data.get("user_id") or data.get("telegram_user_id")
            keys.update(recipient_keys_for_values(excel_row, phone, site, user_id))
    return keys


def find_col_by_header(ws, header_name: str) -> Optional[int]:
    wanted = header_name.strip().lower()
    for col in range(1, ws.max_column + 1):
        value = normalize_cell(ws.cell(row=1, column=col).value).lower()
        if value == wanted:
            return col
    return None


def get_or_create_column(ws, title: str) -> int:
    existing = find_col_by_header(ws, title)
    if existing:
        return existing
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = title
    return col


def detect_last_data_row(ws, relevant_cols: list[int]) -> int:
    """Excel иногда считает max_row как 1048576 из-за форматирования. Ищем последнюю реально заполненную строку."""
    for row in range(ws.max_row, 1, -1):
        for col in relevant_cols:
            if normalize_cell(ws.cell(row=row, column=col).value):
                return row
    return 1


def append_lead_to_queue(
    queue_path: Path,
    account_name: str,
    excel_path: Path,
    excel_row: int,
    phone: str,
    site: str,
    user_id: int,
    access_hash: int,
    username: str = "",
    first_name: str = "",
    last_name: str = "",
) -> None:
    queue_path.parent.mkdir(parents=True, exist_ok=True)
    record = {
        "account": account_name,
        "excel_file": str(excel_path),
        "excel_row": excel_row,
        "phone": phone,
        "site": site,
        "user_id": int(user_id),
        "access_hash": int(access_hash),
        "username": username or None,
        "first_name": first_name or None,
        "last_name": last_name or None,
        "legacy_key": legacy_queue_key(excel_row, phone, site),
        "source": "excel_multi_accounts_phone_to_tg_user_id_queues.py",
        "queued_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }

    with queue_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
        f.flush()

    log(account_name, f"[QUEUE] Добавлен в очередь: user_id={user_id}, phone={phone}")


def log(account_name: str, message: str) -> None:
    print(f"[{account_name}] {message}", flush=True)


def read_accounts_config(path: Path, args: argparse.Namespace) -> list[AccountConfig]:
    if not path.exists():
        raise FileNotFoundError(f"Файл accounts config не найден: {path.resolve()}")

    try:
        raw_accounts = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Не удалось прочитать JSON config {path}: {exc}") from exc

    if not isinstance(raw_accounts, list) or not raw_accounts:
        raise RuntimeError("accounts config должен быть непустым JSON-массивом аккаунтов.")

    default_api_id_raw = os.getenv("TELEGRAM_SEARCH_API_ID") or os.getenv("TELEGRAM_API_ID")
    default_api_hash = os.getenv("TELEGRAM_SEARCH_API_HASH") or os.getenv("TELEGRAM_API_HASH")

    accounts: list[AccountConfig] = []
    names: set[str] = set()
    sessions: set[str] = set()

    for idx, item in enumerate(raw_accounts, start=1):
        if not isinstance(item, dict):
            raise RuntimeError(f"Аккаунт #{idx} должен быть JSON-объектом.")

        name = normalize_cell(item.get("name")) or f"account_{idx}"
        if name in names:
            raise RuntimeError(f"Дублируется name аккаунта: {name}")
        names.add(name)

        phone = normalize_phone(item.get("phone"))
        if not phone:
            raise RuntimeError(f"У аккаунта {name} не указан корректный phone, пример: +79990000001")

        session_name = normalize_cell(item.get("session_name")) or f"tg_search_{name}"
        if session_name in sessions:
            raise RuntimeError(f"Дублируется session_name: {session_name}")
        sessions.add(session_name)

        api_id_raw = normalize_cell(item.get("api_id")) or default_api_id_raw
        api_hash = normalize_cell(item.get("api_hash")) or default_api_hash
        if not api_id_raw or not api_hash:
            raise RuntimeError(
                f"Для аккаунта {name} не найдены api_id/api_hash. "
                "Укажите их в accounts.json или в .env как TELEGRAM_SEARCH_API_ID/TELEGRAM_SEARCH_API_HASH."
            )
        try:
            api_id = int(api_id_raw)
        except ValueError as exc:
            raise RuntimeError(f"api_id у аккаунта {name} должен быть числом.") from exc

        xlsx = Path(normalize_cell(item.get("xlsx")))
        if not normalize_cell(item.get("xlsx")):
            raise RuntimeError(f"У аккаунта {name} не указан xlsx.")
        if not xlsx.exists():
            raise FileNotFoundError(f"Excel-файл аккаунта {name} не найден: {xlsx.resolve()}")

        queue = Path(normalize_cell(item.get("queue")) or f"leads_queue_{name}.jsonl")
        processed_state = Path(normalize_cell(item.get("processed_state")) or DEFAULT_PROCESSED_STATE_FILE)

        accounts.append(
            AccountConfig(
                name=name,
                phone=phone,
                session_name=session_name,
                xlsx=xlsx,
                queue=queue,
                processed_state=processed_state,
                sheet=item.get("sheet", args.sheet),
                start_row=int(item.get("start_row", args.start_row)),
                limit=item.get("limit", args.limit),
                phone_header=normalize_cell(item.get("phone_header")) or args.phone_header,
                site_header=normalize_cell(item.get("site_header")) or args.site_header,
                user_id_header=normalize_cell(item.get("user_id_header")) or args.user_id_header,
                access_hash_header=normalize_cell(item.get("access_hash_header")) or args.access_hash_header,
                username_header=normalize_cell(item.get("username_header")) or args.username_header,
                fallback_site=normalize_cell(item.get("fallback_site")) if "fallback_site" in item else args.fallback_site,
                api_id=api_id,
                api_hash=api_hash,
                delay_seconds=int(item.get("delay_seconds", args.delay_seconds)),
            )
        )

    return accounts


async def get_or_create_search_telegram_client(account: AccountConfig) -> TelegramClient:
    client = TelegramClient(account.session_name, account.api_id, account.api_hash)
    log(account.name, f"[TG-SEARCH] Подключаюсь к Telegram. Session: {account.session_name}")
    await client.connect()

    if not await client.is_user_authorized():
        async with AUTH_INPUT_LOCK:
            # Пока ждали lock, другой код не мог авторизовать именно эту сессию, но перепроверка не мешает.
            if not await client.is_user_authorized():
                log(account.name, "[TG-SEARCH] Сессия ещё не авторизована. Отправляю код входа.")
                await client.send_code_request(account.phone)
                code = input(f"[{account.name}] Введите код из Telegram/SMS для {account.phone}: ").strip()
                try:
                    await client.sign_in(account.phone, code)
                except SessionPasswordNeededError:
                    password = getpass(f"[{account.name}] Включена 2FA. Введите пароль Telegram: ")
                    await client.sign_in(password=password)

    log(account.name, "[TG-SEARCH] Telegram-аккаунт подключён и авторизован.")
    return client


async def find_telegram_user_by_phone(account_name: str, client: TelegramClient, target_phone: str):
    log(account_name, f"[TG-SEARCH] Ищу user_id по номеру: {target_phone}")

    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=target_phone,
        first_name="Temporary",
        last_name="Contact",
    )

    result = await client(functions.contacts.ImportContactsRequest([contact]))
    if not result.users:
        log(account_name, "[TG-SEARCH] Пользователь не найден или скрыт настройками приватности.")
        return None

    user = result.users[0]
    log(
        account_name,
        "[TG-SEARCH] Найден: "
        f"id={getattr(user, 'id', None)}, "
        f"access_hash={getattr(user, 'access_hash', None)}, "
        f"username={getattr(user, 'username', None)}",
    )

    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
        log(account_name, "[TG-SEARCH] Временный контакт удалён.")
    except Exception as exc:
        log(account_name, f"[TG-SEARCH][WARN] Не удалось удалить временный контакт: {type(exc).__name__}: {exc}")

    return user


async def maybe_wait_before_tg_search(account: AccountConfig, last_search_at: float) -> float:
    if last_search_at > 0:
        elapsed = time.monotonic() - last_search_at
        remaining = account.delay_seconds - elapsed
        if remaining > 0:
            log(account.name, f"[TG-SEARCH][DELAY] Жду {remaining:.1f} секунд перед следующим поиском по номеру...")
            await asyncio.sleep(remaining)
    return time.monotonic()


async def process_account(account: AccountConfig, dry_run: bool) -> WorkerStats:
    stats = WorkerStats(name=account.name)
    tg_client: Optional[TelegramClient] = None
    wb = None

    try:
        processed_keys = load_processed_keys(account.processed_state)
        queued_keys = load_queued_keys(account.queue)

        log(account.name, "Запущен worker")
        log(account.name, f"Python: {sys.version.split()[0]}")
        log(account.name, f"Excel: {account.xlsx.resolve()}")
        log(account.name, f"Очередь: {account.queue.resolve()}")
        log(account.name, f"Processed: {account.processed_state.resolve()}")
        log(account.name, f"Dry-run: {'да' if dry_run else 'нет'}")
        log(account.name, f"[STATE] Уже отправленных ключей загружено: {len(processed_keys)}")
        log(account.name, f"[QUEUE] Уже стоящих в очереди ключей загружено: {len(queued_keys)}")

        wb = load_workbook(account.xlsx)
        ws = wb[account.sheet] if account.sheet else wb.active

        phone_col = find_col_by_header(ws, account.phone_header)
        site_col = find_col_by_header(ws, account.site_header)
        if not phone_col:
            raise RuntimeError(f"Не нашёл колонку с заголовком {account.phone_header!r} в первой строке Excel.")
        if not site_col:
            raise RuntimeError(f"Не нашёл колонку с заголовком {account.site_header!r} в первой строке Excel.")

        user_id_col = get_or_create_column(ws, account.user_id_header)
        access_hash_col = get_or_create_column(ws, account.access_hash_header)
        username_col = get_or_create_column(ws, account.username_header)

        last_row = detect_last_data_row(ws, [phone_col, site_col, user_id_col, access_hash_col])
        log(account.name, f"[EXCEL] Лист: {ws.title}")
        log(account.name, f"[EXCEL] Колонка телефона: #{phone_col}")
        log(account.name, f"[EXCEL] Колонка сайта: #{site_col}")
        log(account.name, f"[EXCEL] Колонка user_id: #{user_id_col}")
        log(account.name, f"[EXCEL] Колонка access_hash: #{access_hash_col}")
        log(account.name, f"[EXCEL] Последняя строка с данными: {last_row}")

        if not dry_run:
            wb.save(account.xlsx)

        tg_client = await get_or_create_search_telegram_client(account)
        last_tg_search_at = 0.0

        for row in range(account.start_row, last_row + 1):
            raw_phone = ws.cell(row=row, column=phone_col).value
            raw_site = ws.cell(row=row, column=site_col).value

            phone = normalize_phone(raw_phone)
            site = normalize_site(raw_site, fallback_site=account.fallback_site)

            if not phone:
                continue
            if not site:
                log(account.name, f"[{row}] [SKIP] Есть телефон, но нет сайта, а fallback_site пустой: {raw_phone!r}")
                stats.skipped_rows += 1
                continue

            stats.attempted_rows += 1
            if account.limit is not None and stats.attempted_rows > int(account.limit):
                log(account.name, f"[LIMIT] Достигнут limit {account.limit}. Останавливаю worker.")
                break

            existing_user_id = safe_int(ws.cell(row=row, column=user_id_col).value)
            existing_access_hash = safe_int(ws.cell(row=row, column=access_hash_col).value)
            existing_username = normalize_cell(ws.cell(row=row, column=username_col).value)

            log(account.name, f"[{row}] phone={phone!r}, site={site!r}")

            base_keys = recipient_keys_for_values(row, phone, site, existing_user_id)
            if not base_keys.isdisjoint(processed_keys):
                log(account.name, "    [SKIP] Уже есть в processed-state по старой/новой схеме. В очередь не добавляю.")
                stats.skipped_rows += 1
                continue
            if not base_keys.isdisjoint(queued_keys):
                log(account.name, "    [SKIP] Уже есть в очереди этого аккаунта. Дубль не добавляю.")
                stats.skipped_rows += 1
                continue

            if existing_user_id and existing_access_hash:
                log(account.name, f"    [EXCEL] Уже есть user_id/access_hash в Excel: {existing_user_id}")
                if not dry_run:
                    append_lead_to_queue(
                        queue_path=account.queue,
                        account_name=account.name,
                        excel_path=account.xlsx,
                        excel_row=row,
                        phone=phone,
                        site=site,
                        user_id=existing_user_id,
                        access_hash=existing_access_hash,
                        username=existing_username,
                    )
                    queued_keys.update(recipient_keys_for_values(row, phone, site, existing_user_id))
                else:
                    log(account.name, "    [DRY-RUN] В очередь не пишу.")
                stats.added_rows += 1
                continue

            last_tg_search_at = await maybe_wait_before_tg_search(account, last_tg_search_at)
            user = await find_telegram_user_by_phone(account.name, tg_client, phone)
            last_tg_search_at = time.monotonic()

            if user is None:
                log(account.name, "    [QUEUE] В очередь не добавляю: Telegram user_id не найден.")
                stats.skipped_rows += 1
                continue

            user_id = getattr(user, "id", None)
            access_hash = getattr(user, "access_hash", None)
            username = getattr(user, "username", None)
            first_name = getattr(user, "first_name", None)
            last_name = getattr(user, "last_name", None)

            if not user_id or not access_hash:
                log(account.name, "    [QUEUE] В очередь не добавляю: Telegram вернул пользователя без user_id/access_hash.")
                stats.skipped_rows += 1
                continue

            if not dry_run:
                ws.cell(row=row, column=user_id_col).value = str(user_id)
                ws.cell(row=row, column=access_hash_col).value = str(access_hash)
                ws.cell(row=row, column=username_col).value = username
                wb.save(account.xlsx)

                append_lead_to_queue(
                    queue_path=account.queue,
                    account_name=account.name,
                    excel_path=account.xlsx,
                    excel_row=row,
                    phone=phone,
                    site=site,
                    user_id=int(user_id),
                    access_hash=int(access_hash),
                    username=username or "",
                    first_name=first_name or "",
                    last_name=last_name or "",
                )
                queued_keys.update(recipient_keys_for_values(row, phone, site, user_id))
            else:
                log(account.name, f"    [DRY-RUN] Нашёл бы и записал бы: user_id={user_id}, access_hash={access_hash}")

            stats.added_rows += 1

    except CRITICAL_TELEGRAM_ERRORS as exc:
        stats.stopped_by_telegram_limit = True
        stats.error = f"{type(exc).__name__}: {exc}"
        log(account.name, f"[TG-SEARCH][CRITICAL] Telegram ограничил поиск: {stats.error}")
        if isinstance(exc, FloodWaitError):
            log(account.name, f"[TG-SEARCH][CRITICAL] Telegram просит подождать секунд: {exc.seconds}")
        log(account.name, "Останавливаю только этот worker, остальные аккаунты продолжают работу.")
    except Exception as exc:
        stats.error = f"{type(exc).__name__}: {exc}"
        log(account.name, f"[ERROR] Worker остановлен: {stats.error}")
    finally:
        if wb is not None:
            if not dry_run:
                wb.save(account.xlsx)
            wb.close()
        if tg_client is not None:
            await tg_client.disconnect()
            log(account.name, "[TG-SEARCH] Отключился от Telegram.")

    log(account.name, "Готово.")
    log(account.name, f"Строк с телефонами просмотрено: {stats.attempted_rows}")
    log(account.name, f"Добавлено/готово к добавлению в очередь: {stats.added_rows}")
    log(account.name, f"Пропущено: {stats.skipped_rows}")
    return stats


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Параллельно запускает несколько Telegram-аккаунтов поиска. Каждый читает свой Excel и пишет свою очередь."
    )
    parser.add_argument("--accounts-config", required=True, help="Путь к JSON-файлу со списком аккаунтов.")
    parser.add_argument("--sheet", default=None, help="Общее название листа, если не задано внутри аккаунта.")
    parser.add_argument("--start-row", type=int, default=2, help="Общая стартовая строка, если не задана внутри аккаунта.")
    parser.add_argument("--limit", type=int, default=None, help="Общий лимит строк на аккаунт, если не задан внутри аккаунта.")
    parser.add_argument("--phone-header", default="Телефон", help="Заголовок колонки с телефоном.")
    parser.add_argument("--site-header", default="Сайт", help="Заголовок колонки с сайтом.")
    parser.add_argument("--user-id-header", default="Telegram user_id", help="Колонка для user_id.")
    parser.add_argument("--access-hash-header", default="Telegram access_hash", help="Колонка для access_hash.")
    parser.add_argument("--username-header", default="Telegram username", help="Колонка для username.")
    parser.add_argument("--fallback-site", default="https://example.com", help="Что ставить в site, если сайт в Excel пустой. Пустая строка = пропускать без сайта.")
    parser.add_argument("--delay-seconds", type=int, default=TELEGRAM_PHONE_SEARCH_DELAY_SECONDS, help="Пауза между поисками внутри одного аккаунта.")
    parser.add_argument("--dry-run", action="store_true", help="Ничего не писать в очередь и Excel, только показать что было бы сделано.")
    return parser


async def async_main() -> None:
    args = build_arg_parser().parse_args()
    accounts = read_accounts_config(Path(args.accounts_config), args)

    print("\nexcel_multi_accounts_phone_to_tg_user_id_queues.py запущен", flush=True)
    print(f"Аккаунтов в config: {len(accounts)}", flush=True)
    print(f"Dry-run: {'да' if args.dry_run else 'нет'}", flush=True)

    results = await asyncio.gather(
        *(process_account(account, dry_run=args.dry_run) for account in accounts),
        return_exceptions=False,
    )

    print("\nИтог по всем аккаунтам:", flush=True)
    total_attempted = 0
    total_added = 0
    total_skipped = 0
    for stats in results:
        total_attempted += stats.attempted_rows
        total_added += stats.added_rows
        total_skipped += stats.skipped_rows
        status = "OK"
        if stats.stopped_by_telegram_limit:
            status = "TELEGRAM_LIMIT"
        elif stats.error:
            status = "ERROR"
        print(
            f"- {stats.name}: status={status}, viewed={stats.attempted_rows}, "
            f"added={stats.added_rows}, skipped={stats.skipped_rows}"
            + (f", error={stats.error}" if stats.error else ""),
            flush=True,
        )

    print(f"Всего просмотрено: {total_attempted}", flush=True)
    print(f"Всего добавлено/готово к добавлению: {total_added}", flush=True)
    print(f"Всего пропущено: {total_skipped}", flush=True)


def main() -> None:
    asyncio.run(async_main())


if __name__ == "__main__":
    main()
