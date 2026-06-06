import argparse
import random
import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency. Install requirements first:\n"
        "  python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
BASE_URL = "https://www.rusprofile.ru"

DEFAULT_XLSX = BASE_DIR / "rusprofile_users.xlsx"
DEFAULT_PHONE_HEADER = "Телефон"
DEFAULT_SITE_HEADER = "Сайт"
LINK_HEADERS = ("Ссылка", "ссылка", "link", "url")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
]


class CaptchaDetected(RuntimeError):
    pass


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value) -> str:
    return normalize_cell(value).lower()


def find_col_by_headers(ws, headers: tuple[str, ...]) -> int | None:
    wanted = {normalize_header(header) for header in headers}
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=1, column=col).value) in wanted:
            return col
    return None


def get_or_create_col(ws, header: str) -> int:
    existing = find_col_by_headers(ws, (header,))
    if existing:
        return existing
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header
    return col


def make_rusprofile_url(raw_link: str) -> str:
    link = normalize_cell(raw_link)
    if not link:
        return ""
    if link.startswith("http://") or link.startswith("https://"):
        return link
    return urljoin(BASE_URL + "/", link)


def clean_phone(raw_phone: str) -> str:
    phone = normalize_cell(raw_phone)
    if not phone:
        return ""
    if phone.startswith("+"):
        digits = re.sub(r"\D", "", phone)
        return f"+{digits}" if digits else ""

    digits = re.sub(r"\D", "", phone)
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    return "+" + digits


def is_captcha_page(html: str) -> bool:
    markers = (
        "captcha-section",
        "g-recaptcha",
        "captcha-validate",
        "Активность с вашего IP-адреса была распознана как автоматическая",
        "Я не робот",
    )
    return any(marker in html for marker in markers)


def fetch_html(url: str, timeout: int) -> str:
    response = requests.get(
        url,
        headers={"User-Agent": random.choice(USER_AGENTS)},
        timeout=timeout,
    )
    response.raise_for_status()
    return response.text


def extract_phone_and_site(html: str) -> tuple[str, str]:
    if is_captcha_page(html):
        raise CaptchaDetected("Rusprofile returned captcha page")

    soup = BeautifulSoup(html, "html.parser")
    phone = ""
    site = ""

    phone_tag = soup.select_one('a[href^="tel:"]')
    if phone_tag:
        phone = phone_tag.get("href", "").replace("tel:", "").strip()

    if not phone:
        match = re.search(r'href=["\']tel:([^"\']+)', html, flags=re.IGNORECASE)
        if match:
            phone = match.group(1)

    site_tag = soup.select_one('.company-info__contact.site a[href^="http"]')
    if site_tag:
        site = site_tag.get("href", "").strip()

    if not site:
        for contact in soup.select(".company-info__contact"):
            classes = contact.get("class") or []
            text = contact.get_text(" ", strip=True)
            if "site" not in classes and "Сайт" not in text:
                continue
            link = contact.select_one('a[href^="http"]')
            if link:
                site = link.get("href", "").strip()
                break

    if not site:
        match = re.search(
            r'class=["\'][^"\']*company-info__contact[^"\']*site[^"\']*["\'][\s\S]*?href=["\'](https?://[^"\']+)',
            html,
            flags=re.IGNORECASE,
        )
        if match:
            site = match.group(1).strip()

    return clean_phone(phone), site


def detect_last_row(ws, columns: list[int]) -> int:
    for row in range(ws.max_row, 1, -1):
        if any(normalize_cell(ws.cell(row=row, column=col).value) for col in columns):
            return row
    return 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fill phone/site columns in Excel from Rusprofile links."
    )
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Excel file path.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to active sheet.")
    parser.add_argument("--start-row", type=int, default=2, help="First data row.")
    parser.add_argument("--limit", type=int, default=None, help="Max rows to process.")
    parser.add_argument("--force", action="store_true", help="Parse even when phone and site already exist.")
    parser.add_argument("--dry-run", action="store_true", help="Do not write to Excel.")
    parser.add_argument("--delay-min", type=float, default=4.0, help="Min delay between requests.")
    parser.add_argument("--delay-max", type=float, default=12.0, help="Max delay between requests.")
    parser.add_argument("--timeout", type=int, default=40, help="HTTP timeout seconds.")
    parser.add_argument("--captcha-sleep", type=int, default=1800, help="Sleep seconds after captcha.")
    parser.add_argument("--captcha-retries", type=int, default=1, help="Retries for the same row after captcha.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path.resolve()}")

    wb = load_workbook(xlsx_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    link_col = find_col_by_headers(ws, LINK_HEADERS)
    if not link_col:
        raise SystemExit("Could not find link column. Expected header: Ссылка")

    phone_col = get_or_create_col(ws, DEFAULT_PHONE_HEADER)
    site_col = get_or_create_col(ws, DEFAULT_SITE_HEADER)
    last_row = detect_last_row(ws, [link_col, phone_col, site_col])

    print(f"Excel: {xlsx_path.resolve()}")
    print(f"Sheet: {ws.title}")
    print(f"Link column: #{link_col}")
    print(f"Phone column: #{phone_col}")
    print(f"Site column: #{site_col}")
    print(f"Last data row: {last_row}")
    print(f"Dry-run: {'yes' if args.dry_run else 'no'}")

    processed = 0
    changed = 0

    try:
        for row in range(args.start_row, last_row + 1):
            if args.limit is not None and processed >= args.limit:
                print(f"Limit reached: {args.limit}")
                break

            raw_link = ws.cell(row=row, column=link_col).value
            url = make_rusprofile_url(raw_link)
            if not url:
                continue

            existing_phone = normalize_cell(ws.cell(row=row, column=phone_col).value)
            existing_site = normalize_cell(ws.cell(row=row, column=site_col).value)
            if existing_phone and existing_site and not args.force:
                continue

            processed += 1
            print(f"\n[{row}] {url}")

            captcha_attempts = 0
            while True:
                try:
                    html = fetch_html(url, timeout=args.timeout)
                    phone, site = extract_phone_and_site(html)
                    break
                except CaptchaDetected as exc:
                    captcha_attempts += 1
                    print(f"  CAPTCHA: {exc}")
                    if captcha_attempts > args.captcha_retries:
                        print("  Stop on repeated captcha. Run later with --start-row", row)
                        raise
                    print(f"  Sleeping {args.captcha_sleep} seconds before retry...")
                    time.sleep(args.captcha_sleep)

            final_phone = phone or existing_phone
            final_site = site or existing_site

            print(f"  phone: {final_phone or '-'}")
            print(f"  site:  {final_site or '-'}")

            if not args.dry_run:
                ws.cell(row=row, column=phone_col).value = final_phone
                ws.cell(row=row, column=site_col).value = final_site
                wb.save(xlsx_path)
            changed += 1

            delay = random.uniform(args.delay_min, args.delay_max)
            print(f"  delay: {delay:.1f}s")
            time.sleep(delay)
    except KeyboardInterrupt:
        print("\nInterrupted by user.")
    finally:
        if not args.dry_run:
            wb.save(xlsx_path)
        wb.close()

    print("\nDone.")
    print(f"Rows parsed: {processed}")
    print(f"Rows updated/kept: {changed}")


if __name__ == "__main__":
    sys.exit(main())
