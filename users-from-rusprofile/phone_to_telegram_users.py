import argparse
import asyncio
import csv
import os
import random
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from getpass import getpass
from pathlib import Path

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    from openpyxl import load_workbook
    from telethon import TelegramClient, errors, functions, types
except ImportError as exc:
    raise SystemExit(
        "Missing dependency. Install requirements first:\n"
        "  python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR / "rusprofile_users.xlsx"
DEFAULT_OUTPUT = BASE_DIR / "telegram_users_from_phones.csv"
DEFAULT_API_ID = "34825825"
DEFAULT_API_HASH = "60176f7ad0bcd77e63d4a64ca8d50a38"

PHONE_RE = re.compile(r"^\+\d{10,15}$")


def load_env_files() -> list[Path]:
    if not load_dotenv:
        return []

    candidates = [
        BASE_DIR / ".env",
        BASE_DIR.parent / ".env",
        BASE_DIR.parent.parent / ".env",
        Path.cwd() / ".env",
    ]
    loaded = []
    seen = set()

    for path in candidates:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists() and load_dotenv(resolved, override=False):
            loaded.append(resolved)

    return loaded


LOADED_ENV_FILES = load_env_files()


@dataclass
class AccountConfig:
    account: str
    api_id: int
    api_hash: str
    phone: str
    session_path: Path


def normalize_account_name(account: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", account.strip())
    return cleaned or "default"


def env_for_account(account: str, key: str) -> str:
    prefix = normalize_account_name(account).upper()
    candidates = [
        f"TG_{prefix}_{key}",
        f"TELEGRAM_{prefix}_{key}",
    ]
    if account.lower() in {"search", "sender", "comments"}:
        candidates.extend(
            [
                f"TELEGRAM_{account.upper()}_{key}",
                f"TG_{account.upper()}_{key}",
            ]
        )
    candidates.extend([f"TG_{key}", f"TELEGRAM_{key}"])
    for name in candidates:
        value = os.getenv(name)
        if value:
            return value
    return ""


def resolve_account(account: str) -> AccountConfig:
    api_id_raw = env_for_account(account, "API_ID") or DEFAULT_API_ID
    api_hash = env_for_account(account, "API_HASH") or DEFAULT_API_HASH
    phone = env_for_account(account, "PHONE")
    session = env_for_account(account, "SESSION") or env_for_account(account, "SESSION_NAME")

    try:
        api_id = int(api_id_raw)
    except ValueError as exc:
        raise SystemExit("Telegram API_ID must be a number.") from exc

    if session:
        session_path = Path(session)
        if not session_path.is_absolute():
            session_path = BASE_DIR / session_path
    else:
        session_path = BASE_DIR / "sessions" / f"{normalize_account_name(account)}.session"

    session_path.parent.mkdir(parents=True, exist_ok=True)
    return AccountConfig(account, api_id, api_hash, phone, session_path)


def print_terminal_qr(url: str) -> None:
    qrencode = shutil.which("qrencode")
    if qrencode:
        subprocess.run([qrencode, "-t", "ANSIUTF8", url], check=False)
        return
    print("QR generator is not installed. Open this link from an already logged-in Telegram app:")
    print(url)


async def login_with_qr(client: TelegramClient) -> None:
    print("Starting QR login.")
    print("Open Telegram: Settings -> Devices -> Link Desktop Device.")
    qr_login = await client.qr_login()
    print_terminal_qr(qr_login.url)
    try:
        user = await qr_login.wait(timeout=120)
    except asyncio.TimeoutError as exc:
        raise SystemExit("QR login timed out.") from exc
    except errors.SessionPasswordNeededError:
        password = getpass("Enter Telegram 2FA password: ")
        await client.sign_in(password=password)
        user = await client.get_me()
    print(f"Logged in as @{user.username or user.id}")


async def ensure_authorized(client: TelegramClient, account: AccountConfig) -> None:
    if await client.is_user_authorized():
        me = await client.get_me()
        print(f"[TG] Account '{account.account}' authorized as @{me.username or me.id}")
        return

    method = input("Type 'qr' for QR login, or press Enter for phone-code login: ").strip().lower()
    if method == "qr":
        await login_with_qr(client)
        return

    if not account.phone:
        account.phone = input(f"Enter Telegram phone for account '{account.account}' (+79991234567): ").strip()
    if not account.phone:
        raise SystemExit("Telegram phone is empty.")

    try:
        sent = await client.send_code_request(account.phone)
    except errors.PhoneNumberBannedError as exc:
        raise SystemExit("Telegram says this phone number is banned.") from exc
    except errors.PhoneNumberInvalidError as exc:
        raise SystemExit("Telegram says this phone number is invalid.") from exc
    except errors.FloodWaitError as exc:
        raise SystemExit(f"Telegram rate-limited login. Wait {exc.seconds} seconds.") from exc

    print(f"Code requested. Delivery type: {type(sent.type).__name__}")
    code = input("Enter Telegram login code: ").strip().replace(" ", "")
    try:
        await client.sign_in(phone=account.phone, code=code, phone_code_hash=sent.phone_code_hash)
    except errors.SessionPasswordNeededError:
        password = getpass("Enter Telegram 2FA password: ")
        await client.sign_in(password=password)


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_phone(raw_phone: str) -> str:
    phone = normalize_cell(raw_phone)
    if not phone:
        return ""
    if phone.startswith("+"):
        cleaned = "+" + re.sub(r"\D", "", phone)
    else:
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("8"):
            cleaned = "+7" + digits[1:]
        elif len(digits) == 11 and digits.startswith("7"):
            cleaned = "+" + digits
        elif len(digits) == 10:
            cleaned = "+7" + digits
        else:
            cleaned = "+" + digits if digits else ""
    return cleaned if PHONE_RE.fullmatch(cleaned) else ""


def find_col(ws, header: str) -> int | None:
    wanted = header.strip().lower()
    for col in range(1, ws.max_column + 1):
        if normalize_cell(ws.cell(row=1, column=col).value).lower() == wanted:
            return col
    return None


def get_or_create_col(ws, header: str) -> int:
    existing = find_col(ws, header)
    if existing:
        return existing
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header
    return col


def detect_last_row(ws, cols: list[int]) -> int:
    for row in range(ws.max_row, 1, -1):
        if any(normalize_cell(ws.cell(row=row, column=col).value) for col in cols):
            return row
    return 1


async def find_user_by_phone(client: TelegramClient, phone: str):
    contact = types.InputPhoneContact(
        client_id=random.randrange(1, 10_000_000),
        phone=phone,
        first_name="Temporary",
        last_name="Contact",
    )
    result = await client(functions.contacts.ImportContactsRequest([contact]))
    if not result.users:
        return None
    user = result.users[0]
    try:
        await client(functions.contacts.DeleteContactsRequest(id=[user]))
    except Exception as exc:
        print(f"  [WARN] Could not delete temporary contact: {type(exc).__name__}: {exc}")
    return user


def write_results_csv(ws, path: Path, cols: dict[str, int], last_row: int) -> None:
    rows = []
    for row in range(2, last_row + 1):
        phone = normalize_phone(ws.cell(row=row, column=cols["phone"]).value)
        user_id = normalize_cell(ws.cell(row=row, column=cols["user_id"]).value)
        access_hash = normalize_cell(ws.cell(row=row, column=cols["access_hash"]).value)
        username = normalize_cell(ws.cell(row=row, column=cols["username"]).value)
        if not phone or not user_id:
            continue
        rows.append(
            {
                "excel_row": row,
                "fio": normalize_cell(ws.cell(row=row, column=cols["fio"]).value) if cols.get("fio") else "",
                "link": normalize_cell(ws.cell(row=row, column=cols["link"]).value) if cols.get("link") else "",
                "phone": phone,
                "user_id": user_id,
                "access_hash": access_hash,
                "username": username,
            }
        )

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["excel_row", "fio", "link", "phone", "user_id", "access_hash", "username"],
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"[CSV] Saved {len(rows)} Telegram users to {path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Find Telegram user_id/access_hash by phone numbers from Excel.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Excel file path.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to active sheet.")
    parser.add_argument("--account", default="search", help="Telegram account name. Example: search, sender, main.")
    parser.add_argument("--start-row", type=int, default=2, help="First Excel row.")
    parser.add_argument("--limit", type=int, default=None, help="Max phone rows to try.")
    parser.add_argument("--delay", type=float, default=float(os.getenv("TG_PHONE_SEARCH_DELAY", "60")), help="Delay between phone searches.")
    parser.add_argument("--phone-header", default="Телефон")
    parser.add_argument("--fio-header", default="ФИО")
    parser.add_argument("--link-header", default="Ссылка")
    parser.add_argument("--user-id-header", default="Telegram user_id")
    parser.add_argument("--access-hash-header", default="Telegram access_hash")
    parser.add_argument("--username-header", default="Telegram username")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output CSV path.")
    parser.add_argument("--force", action="store_true", help="Search again even if Telegram columns are filled.")
    parser.add_argument("--dry-run", action="store_true", help="Do not write Excel/CSV.")
    return parser


async def main() -> None:
    args = build_parser().parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path.resolve()}")

    account = resolve_account(args.account)
    client = TelegramClient(str(account.session_path.with_suffix("")), account.api_id, account.api_hash)
    await client.connect()

    wb = load_workbook(xlsx_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    phone_col = find_col(ws, args.phone_header)
    if not phone_col:
        raise SystemExit(f"Phone column not found: {args.phone_header!r}")

    cols = {
        "phone": phone_col,
        "fio": find_col(ws, args.fio_header),
        "link": find_col(ws, args.link_header),
        "user_id": get_or_create_col(ws, args.user_id_header),
        "access_hash": get_or_create_col(ws, args.access_hash_header),
        "username": get_or_create_col(ws, args.username_header),
    }
    last_row = detect_last_row(ws, [phone_col, cols["user_id"], cols["access_hash"]])

    print(f"Excel: {xlsx_path.resolve()}")
    print(f"Sheet: {ws.title}")
    print(f"Account: {account.account}")
    print(f"Session: {account.session_path}")
    print(f"Loaded .env files: {', '.join(str(path) for path in LOADED_ENV_FILES) or '-'}")
    print(f"Last data row: {last_row}")

    attempted = 0
    found = 0
    skipped = 0
    last_search_at = 0.0

    try:
        await ensure_authorized(client, account)

        for row in range(args.start_row, last_row + 1):
            phone = normalize_phone(ws.cell(row=row, column=phone_col).value)
            if not phone:
                continue

            existing_user_id = normalize_cell(ws.cell(row=row, column=cols["user_id"]).value)
            existing_access_hash = normalize_cell(ws.cell(row=row, column=cols["access_hash"]).value)
            if existing_user_id and existing_access_hash and not args.force:
                skipped += 1
                continue

            if args.limit is not None and attempted >= args.limit:
                print(f"Limit reached: {args.limit}")
                break

            attempted += 1
            now = time.monotonic()
            if last_search_at:
                wait = args.delay - (now - last_search_at)
                if wait > 0:
                    print(f"[DELAY] Sleeping {wait:.1f}s before next Telegram phone search...")
                    await asyncio.sleep(wait)

            print(f"\n[{row}] phone={phone}")
            try:
                user = await find_user_by_phone(client, phone)
                last_search_at = time.monotonic()
            except errors.FloodWaitError as exc:
                print(f"[TG][STOP] FloodWaitError. Telegram asks to wait {exc.seconds} seconds.")
                break
            except errors.PeerFloodError as exc:
                print(f"[TG][STOP] PeerFloodError. Account is limited for too many actions: {exc}")
                break
            except errors.PhoneNumberBannedError as exc:
                print(f"[TG][STOP] PhoneNumberBannedError. Your Telegram phone/account may be banned: {exc}")
                break
            except errors.UserDeactivatedBanError as exc:
                print(f"[TG][STOP] UserDeactivatedBanError. Account is deactivated/banned: {exc}")
                break
            except errors.RPCError as exc:
                print(f"[TG][ERROR] {type(exc).__name__}: {exc}")
                skipped += 1
                continue

            if not user:
                print("  not found")
                skipped += 1
                continue

            user_id = getattr(user, "id", None)
            access_hash = getattr(user, "access_hash", None)
            username = getattr(user, "username", None) or ""
            print(f"  user_id={user_id}, access_hash={access_hash}, username={username or '-'}")

            if not user_id or not access_hash:
                skipped += 1
                continue

            if not args.dry_run:
                ws.cell(row=row, column=cols["user_id"]).value = str(user_id)
                ws.cell(row=row, column=cols["access_hash"]).value = str(access_hash)
                ws.cell(row=row, column=cols["username"]).value = username
                wb.save(xlsx_path)
            found += 1
    finally:
        if not args.dry_run:
            wb.save(xlsx_path)
            write_results_csv(ws, Path(args.output), cols, last_row)
        wb.close()
        await client.disconnect()

    print("\nDone.")
    print(f"Attempted: {attempted}")
    print(f"Found: {found}")
    print(f"Skipped: {skipped}")


if __name__ == "__main__":
    asyncio.run(main())
