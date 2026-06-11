import argparse
import re
import sys
from pathlib import Path
from urllib.parse import urlparse

try:
    from bs4 import BeautifulSoup
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit(
        "Missing dependency. Install requirements first:\n"
        "  python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR.parent / "test.txt"
DEFAULT_XLSX = BASE_DIR / "rusprofile_users_from_test.xlsx"
DEFAULT_SHEET = "Уникальные записи"

HEADERS = ("Порядковый номер", "ФИО", "Ссылка")
ID_LINK_RE = re.compile(r"/id/(\d+)")
PERSON_LINE_RE = re.compile(r"^(.+?)\s+-\s+.+$")


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_id_link(raw_link: str) -> str:
    link = normalize_text(raw_link)
    if not link:
        return ""

    parsed_path = urlparse(link).path if link.startswith(("http://", "https://")) else link
    match = ID_LINK_RE.search(parsed_path)
    return f"/id/{match.group(1)}" if match else ""


def extract_person_name(card) -> str:
    for text_tag in card.select(".list-element__text"):
        text = normalize_text(text_tag.get_text(" ", strip=True))
        match = PERSON_LINE_RE.match(text)
        if not match:
            continue

        fio = normalize_text(match.group(1))
        if re.search(r"[А-Яа-яЁё]", fio) and not re.search(r"\d", fio):
            return fio

    return ""


def parse_rusprofile_search_html(html: str) -> list[tuple[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[tuple[str, str]] = []
    seen_links: set[str] = set()

    for card in soup.select(".list-element"):
        link_tag = card.select_one('a[href*="/id/"]')
        if not link_tag:
            continue

        link = normalize_id_link(link_tag.get("href", ""))
        if not link or link in seen_links:
            continue

        fio = extract_person_name(card)
        if not fio:
            continue

        rows.append((fio, link))
        seen_links.add(link)

    return rows


def write_excel(rows: list[tuple[str, str]], xlsx_path: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(fill_type="solid", fgColor="0F5E7E")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for index, (fio, link) in enumerate(rows, start=1):
        row = index + 1
        ws.cell(row=row, column=1, value=index)
        ws.cell(row=row, column=2, value=fio)
        ws.cell(row=row, column=3, value=link)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 24

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        row[0].alignment = Alignment(horizontal="right")
        row[1].alignment = Alignment(horizontal="left")
        row[2].alignment = Alignment(horizontal="left")

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    wb.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert saved Rusprofile search HTML from test.txt to an Excel lead list."
    )
    parser.add_argument("--txt", default=str(DEFAULT_INPUT), help="Input HTML/text file path.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Output Excel file path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Output sheet name.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    txt_path = Path(args.txt)
    xlsx_path = Path(args.xlsx)

    if not txt_path.exists():
        raise SystemExit(f"Input file not found: {txt_path.resolve()}")

    html = txt_path.read_text(encoding="utf-8-sig")
    rows = parse_rusprofile_search_html(html)
    if not rows:
        raise SystemExit("No rows found. Expected Rusprofile cards with /id/... links and person names.")

    write_excel(rows, xlsx_path, args.sheet)

    print(f"Input: {txt_path.resolve()}")
    print(f"Output: {xlsx_path.resolve()}")
    print(f"Rows written: {len(rows)}")


if __name__ == "__main__":
    sys.exit(main())
