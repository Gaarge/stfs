import argparse
import html as html_module
import os
import random
import re
import shutil
import subprocess
import sys
import time
from pathlib import Path
from urllib.parse import unquote, urljoin

try:
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency. Install requirements first:\n"
        "  python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
BASE_URL = "https://www.rusprofile.ru"

DEFAULT_XLSX = BASE_DIR / "rusprofile_users.xlsx"
DEFAULT_QUEUE = BASE_DIR / "rusprofile_email_queue.txt"
DEFAULT_EMAIL_HEADER = "Email"
DEFAULT_COOKIE = os.getenv("RUSPROFILE_COOKIE", "")
DEFAULT_COOKIE_FILE = os.getenv("RUSPROFILE_COOKIE_FILE", "")
LINK_HEADERS = ("Ссылка", "ссылка", "link", "url")
FIO_HEADERS = ("ФИО", "фио", "name")
HTTP_STATUS_MARKER = "\n__RUSPROFILE_HTTP_STATUS__:"
CURL_RETRY_EXIT_CODES = {35, 56, 92}

EMAIL_RE = re.compile(
    r"(?<![\w.+-])([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})(?![\w.-])",
    flags=re.IGNORECASE,
)

USER_AGENTS = [
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/132.0.0.0 YaBrowser/25.2.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (iPad; CPU OS 17_5 like Mac OS X) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 14; Pixel 8 Pro) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 14; SM-S928B) AppleWebKit/537.36 "
    "(KHTML, like Gecko) SamsungBrowser/25.0 Chrome/121.0.0.0 Mobile Safari/537.36",
]


class CaptchaDetected(RuntimeError):
    pass


class CurlFetchError(RuntimeError):
    def __init__(self, attempts: list[str]) -> None:
        super().__init__("\n".join(attempts))
        self.attempts = attempts


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value) -> str:
    return normalize_cell(value).lower()


def find_col_by_headers(ws, headers: tuple[str, ...]) -> int | None:
    wanted = {normalize_header(header) for header in headers}
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=1, column=col).value) in wanted:
            return col
    return None


def get_or_create_col(ws, header: str) -> int:
    existing = find_col_by_headers(ws, (header,))
    if existing:
        return existing
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header
    return col


def detect_last_row(ws, columns: list[int]) -> int:
    for row in range(ws.max_row, 1, -1):
        if any(normalize_cell(ws.cell(row=row, column=col).value) for col in columns):
            return row
    return 1


def make_rusprofile_id_url(raw_link: str) -> str:
    link = normalize_cell(raw_link)
    if not link:
        return ""
    if link.startswith("//"):
        link = "https:" + link

    match = re.search(r"/(?:id|ip)/(\d+)", link)
    if match:
        return f"{BASE_URL}/id/{match.group(1)}"

    if link.isdigit():
        return f"{BASE_URL}/id/{link}"

    if link.startswith("http://") or link.startswith("https://"):
        return link

    return urljoin(BASE_URL + "/", link.lstrip("/"))


def is_captcha_page(page_html: str) -> bool:
    markers = (
        "captcha-section",
        "g-recaptcha",
        "captcha-validate",
        "Активность с вашего IP-адреса была распознана как автоматическая",
        "Я не робот",
    )
    return any(marker in page_html for marker in markers)


def is_subscription_locked_page(page_html: str) -> bool:
    text = page_html.lower()
    markers = (
        "оформите подписку",
        "оформить подписку",
        "нужна подписка",
        "доступно по подписке",
        "email доступен",
        "электронная почта доступна",
    )
    return any(marker in text for marker in markers)


def normalize_raw_cookie(raw_cookie: str) -> str:
    cookie = normalize_cell(raw_cookie)
    if cookie.lower().startswith("cookie:"):
        cookie = cookie.split(":", 1)[1].strip()
    return " ".join(cookie.splitlines()).strip()


def resolve_cookie_args(raw_cookie: str, raw_cookie_file: str) -> tuple[str, Path | None, str]:
    cookie = normalize_raw_cookie(raw_cookie)
    if not raw_cookie_file:
        return cookie, None, "raw" if cookie else "none"

    cookie_file = Path(raw_cookie_file).expanduser()
    if not cookie_file.exists():
        raise SystemExit(f"Cookie file not found: {cookie_file.resolve()}")

    content = cookie_file.read_text(encoding="utf-8-sig").strip()
    first_line = next(
        (line.strip() for line in content.splitlines() if line.strip() and not line.startswith("#")),
        "",
    )

    if first_line.lower().startswith("cookie:") or ("=" in first_line and "\t" not in first_line):
        file_cookie = normalize_raw_cookie(content)
        return file_cookie, None, "raw-file"

    return cookie, cookie_file, "curl-file"


def fetch_html_with_curl(
    url: str,
    user_agent: str,
    timeout: int,
    cookie: str = "",
    cookie_file: Path | None = None,
    cookie_jar: Path | None = None,
    referer: str = "",
) -> tuple[str, int | None]:
    base_cmd = [
        "curl",
        "--location",
        "--silent",
        "--show-error",
        "--compressed",
        "--max-time",
        str(timeout),
        "--connect-timeout",
        str(min(timeout, 15)),
        "-A",
        user_agent,
        "-H",
        "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "-H",
        "Accept-Language: ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "-H",
        "Cache-Control: max-age=0",
        "-H",
        "Upgrade-Insecure-Requests: 1",
        "-H",
        "Sec-Fetch-Dest: document",
        "-H",
        "Sec-Fetch-Mode: navigate",
        "-H",
        "Sec-Fetch-Site: same-origin",
        "-H",
        "Sec-Fetch-User: ?1",
        "--write-out",
        f"{HTTP_STATUS_MARKER}%{{http_code}}",
    ]
    if referer:
        base_cmd.extend(["--referer", referer])
    if cookie:
        base_cmd.extend(["--cookie", cookie])
    if cookie_file:
        base_cmd.extend(["--cookie", str(cookie_file)])
    if cookie_jar:
        base_cmd.extend(["--cookie-jar", str(cookie_jar)])

    attempts = [
        ("default", []),
        ("http1.1", ["--http1.1"]),
        ("http1.1/tls1.2", ["--http1.1", "--tlsv1.2"]),
        ("ipv4/http1.1/tls1.2", ["--ipv4", "--http1.1", "--tlsv1.2"]),
    ]
    errors: list[str] = []

    completed = None
    for label, extra_args in attempts:
        cmd = [*base_cmd, *extra_args, url]
        try:
            completed = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=timeout + 5,
                check=False,
            )
        except subprocess.TimeoutExpired:
            errors.append(f"{label}: timed out after {timeout} seconds")
            continue

        if completed.returncode == 0:
            break

        stderr = completed.stderr.strip()
        details = f": {stderr}" if stderr else ""
        errors.append(f"{label}: curl exit {completed.returncode}{details}")
        if completed.returncode not in CURL_RETRY_EXIT_CODES:
            break
    else:
        completed = None

    if completed is None or completed.returncode != 0:
        raise CurlFetchError(errors)

    page_html = completed.stdout
    status = None
    if HTTP_STATUS_MARKER in page_html:
        page_html, status_raw = page_html.rsplit(HTTP_STATUS_MARKER, 1)
        status_raw = status_raw.strip()
        if status_raw:
            try:
                status = int(status_raw.splitlines()[-1])
            except ValueError:
                status = None

    return page_html, status


def clean_email(raw_email: str) -> str:
    email = html_module.unescape(normalize_cell(raw_email)).replace("\xa0", " ")
    if email.lower().startswith("mailto:"):
        email = email[7:]
    email = unquote(email.split("?", 1)[0]).strip(" \t\r\n\"'<>")
    match = EMAIL_RE.search(email)
    return match.group(1).lower() if match else ""


def add_email(raw_email: str, emails: list[str], seen: set[str]) -> None:
    email = clean_email(raw_email)
    if email and email not in seen:
        emails.append(email)
        seen.add(email)


def extract_emails(page_html: str) -> list[str]:
    if is_captcha_page(page_html):
        raise CaptchaDetected("Rusprofile returned captcha page")

    soup = BeautifulSoup(page_html, "html.parser")
    emails: list[str] = []
    seen: set[str] = set()

    for link in soup.select('.company-info__contact.mail a[href^="mailto:"]'):
        add_email(link.get("href", ""), emails, seen)
        add_email(link.get_text(" ", strip=True), emails, seen)

    for link in soup.select('a[href^="mailto:"]'):
        add_email(link.get("href", ""), emails, seen)
        add_email(link.get_text(" ", strip=True), emails, seen)

    if not emails:
        for contact in soup.select(".company-info__contact"):
            classes = contact.get("class") or []
            text = contact.get_text(" ", strip=True)
            if "mail" not in classes and "Электронная почта" not in text:
                continue
            for match in EMAIL_RE.findall(text):
                add_email(match, emails, seen)

    if not emails:
        for match in re.finditer(r"mailto:([^\"'<>\s]+)", page_html, flags=re.IGNORECASE):
            add_email(f"mailto:{match.group(1)}", emails, seen)

    if not emails:
        clean_html = html_module.unescape(page_html)
        for match in EMAIL_RE.findall(clean_html):
            add_email(match, emails, seen)

    return emails


def split_emails(raw_value: str) -> list[str]:
    emails: list[str] = []
    seen: set[str] = set()
    for match in EMAIL_RE.findall(html_module.unescape(normalize_cell(raw_value))):
        add_email(match, emails, seen)
    return emails


def load_queue_emails(path: Path) -> set[str]:
    if not path.exists():
        return set()
    emails: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        emails.update(split_emails(line))
    return emails


def queue_emails(path: Path, emails: list[str], queued_emails: set[str], dry_run: bool) -> int:
    new_emails = [email for email in emails if email not in queued_emails]
    if not new_emails:
        return 0

    if not dry_run:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as f:
            for email in new_emails:
                f.write(email + "\n")

    queued_emails.update(new_emails)
    return len(new_emails)


def merge_emails(first: list[str], second: list[str]) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for email in first + second:
        add_email(email, merged, seen)
    return merged


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fill Email column and queue emails from Rusprofile /id/ links."
    )
    parser.add_argument("--url", default="", help="Fetch one Rusprofile /id/... URL and print emails without Excel.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Excel file path.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to active sheet.")
    parser.add_argument("--queue", default=str(DEFAULT_QUEUE), help="Email queue path. One email per line.")
    parser.add_argument("--start-row", type=int, default=2, help="First data row.")
    parser.add_argument("--limit", type=int, default=None, help="Max rows to fetch.")
    parser.add_argument("--force", action="store_true", help="Parse even when Email already exists.")
    parser.add_argument("--dry-run", action="store_true", help="Do not write to Excel or queue.")
    parser.add_argument("--link-header", default="Ссылка", help="Link column header.")
    parser.add_argument("--email-header", default=DEFAULT_EMAIL_HEADER, help="Email output column header.")
    parser.add_argument("--fio-header", default="ФИО", help="FIO column header for logs.")
    parser.add_argument("--delay-min", type=float, default=4.0, help="Min delay between requests.")
    parser.add_argument("--delay-max", type=float, default=12.0, help="Max delay between requests.")
    parser.add_argument("--timeout", type=int, default=40, help="curl timeout seconds.")
    parser.add_argument(
        "--user-agent",
        default="",
        help="Fixed User-Agent. If empty, one of USER_AGENTS is chosen randomly for each request.",
    )
    parser.add_argument(
        "--referer",
        default=BASE_URL + "/",
        help="Referer header for Rusprofile requests. Pass empty string to disable.",
    )
    parser.add_argument(
        "--cookie",
        default=DEFAULT_COOKIE,
        help="Raw Rusprofile Cookie header. Can also be set through RUSPROFILE_COOKIE.",
    )
    parser.add_argument(
        "--cookie-file",
        default=DEFAULT_COOKIE_FILE,
        help="Path to a curl/Netscape cookie file. Can also be set through RUSPROFILE_COOKIE_FILE.",
    )
    parser.add_argument(
        "--cookie-jar",
        default="",
        help="Optional path where curl should save updated cookies after requests.",
    )
    parser.add_argument("--captcha-sleep", type=int, default=1800, help="Sleep seconds after captcha.")
    parser.add_argument("--captcha-retries", type=int, default=1, help="Retries for the same row after captcha.")
    return parser


def fetch_emails_for_url(args, url: str, cookie: str, cookie_file: Path | None, cookie_jar: Path | None) -> list[str]:
    user_agent = args.user_agent or random.choice(USER_AGENTS)
    print(f"URL: {url}")
    print(f"User-Agent: {user_agent}")
    try:
        page_html, status = fetch_html_with_curl(
            url,
            user_agent,
            timeout=args.timeout,
            cookie=cookie,
            cookie_file=cookie_file,
            cookie_jar=cookie_jar,
            referer=args.referer,
        )
    except CurlFetchError as exc:
        print("Curl failed after retries:")
        for attempt in exc.attempts:
            print(f"  {attempt}")
        raise SystemExit("Stop: cannot fetch URL.") from exc

    if status and status >= 400:
        print(f"HTTP status: {status}")

    emails = extract_emails(page_html)
    if emails:
        print("Emails:")
        for email in emails:
            print(email)
    else:
        print("Emails: -")
        if is_subscription_locked_page(page_html):
            print("Subscription: page still looks locked. Check that Rusprofile cookies are fresh.")
    return emails


def main() -> None:
    args = build_parser().parse_args()
    if args.delay_min < 0 or args.delay_max < 0 or args.delay_min > args.delay_max:
        raise SystemExit("Delay values must be non-negative and delay-min must be <= delay-max.")
    if not shutil.which("curl"):
        raise SystemExit("curl is not installed or not available in PATH.")

    xlsx_path = Path(args.xlsx)
    queue_path = Path(args.queue)
    cookie, cookie_file, cookie_mode = resolve_cookie_args(args.cookie, args.cookie_file)
    cookie_jar = Path(args.cookie_jar).expanduser() if args.cookie_jar else None
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path.resolve()}")
    if cookie_jar:
        cookie_jar.parent.mkdir(parents=True, exist_ok=True)

    if args.url:
        url = make_rusprofile_id_url(args.url)
        if not url:
            raise SystemExit("URL is empty or invalid.")
        print(f"Cookie auth: {'yes' if cookie or cookie_file else 'no'} ({cookie_mode})")
        fetch_emails_for_url(args, url, cookie, cookie_file, cookie_jar)
        return

    wb = load_workbook(xlsx_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    link_headers = tuple(dict.fromkeys((args.link_header, *LINK_HEADERS)))
    fio_headers = tuple(dict.fromkeys((args.fio_header, *FIO_HEADERS)))
    link_col = find_col_by_headers(ws, link_headers)
    if not link_col:
        raise SystemExit(f"Could not find link column. Expected header: {args.link_header!r}")

    email_col = get_or_create_col(ws, args.email_header)
    fio_col = find_col_by_headers(ws, fio_headers)
    last_row = detect_last_row(ws, [link_col, email_col])
    queued_emails = load_queue_emails(queue_path)

    print(f"Excel: {xlsx_path.resolve()}")
    print(f"Sheet: {ws.title}")
    print(f"Link column: #{link_col}")
    print(f"Email column: #{email_col}")
    print(f"Queue: {queue_path.resolve()}")
    print(f"User agents: {len(USER_AGENTS)}")
    print(f"User agent mode: {'fixed' if args.user_agent else 'random'}")
    print(f"Cookie auth: {'yes' if cookie or cookie_file else 'no'} ({cookie_mode})")
    print(f"Last data row: {last_row}")
    print(f"Dry-run: {'yes' if args.dry_run else 'no'}")

    fetched = 0
    rows_with_email = 0
    excel_updated = 0
    queued = 0
    skipped_existing = 0

    try:
        for row in range(args.start_row, last_row + 1):
            if args.limit is not None and fetched >= args.limit:
                print(f"Limit reached: {args.limit}")
                break

            raw_link = ws.cell(row=row, column=link_col).value
            url = make_rusprofile_id_url(raw_link)
            if not url:
                continue

            existing_emails = split_emails(ws.cell(row=row, column=email_col).value)
            if existing_emails and not args.force:
                queued += queue_emails(queue_path, existing_emails, queued_emails, args.dry_run)
                skipped_existing += 1
                continue

            fetched += 1
            fio = normalize_cell(ws.cell(row=row, column=fio_col).value) if fio_col else ""
            print(f"\n[{row}] {fio or '-'}")
            print(f"  url: {url}")

            captcha_attempts = 0
            while True:
                user_agent = args.user_agent or random.choice(USER_AGENTS)
                print(f"  user-agent: {user_agent}")
                try:
                    page_html, status = fetch_html_with_curl(
                        url,
                        user_agent,
                        timeout=args.timeout,
                        cookie=cookie,
                        cookie_file=cookie_file,
                        cookie_jar=cookie_jar,
                        referer=args.referer,
                    )
                    if status and status >= 400:
                        print(f"  HTTP status: {status}")
                    emails = extract_emails(page_html)
                    break
                except CaptchaDetected as exc:
                    captcha_attempts += 1
                    print(f"  CAPTCHA: {exc}")
                    if captcha_attempts > args.captcha_retries:
                        print("  Stop on repeated captcha. Run later with --start-row", row)
                        raise
                    print(f"  Sleeping {args.captcha_sleep} seconds before retry...")
                    time.sleep(args.captcha_sleep)
                except CurlFetchError as exc:
                    print("  curl failed after retries:")
                    for attempt in exc.attempts:
                        print(f"    {attempt}")
                    raise SystemExit(f"Stop on curl error. Run later with --start-row {row}") from exc

            if not emails:
                print("  email: -")
                if is_subscription_locked_page(page_html):
                    print("  subscription: page still looks locked. Check that Rusprofile cookies are fresh.")
            else:
                rows_with_email += 1
                final_emails = merge_emails(existing_emails, emails)
                final_value = ", ".join(final_emails)
                print(f"  email: {final_value}")

                if normalize_cell(ws.cell(row=row, column=email_col).value) != final_value:
                    excel_updated += 1
                    if not args.dry_run:
                        ws.cell(row=row, column=email_col).value = final_value
                        wb.save(xlsx_path)

                queued += queue_emails(queue_path, final_emails, queued_emails, args.dry_run)

            delay = random.uniform(args.delay_min, args.delay_max)
            print(f"  delay: {delay:.1f}s")
            time.sleep(delay)
    except KeyboardInterrupt:
        print("\nInterrupted by user.")
    finally:
        if not args.dry_run:
            wb.save(xlsx_path)
        wb.close()

    print("\nDone.")
    print(f"Rows fetched: {fetched}")
    print(f"Rows with email: {rows_with_email}")
    print(f"Rows skipped with existing email: {skipped_existing}")
    print(f"Excel rows updated: {excel_updated}")
    print(f"Emails queued: {queued}")


if __name__ == "__main__":
    sys.exit(main())
